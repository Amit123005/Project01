from flask import current_app, jsonify
import pymysql.cursors
from datetime import date, datetime, timedelta
from icecream import ic
import time
import os
import re
from openpyxl import load_workbook
from config import Config
from flask_mail import Mail, Message
from Email_Extension import mail
from flask import current_app
import smtplib
from email.message import EmailMessage
from config import Config
import json
import requests
from werkzeug.security import generate_password_hash, check_password_hash
from pymodbus.client import ModbusTcpClient

def get_db_connection(db_name_key='DB_NAME_Line8'):
    return pymysql.connect(
        host=Config.DB_HOST,
        user=Config.DB_USER,
        password=Config.DB_PASSWORD,
        db=getattr(Config, db_name_key),
        cursorclass=pymysql.cursors.DictCursor
    )



class Production:

    def test():
        ic("Test Function")

    def sanitize_filename(name):
        return re.sub(r'[<>:"/\\|?*]', '_', name)
    
    def password_hash2(password):
        plain_password = password
        hashed_password = generate_password_hash(plain_password, method='scrypt')
        return hashed_password

    def get_pqcr_report():
        generated_files = []
        gl_name = "Saurabh Vishwakarma"
        line = "8"
        date = datetime.now()
        mysql_date = date.strftime('%Y-%m-%d')
        time_only = date.time()
        shift_a_start = datetime.strptime("06:00", "%H:%M").time()
        shift_b_start = datetime.strptime("14:00", "%H:%M").time()
        shift_c_start = datetime.strptime("22:00", "%H:%M").time()

        if shift_a_start <= time_only < shift_b_start:
            shift = "A"
            shift_start = date.replace(hour=6, minute=0, second=0)
            shift_end = date.replace(hour=14, minute=0, second=0)

        elif shift_b_start <= time_only < shift_c_start:
            shift = "B"
            shift_start = date.replace(hour=14, minute=0, second=0)
            shift_end = date.replace(hour=22, minute=0, second=0)

        else:
            shift = "C"
            if time_only >= shift_c_start:
                shift_start = date.replace(hour=22, minute=0, second=0)
                shift_end = (shift_start + timedelta(hours=8))
            else:
                # Early morning hours (00:00 to 06:00)
                shift_end = date.replace(hour=6, minute=0, second=0)
                shift_start = shift_end - timedelta(hours=8)

        connection = get_db_connection()
        cursor = connection.cursor()

        unique_ids_query = """
            SELECT DISTINCT id FROM (
                SELECT id FROM l8ext1 WHERE timestamp BETWEEN %s AND %s AND shift = %s
                UNION
                SELECT id FROM l8ext2 WHERE timestamp BETWEEN %s AND %s AND shift = %s
                UNION
                SELECT id FROM l8ext3 WHERE timestamp BETWEEN %s AND %s AND shift = %s
                UNION
                SELECT id FROM l8ext4 WHERE timestamp BETWEEN %s AND %s AND shift = %s
            ) AS combined_ids;
        """

        cursor.execute(unique_ids_query, (
            shift_start, shift_end, shift,
            shift_start, shift_end, shift,
            shift_start, shift_end, shift,
            shift_start, shift_end, shift,
        ))
        all_ids = [row['id'] for row in cursor.fetchall()]
        # ic(all_ids)
        for current_id in all_ids:
            cursor.execute( "Select partna, partno, model from planning where id = %s", (current_id,))
            plannig_data = cursor.fetchone()
            part_name  = plannig_data['partna']
            part_no  = plannig_data['partno']
            model  = plannig_data['model']
            model = Production.sanitize_filename(model)
            part_name = Production.sanitize_filename(part_name)

            cursor.execute("Select min(hour) as hour from l8ext1 WHERE timestamp BETWEEN %s AND %s AND shift = %s and id = %s", (shift_start, shift_end, shift, current_id))
            hour_result = cursor.fetchone()
            hour = hour_result['hour']
            ic(current_id, plannig_data, hour)

            ext1_query = """SELECT 
                            HOUR(timestamp) AS hour_block,
                            MAX(BZ1_Set) AS BZ1_Set, MAX(BZ1_Actual) AS BZ1_Actual, MAX(BZ2_Set) AS BZ2_Set, MAX(BZ2_Actual) AS BZ2_Actual, MAX(BZ3_Set) AS BZ3_Set,
                            MAX(BZ3_Actual) AS BZ3_Actual, MAX(BZ4_Set) AS BZ4_Set, MAX(BZ4_Actual) AS BZ4_Actual, MAX(DZ1_Set) AS DZ1_Set,
                            MAX(DZ1_Actual) AS DZ1_Actual, MAX(DZ2_Set) AS DZ2_Set, MAX(DZ2_Actual) AS DZ2_Actual, MAX(DZ3_Set) AS DZ3_Set,
                            MAX(DZ3_Actual) AS DZ3_Actual, MAX(Main_Motor_Set) AS Main_Motor_SetE1, MAX(Main_Motor_Actual) AS Main_Motor_ActualE1
                        FROM l8ext1
                        WHERE timestamp BETWEEN %s AND %s and shift = %s AND id = %s
                        GROUP BY HOUR(timestamp)
                        ORDER BY hour_block;
                        """
            ic
            cursor.execute(ext1_query, (shift_start, shift_end, shift, current_id))
            result1= cursor.fetchall()

            ext2_query = """SELECT 
                            HOUR(timestamp) AS hour_block,
                            MAX(BZ1_Set) AS BZ1_Set, MAX(BZ1_Actual) AS BZ1_Actual, MAX(BZ2_Set) AS BZ2_Set, MAX(BZ2_Actual) AS BZ2_Actual, MAX(BZ3_Set) AS BZ3_Set,
                            MAX(BZ3_Actual) AS BZ3_Actual, MAX(BZ4_Set) AS BZ4_Set, MAX(BZ4_Actual) AS BZ4_Actual, MAX(DZ1_Set) AS DZ1_Set,
                            MAX(DZ1_Actual) AS DZ1_Actual, MAX(DZ2_Set) AS DZ2_Set, MAX(DZ2_Actual) AS DZ2_Actual, MAX(DZ3_Set) AS DZ3_Set,
                            MAX(DZ3_Actual) AS DZ3_Actual, MAX(Main_Motor_Set) AS Main_Motor_SetE2, MAX(Main_Motor_Actual) AS Main_Motor_ActualE2
                        FROM l8ext2
                        WHERE timestamp BETWEEN %s AND %s and shift = %s AND id = %s
                        GROUP BY HOUR(timestamp)
                        ORDER BY hour_block;
                        """
            cursor.execute(ext2_query, (shift_start, shift_end, shift, current_id))
            result2= cursor.fetchall()

            ext3_query = """SELECT 
                            HOUR(timestamp) AS hour_block,
                            MAX(BZ1_Set) AS BZ1_Set, MAX(BZ1_Actual) AS BZ1_Actual, MAX(BZ2_Set) AS BZ2_Set, MAX(BZ2_Actual) AS BZ2_Actual, MAX(BZ3_Set) AS BZ3_Set,
                            MAX(BZ3_Actual) AS BZ3_Actual, MAX(BZ4_Set) AS BZ4_Set, MAX(BZ4_Actual) AS BZ4_Actual, MAX(DZ1_Set) AS DZ1_Set,
                            MAX(DZ1_Actual) AS DZ1_Actual, MAX(DZ2_Set) AS DZ2_Set, MAX(DZ2_Actual) AS DZ2_Actual,
                            MAX(Main_Motor_Set) AS Main_Motor_SetE3, MAX(Main_Motor_Actual) AS Main_Motor_ActualE3
                        FROM l8ext3
                        WHERE timestamp BETWEEN %s AND %s and shift = %s AND id = %s
                        GROUP BY HOUR(timestamp)
                        ORDER BY hour_block;
                        """
            cursor.execute(ext3_query, (shift_start, shift_end, shift, current_id))
            result3= cursor.fetchall()

            ext4_query ="""SELECT 
                            HOUR(timestamp) AS hour_block,
                            MAX(BZ1_Set) AS BZ1_Set, MAX(BZ1_Actual) AS BZ1_Actual, MAX(BZ2_Set) AS BZ2_Set, MAX(BZ2_Actual) AS BZ2_Actual, MAX(BZ3_Set) AS BZ3_Set,
                            MAX(BZ3_Actual) AS BZ3_Actual, MAX(BZ4_Set) AS BZ4_Set, MAX(BZ4_Actual) AS BZ4_Actual, MAX(DZ1_Set) AS DZ1_Set,
                            MAX(DZ1_Actual) AS DZ1_Actual, MAX(DZ2_Set) AS DZ2_Set, MAX(DZ2_Actual) AS DZ2_Actual, MAX(DZ3_Set) AS DZ3_Set,
                            MAX(DZ3_Actual) AS DZ3_Actual, MAX(Main_Motor_Set) AS Main_Motor_SetE4, MAX(Main_Motor_Actual) AS Main_Motor_ActualE4
                        FROM l8ext4
                        WHERE timestamp BETWEEN %s AND %s and shift = %s AND id = %s
                        GROUP BY HOUR(timestamp)
                        ORDER BY hour_block;
                        """
            cursor.execute(ext4_query, (shift_start, shift_end, shift, current_id))
            result4= cursor.fetchall()

            common_query = """SELECT 
                            HOUR(datetime) AS hour_block,
                            MAX(Main_Die_Set) AS Main_Die_Set, MAX(Main_Die_Actual) AS Main_Die_Actual, 
                            MAX(Main_Puller_Set) AS Main_Puller_Set, MAX(Main_Puller_Actual) AS Main_Puller_Actual,
                            MAX(Mini_Puller_Set) AS Mini_Puller_Set, MAX(Mini_Puller_Actual) AS Mini_Puller_Actual,
                            MAX(Cutting_Machine_Set) AS Cutting_Machine_Set, MAX(Cutting_Machine_Actual) AS Cutting_Machine_Actual,
                            MAX(Cooling_Trough_Actual) AS Cooling_Trough_Actual, 
                            MAX(factor) AS factor, MAX(comp_press) AS comp_press, MAX(root_press) AS root_press
                        FROM common
                        WHERE datetime BETWEEN %s AND %s and shift = %s AND id = %s
                        GROUP BY HOUR(datetime)
                        ORDER BY hour_block;
                        """
            cursor.execute(common_query, (shift_start, shift_end, shift, current_id))
            result5= cursor.fetchall()

            last_query = """SELECT HOUR(timestamp) AS hour_block, MAX(line_speed) as line_speed 
                            FROM L8_Running_Status 
                            WHERE timestamp BETWEEN %s AND %s AND shift = %s AND plan_id = %s
                            GROUP BY HOUR(timestamp)
                            ORDER BY hour_block;
                            """
            cursor.execute(last_query, (shift_start, shift_end, shift, current_id))
            result6= cursor.fetchall()

            file_path = r"D:\PQCR\PQCR2_Format.xlsx"
            # new_file_path = f"D:\\PQCR\\PQCR_{mysql_date}_{shift}.xlsx"
            # download_path = os.path.join(os.path.expanduser("~"), "Downloads", f"PQCR{mysql_date}_{shift}.xlsx")

            wb = load_workbook(file_path)
            ws = wb.active

            # Update header information
            ws['B2'] = f"{mysql_date}"
            ws['B4'] = f"{shift}"
            ws['B6'] = f"{line}"
            ws['K2'] = f"{part_name}"
            ws['K4'] = f"{part_no}"
            ws['K6'] = f"{model}"

            for idx, result in enumerate(result1):
                row = 9 + hour + idx
                ws[f'B{row}'] = result['BZ1_Actual']
                ws[f'C{row}'] = result['BZ1_Set']
                ws[f'D{row}'] = result['BZ2_Actual']
                ws[f'E{row}'] = result['BZ2_Set']
                ws[f'F{row}'] = result['BZ3_Actual']
                ws[f'G{row}'] = result['BZ3_Set']
                ws[f'H{row}'] = result['BZ4_Actual']
                ws[f'I{row}'] = result['BZ4_Set']
                ws[f'J{row}'] = result['DZ1_Actual']
                ws[f'K{row}'] = result['DZ1_Set']
                ws[f'L{row}'] = result['DZ2_Actual']
                ws[f'M{row}'] = result['DZ2_Set']
                ws[f'N{row}'] = result['DZ3_Actual']
                ws[f'O{row}'] = result['DZ3_Set']
                ws[f'P{row}'] = result['Main_Motor_SetE1']
                ws[f'Q{row}'] = result['Main_Motor_ActualE1']
            
            for idx, result in enumerate(result2):
                row = 19 + hour + idx
                ws[f'B{row}'] = result['BZ1_Actual']
                ws[f'C{row}'] = result['BZ1_Set']
                ws[f'D{row}'] = result['BZ2_Actual']
                ws[f'E{row}'] = result['BZ2_Set']
                ws[f'F{row}'] = result['BZ3_Actual']
                ws[f'G{row}'] = result['BZ3_Set']
                ws[f'H{row}'] = result['BZ4_Actual']
                ws[f'I{row}'] = result['BZ4_Set']
                ws[f'J{row}'] = result['DZ1_Actual']
                ws[f'K{row}'] = result['DZ1_Set']
                ws[f'L{row}'] = result['DZ2_Actual']
                ws[f'M{row}'] = result['DZ2_Set']
                ws[f'N{row}'] = result['DZ3_Actual']
                ws[f'O{row}'] = result['DZ3_Set']
                ws[f'P{row}'] = result['Main_Motor_SetE2']
                ws[f'Q{row}'] = result['Main_Motor_ActualE2']

            for idx, result in enumerate(result3):
                row = 29 + hour + idx
                ws[f'B{row}'] = result['BZ1_Actual']
                ws[f'C{row}'] = result['BZ1_Set']
                ws[f'D{row}'] = result['BZ2_Actual']
                ws[f'E{row}'] = result['BZ2_Set']
                ws[f'F{row}'] = result['BZ3_Actual']
                ws[f'G{row}'] = result['BZ3_Set']
                ws[f'H{row}'] = result['BZ4_Actual']
                ws[f'I{row}'] = result['BZ4_Set']
                ws[f'J{row}'] = result['DZ1_Actual']
                ws[f'K{row}'] = result['DZ1_Set']
                ws[f'L{row}'] = result['DZ2_Actual']
                ws[f'M{row}'] = result['DZ2_Set']
                ws[f'N{row}'] = result['Main_Motor_SetE3']
                ws[f'P{row}'] = result['Main_Motor_ActualE3']
            
            for idx, result in enumerate(result4):
                row = 39 + hour + idx
                ws[f'B{row}'] = result['BZ1_Actual']
                ws[f'C{row}'] = result['BZ1_Set']
                ws[f'D{row}'] = result['BZ2_Actual']
                ws[f'E{row}'] = result['BZ2_Set']
                ws[f'F{row}'] = result['BZ3_Actual']
                ws[f'G{row}'] = result['BZ3_Set']
                ws[f'H{row}'] = result['BZ4_Actual']
                ws[f'I{row}'] = result['BZ4_Set']
                ws[f'J{row}'] = result['DZ1_Actual']
                ws[f'K{row}'] = result['DZ1_Set']
                ws[f'L{row}'] = result['DZ2_Actual']
                ws[f'M{row}'] = result['DZ2_Set']
                ws[f'N{row}'] = result['DZ3_Actual']
                ws[f'O{row}'] = result['DZ3_Set']
                ws[f'P{row}'] = result['Main_Motor_SetE4']
                ws[f'Q{row}'] = result['Main_Motor_ActualE4']

            for idx, result in enumerate(result5):
                row = 49 + hour + idx
                ws[f'D{row}'] = result['factor']
                ws[f'F{row}'] = result['Main_Die_Set']
                ws[f'G{row}'] = result['Main_Die_Actual']
                ws[f'H{row}'] = result['Main_Puller_Set']
                ws[f'I{row}'] = result['Main_Puller_Actual']
                ws[f'J{row}'] = result['Mini_Puller_Set']
                ws[f'K{row}'] = result['Mini_Puller_Actual']
                ws[f'L{row}'] = result['Cutting_Machine_Set']
                ws[f'M{row}'] = result['Cutting_Machine_Actual']
                ws[f'N{row}'] = result['Cooling_Trough_Actual']
                ws[f'P{row}'] = result['root_press']
                ws[f'Q{row}'] = result['comp_press']

            for idx, result in enumerate(result6):
                row = 49 + hour + idx
                ws[f'C{row}'] = result['line_speed']
            # new_file_path = f"D:\\PQCR\\PQCR_{mysql_date}({shift})_{model}_{part_name}({current_id}).xlsx"
            save_folder = r"D:\PQCR"  # You can change this path as needed
            os.makedirs(save_folder, exist_ok=True)  # Ensure the folder exists
            save_path = os.path.join(save_folder, f"PQCR_{mysql_date}({shift})_{model}_{part_name}({current_id}).xlsx")
            wb.save(save_path)
            generated_files.append(save_path)

            # return send_file(
            #     new_file_path,
            #     as_attachment=True,
            #     download_name=f"PQCR_{mysql_date}({shift})_{model}_{part_name}({current_id}).xlsx",
            #     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            # )
        cursor.close()
        ic("File generated successfully")
        spr_gen = Production.get_spr_report(shift, shift_start, shift_end, gl_name, line, mysql_date)
        spr_send  = Production.send_spr_email(spr_gen)
        ic("SPR Report Generated Successfully")
        return generated_files
    
    def get_spr_report(shift, shift_start, shift_end, gl_name, line, mysql_date):
        results = None
        # Convert date to match MySQL timestamp format
        # date_obj = datetime.strptime(date, '%Y-%m-%dT%H:%M:%S.%fZ')
        # date_obj += timedelta(days=1)
        # mysql_date = date_obj.strftime('%Y-%m-%d')
        generated_spr = []
        connection = get_db_connection()
        cursor = connection.cursor()

        model_count_query = """
        SELECT COUNT(DISTINCT L8.model, L8.part_name) as count
        FROM L8_Running_Status AS L8
        WHERE DATE(L8.timestamp) = %s AND L8.shift = %s
        """
        
        cursor.execute(model_count_query, (mysql_date, shift))
        countrow = cursor.fetchone()
        model_count = countrow['count'] if countrow else 0
        print("Model Count is : ",model_count)
        
        ###Production Data Query###
        query = """
        SELECT 
            max(L8.cutter1_count) AS cutter1,
            max(L8.cutter2_count) AS cutter2,
            max(L8.line_rej) AS line_rej,
            max(L8.startup_rej) AS startup_rej,
            L8.model,
            L8.part_name,
            P.target_qty,
            MIN(L8.timestamp) AS first_timestamp,
            MAX(L8.timestamp) AS last_timestamp,
            TS.startup_rej AS startup_rej_target,
            TS.line_rej AS line_rej_target,
            MIN(L8.energy) AS first_energy,  
            MAX(L8.energy) AS last_energy
        FROM 
            L8_Running_Status AS L8
        LEFT JOIN 
            planning AS P 
        ON 
            L8.model = P.model AND DATE(L8.timestamp) = DATE(P.date) AND L8.shift = P.shift
        LEFT JOIN 
            target_sheet AS TS 
        ON 
            L8.model = TS.model AND L8.part_name = TS.partna
        WHERE 
            DATE(L8.timestamp) = %s AND L8.shift = %s 
        GROUP BY 
            L8.model, L8.part_name, P.target_qty, TS.startup_rej, TS.line_rej,L8.plan_id
        ORDER BY 
            L8.model, L8.part_name,L8.plan_id
        """
        
        cursor.execute(query, (mysql_date, shift))
        results = cursor.fetchall()

        ###Downtime Query###
        downtime_query = """
        SELECT 
            downtime1, type1, reason1, downtime1_start, downtime1_stop, downtime1_duration,
            downtime2, type2, reason2, downtime2_start, downtime2_stop, downtime2_duration,
            downtime3, type3, reason3, downtime3_start, downtime3_stop, downtime3_duration,
            downtime4, type4, reason4, downtime4_start, downtime4_stop, downtime4_duration,
            downtime5, type5, reason5, downtime5_start, downtime5_stop, downtime5_duration
        FROM 
            planning 
        WHERE 
            DATE(date) = %s AND shift = %s
        """
        
        cursor.execute(downtime_query, (mysql_date, shift))
        downtime_results = cursor.fetchall()

        cursor.close()

        data = []
        for row in results:
            cutter1 = row.get('cutter1', 0)
            cutter2 = row.get('cutter2', 0)
            line_rej = row.get('line_rej', 0)
            startup_rej = row.get('startup_rej', 0)
            model = row.get('model')
            part_name = row.get('part_name')
            target_qty = row.get('target_qty', 0)
            first_timestamp = row.get('first_timestamp')
            last_timestamp = row.get('last_timestamp')
            startup_rej_target = row.get('startup_rej_target', 0)
            line_rej_target = row.get('line_rej_target', 0)
            first_energy = row.get('first_energy', 0)
            last_energy = row.get('last_energy', 0)
            prod_act = (cutter1 + cutter2) or 0
            prod_target = target_qty or 0
            gap = prod_target - prod_act
            startup_gap = startup_rej_target - startup_rej
            line_gap = line_rej_target - line_rej

            # Calculate gap_prod directly using datetime objects
            energy_gap = last_energy - first_energy
            gap_prod = (
                (last_timestamp - first_timestamp).total_seconds() / 60.0
                if first_timestamp and last_timestamp
                else 0
            )
            
            data.append({
                'model_part': f"{model} {part_name}",
                'prod_act': prod_act,
                'prod_target': prod_target,
                'prod_gap': gap,
                'startup_act': startup_rej,
                'startup_rej_target': startup_rej_target,
                'startup_gap': startup_gap,
                'line_act': line_rej,
                'line_rej_target': line_rej_target,
                'line_gap': line_gap,
                'first_prod': first_timestamp,
                'last_prod': last_timestamp,
                'gap_prod': gap_prod,
                'first_energy': first_energy,
                'last_energy': last_energy,
                'energy_gap': energy_gap,
            })

        file_path = r"D:\test\SPR.xlsx"
        new_file_path = f"D:\\test\\SPR_{mysql_date}_{shift}.xlsx"
        download_path = os.path.join(os.path.expanduser("~"), "Downloads", f"SPR_{mysql_date}_{shift}.xlsx")

        wb = load_workbook(file_path)
        ws = wb.active

        # Update header information
        ws['A3'] = f"Date : {mysql_date}"
        ws['B3'] = f"Shift : {shift}"
        ws['D3'] = f"Line : {line}"
        ws['G3'] = f"Team Leader : Narender"
        ws['K3'] = f"Group Leader : {gl_name}"
        ws['O3'] = f"Operator : Kaptan"

        # Start writing data from row 6
        for index, item in enumerate(data, start=6):
            ws[f'A{index}'] = item['model_part']
            ws[f'B{index}'] = item['prod_target']
            ws[f'D{index}'] = item['prod_act']
            ws[f'G{index}'] = item['prod_gap']
            ws[f'H{index}'] = item['startup_rej_target']
            ws[f'J{index}'] = item['startup_act']
            ws[f'L{index}'] = item['startup_gap']
            ws[f'M{index}'] = item['line_rej_target']
            ws[f'N{index}'] = item['line_act']
            ws[f'P{index}'] = item['line_gap']

        if data:
            first_energy = data[0]['first_energy']
            last_energy = data[0]['last_energy']
            energy_gap = data[0]['energy_gap']

            ws['R6'] = first_energy  # First energy
            ws['R8'] = last_energy   # Last energy
            ws['R10'] = energy_gap

        if results:
            for idx, row in enumerate(results, start=33):
                model = row.get('model')
                part_name = row.get('part_name')
                first_timestamp = row.get('first_timestamp')
                last_timestamp = row.get('last_timestamp')
                
                # Calculate duration
                if first_timestamp and last_timestamp:
                    duration = round((last_timestamp - first_timestamp).total_seconds() / 60.0, 1)
                else:
                    duration = 0  # Handle cases where timestamps are missing
                
                # Standard quantity and gap standard
                standard_quantity = 480
                gap_standard = f"{(standard_quantity - duration):.1f}"

                # Write to the spreadsheet
                ws[f'A{idx}'] = f"{model} {part_name} (production)"
                ws[f'B{idx}'] = first_timestamp
                ws[f'C{idx}'] = last_timestamp
                ws[f'D{idx}'] = duration
                ws[f'E{idx}'] = standard_quantity
                ws[f'F{idx}'] = gap_standard
        
        downtime_start_row = model_count + 33
        print("Start row for downtime is : ",downtime_start_row)

        # Function to calculate duration between start and stop times
        def calculate_duration_in_minutes(start_dt, stop_dt):
            return round((stop_dt - start_dt).total_seconds() / 60.0, 1)

        # Start writing downtime data from calculated row
        for downtime in downtime_results:
            for i in range(1, 6):  # i from 1 to 5 for downtime1 ... downtime5
                type_col = downtime.get(f"type{i}")
                start_time = downtime.get(f"downtime{i}_start")
                stop_time = downtime.get(f"downtime{i}_stop")

                if type_col != "NA" and start_time and stop_time:
                    ws[f'A{downtime_start_row}'] = f"{model} {part_name} ({type_col})"
                    ws[f'B{downtime_start_row}'] = start_time
                    ws[f'C{downtime_start_row}'] = stop_time

                    downtime_duration_calculated = calculate_duration_in_minutes(start_time, stop_time)
                    ws[f'D{downtime_start_row}'] = downtime_duration_calculated
                    ws[f'G{downtime_start_row}'] = downtime.get(f"reason{i}")

                    downtime_start_row += 1
        
        wb.save(new_file_path)
        # return send_file(
        #     new_file_path,
        #     as_attachment=True,
        #     download_name=f"SPR_{mysql_date}_{shift}.xlsx",
        #     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        # )
        return new_file_path
    
    def send_pqcr_email(files_to_send):
        msg = EmailMessage()
        msg['Subject'] = "Automated PQCR Reports"
        msg['From'] = Config.MAIL_DEFAULT_SENDER
        msg['To'] = "industrialit@ppapco.com"
        msg.set_content("Attached are the latest PQCR reports.")

        for file_path in files_to_send:
            with open(file_path, 'rb') as f:
                file_data = f.read()
                file_name = os.path.basename(file_path)
                msg.add_attachment(file_data,
                                maintype='application',
                                subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                filename=file_name)

        try:
            # Connect to SMTP server using TLS or SSL based on your config
            if Config.MAIL_USE_SSL:
                server = smtplib.SMTP_SSL(Config.MAIL_SERVER, Config.MAIL_PORT)
            else:
                server = smtplib.SMTP(Config.MAIL_SERVER, Config.MAIL_PORT)
                if Config.MAIL_USE_TLS:
                    server.starttls()

            server.login(Config.MAIL_USERNAME, Config.MAIL_PASSWORD)
            server.send_message(msg)
            server.quit()
            print("Email sent successfully")
        except Exception as e:
            print("Failed to send email:", e)

    def send_spr_email(files_to_send):
        if isinstance(files_to_send, str):
            files_to_send = [files_to_send]
        ic(files_to_send)
        msg = EmailMessage()
        msg['Subject'] = "Automated SPR Report"
        msg['From'] = Config.MAIL_DEFAULT_SENDER
        msg['To'] = ["industrialit@ppapco.com", "saurabhvishwakarma@ppapco.com"]
        msg.set_content("Here is the Attached SPR report")

        for file_path in files_to_send:
            with open(file_path, 'rb') as f:
                file_data = f.read()
                file_name = os.path.basename(file_path)
                msg.add_attachment(file_data,
                                maintype='application',
                                subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                filename=file_name)

        try:
            # Connect to SMTP server using TLS or SSL based on your config
            if Config.MAIL_USE_SSL:
                server = smtplib.SMTP_SSL(Config.MAIL_SERVER, Config.MAIL_PORT)
            else:
                server = smtplib.SMTP(Config.MAIL_SERVER, Config.MAIL_PORT)
                if Config.MAIL_USE_TLS:
                    server.starttls()

            server.login(Config.MAIL_USERNAME, Config.MAIL_PASSWORD)
            server.send_message(msg)
            server.quit()
            print("Email sent successfully")
        except Exception as e:
            print("Failed to send email:", e)

    def Add_part(button, data):
        connection = get_db_connection()
        cursor = connection.cursor()
        if button == 'Add':
            ic("Part Addition")
            query = """
                INSERT INTO target_sheet (
                    part_name, model, customer, changeover_time, startup_time, startup_rej,
                    line_rej, part_weight, standard_line_speed, part_length, productivity,
                    target_ideal, target_w_st, target_w_co, rejection_rate, cycle_time
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            values = (
                data.get('part_name'), data.get('model'), data.get('customer'), data.get('changeover_time'),
                data.get('startup_time'), data.get('startup_rej'), data.get('line_rej'), data.get('part_weight'),
                data.get('standard_line_speed'), data.get('part_length'), data.get('productivity'),
                data.get('target_ideal'), data.get('target_w_st'), data.get('target_w_co'),
                data.get('rejection_rate'), data.get('cycle_time')
            )
            cursor.execute(query, values)

        elif button == 'Update':
            ic("Part Update")
            ic(data)
            modelnew = data.get('modelnew')
            partnanew = data.get('partnanew')
            man = data.get('man')
            man_old = data.get('man_old')
            machine = data.get('line')
            machine_old = data.get('line_old')
            material = data.get('ext1')
            material2 = data.get('ext2')
            material3 = data.get('ext3')
            material4 = data.get('ext4')
            material1_old = data.get('old_ext1')
            material2_old = data.get('old_ext2')
            material3_old = data.get('old_ext3')
            material4_old = data.get('old_ext4')
            username = data.get('username')
            part_id = data.get('part_id')
            abnormal_situation = data.get('abnormal_situation')
            action = data.get('action')
            change_desc = data.get('change_desc')
            setup_approval = data.get('setup_approval')
            type_of_change = data.get('type_of_change')
            retroactive_inspection = data.get('retroactive_inspection')
            suspected_lot = data.get('suspected_lot')
            if 'part_id' not in data:
                return False

            update_fields = []
            update_values = []

            # if man:
            #     cursor.execute("select man from site_map where machine = %s", (machine,))
            #     man_data = cursor.fetchone()
            #     man_fetch = man_data['man']
            #     # ic(man_fetch)
            #     operator_fetch = man_fetch.replace('"','').split(',')
            #     change = 0
            #     for item in operator_fetch:
            #         if man == item:
            #             change += 1
            #     if change == 0:
            #         ic('4M change for Man Detected')
            #     else:
            #         ic('Assigned Operator is Switched')
            if man:
                ic('4M change for Man Detected')
                change_type = 'Man'
                cursor.execute("""
                                insert into four_m_log (part_id, change_type, old_value, new_value, changed_by, change_desc, 
                                type_of_change, action, abnormal_situation, setup_approval, retroactive_inspection, suspected_lot)
                                values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                """, (part_id, change_type, man_old, man, username, change_desc, type_of_change, action, abnormal_situation, setup_approval, retroactive_inspection, suspected_lot))
                connection.commit()

            if machine:
                # cursor.execute("select line from target_sheet where partna = %s and model = %s", (partnanew, modelnew))
                # machine_data = cursor.fetchone()
                # # ic(machine_data)
                # machine_fetch = machine_data['line']
                # # ic(machine_fetch)
                # if machine != machine_fetch:
                ic('4M change for Machine Detected')
                change_type = 'Machine'
                cursor.execute("""
                                insert into four_m_log (part_id, model, partna, change_type, old_value, new_value, changed_by)
                                values (%s, %s, %s, %s, %s, %s, %s)
                                """, (part_id, modelnew, partnanew, change_type, machine_old, machine, username))
                connection.commit()

            if material:
                cursor.execute("select ext1 from target_sheet where partna = %s and model = %s", (partnanew, modelnew))
                material_data = cursor.fetchone()
                material_fetch = material_data['ext1'].replace('"','').split(',')
                change = 0
                for item in material_fetch:
                    if material == item:
                        change += 1
                if change == 0:
                    ic('4M change for Extruder 1 Material')
                    change_type = 'Extruder 1 Material'
                    cursor.execute("""
                                   insert into four_m_log (part_id, model, partna, change_type, old_value, new_value, changed_by)
                                   values (%s, %s, %s, %s, %s, %s, %s)
                                   """, (part_id, modelnew, partnanew, change_type, material1_old, material, username))
                    connection.commit()
                else:
                    ic('Assigned Material is Used')

            if material2:
                cursor.execute("select ext2 from target_sheet where partna = %s and model = %s", (partnanew, modelnew))
                material2_data = cursor.fetchone()
                material2_fetch = material2_data['ext2'].replace('"','').split(',')
                change = 0
                for item in material2_fetch:
                    if material2 == item:
                        change += 1
                if change == 0:
                    ic('4M change for Extruder 2 Material')
                    change_type = 'Extruder 2 Material'
                    cursor.execute("""
                                   insert into four_m_log (part_id, model, partna, change_type, old_value, new_value, changed_by)
                                   values (%s, %s, %s, %s, %s, %s, %s)
                                   """, (part_id, modelnew, partnanew, change_type, material2_old, material2, username))
                    connection.commit()
                else:
                    ic('Assigned Material is Used')
                
            if material3:
                cursor.execute("select ext3 from target_sheet where partna = %s and model = %s", (partnanew, modelnew))
                material3_data = cursor.fetchone()
                material3_fetch = material3_data['ext3'].replace('"','').split(',')
                change = 0
                for item in material3_fetch:
                    if material3 == item:
                        change += 1
                if change == 0:
                    ic('4M change for Extruder 3 Material')
                    change_type = 'Extruder 3 Material'
                    cursor.execute("""
                                   insert into four_m_log (part_id, model, partna, change_type, old_value, new_value, changed_by)
                                   values (%s, %s, %s, %s, %s, %s, %s)
                                   """, (part_id, modelnew, partnanew, change_type, material3_old, material3, username))
                    connection.commit()
                else:
                    ic('Assigned Material is Used')

            if material4:
                cursor.execute("select ext4 from target_sheet where partna = %s and model = %s", (partnanew, modelnew))
                material4_data = cursor.fetchone()
                material4_fetch = material4_data['ext4'].replace('"','').split(',')
                # ic(material4_fetch)
                change = 0
                for item in material4_fetch:
                    if material4 == item:
                        change += 1
                if change == 0:
                    ic('4M change for Extruder 4 Material')
                    change_type = 'Extruder 4 Material'
                    cursor.execute("""
                                   insert into four_m_log (part_id, model, partna, change_type, old_value, new_value, changed_by)
                                   values (%s, %s, %s, %s, %s, %s, %s)
                                   """, (part_id, modelnew, partnanew, change_type, material4_old, material4, username))
                    connection.commit()
                else:
                    ic('Assigned Material is Used')

            fields = [
                'partna', 'model', 'customer', 'changeover_time', 'startup_time',
                'startup_rej', 'line_rej', 'part_weight', 'standard_line_speed',
                'part_length', 'productivity', 'target_ideal', 'target_w_st',
                'target_w_co', 'rejection_rate', 'cycle_time'
            ]

            for field in fields:
                val = data.get(field)
                if val not in [None, '']:  # ✅ Only add if non-empty
                    update_fields.append(f"{field} = %s")
                    update_values.append(val)
            ic(update_fields, update_values)

            if update_fields:
                update_values.append(data['part_id'])  # Append ID at the end for WHERE clause
                query = f"""
                    UPDATE target_sheet SET {', '.join(update_fields)}
                    WHERE part_id = %s
                """
                cursor.execute(query, update_values)

        elif button == 'Delete':
            ic("Part Delete")
            if 'part_id' not in data:
                return False
            query = "DELETE FROM target_sheet WHERE part_id = %s"
            cursor.execute(query, (data['part_id'],))

        else:
            ic("Button State Undefined")
            return False

        connection.commit()
        cursor.close()
        connection.close()
        return True
    
    def Add_site(button, data):
        connection = get_db_connection()
        cursor = connection.cursor()
        if button == 'Add':
            ic("Part Addition")
            query = """
                INSERT INTO site_map (
                    site, plant, department, area, sub_area, machine, man
                ) VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            values = (
                data.get('site'), data.get('plant'), data.get('department'), data.get('area'),
                data.get('subarea'), data.get('machine'), data.get('man')
            )
            cursor.execute(query, values)

        elif button == 'Update':
            ic("Part Update")

        elif button == 'Delete':
            ic("Part Delete")
            ic(data)
            query = "DELETE FROM site_map WHERE id = %s"
            cursor.execute(query, (data['part_id'],))

        else:
            ic("Button State Undefined")
            return False

        connection.commit()
        cursor.close()
        connection.close()
        return True
    
    def crud_user(button, data):
        connection = get_db_connection()
        cursor = connection.cursor()
        if button == 'Add':
            ic("User Addition")
            password_old = data['password']
            password_new = Production.password_hash2(password_old)
            query = """
                INSERT INTO login_new (
                    name, username, password, password_hash, email, department, access
                ) VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            values = (data.get('name'), data.get('username'), password_old, password_new, data.get('email'), data.get('department'), data.get('access'))
            cursor.execute(query, values)

        elif button == 'Update':
            ic("User Update")
            if 'user_id' not in data:
                return False

            update_fields = []
            update_values = []

            # Define all possible fields (you can add/remove fields as per your table structure)
            fields = ['name', 'username', 'password', 'email', 'department', 'access']

            for field in fields:
                val = data.get(field)
                if val not in [None, '']:  # Only add if non-empty
                    if field == 'password':
                        # Handle password hashing
                        password_hash = Production.password_hash2(val)
                        update_fields.append("password = %s")
                        update_values.append(val)
                        update_fields.append("password_hash = %s")
                        update_values.append(password_hash)
                    else:
                        update_fields.append(f"{field} = %s")
                        update_values.append(val)

            ic(update_fields, update_values)

            if not update_fields:
                return False  # No update fields provided

            update_values.append(data['user_id'])  # Append ID for WHERE clause
            query = f"""
                UPDATE login_new SET {', '.join(update_fields)}
                WHERE id = %s
            """
            cursor.execute(query, update_values)

        elif button == 'Delete':
            ic("User Delete")
            # ic(data['user_id'])
            if 'user_id' not in data:
                return False
            query = "DELETE FROM login_new WHERE id = %s"
            cursor.execute(query, (data['user_id'],))

        else:
            ic("Button State Undefined")
            return False

        connection.commit()
        cursor.close()
        connection.close()
        return True

    def add_workflow(button, data):
        connection = get_db_connection()
        cursor = connection.cursor()

        if button == 'Add':
            ic("Workflow Addition")
            query = """
                INSERT INTO workflow (
                    name, site, module, department, users, created_by
                ) VALUES (%s, %s, %s, %s, %s, %s)
            """
            values = (
                data.get('workflow_name'),
                data.get('site'),
                data.get('module'),
                data.get('department'),
                data.get('users'),
                data.get('created_by')
            )
            cursor.execute(query, values)

        elif button == 'Update':
            ic("Workflow Update")
            if 'workflow_id' not in data:
                return False

            update_fields = []
            update_values = []

            fields = ['name', 'site', 'module', 'department', 'users', 'created_by']
            for field in fields:
                val = data.get(field)
                if val not in [None, '']:
                    update_fields.append(f"{field} = %s")
                    update_values.append(val)

            if not update_fields:
                return False  # No fields to update

            update_values.append(data['workflow_id'])  # WHERE clause
            query = f"""
                UPDATE workflow SET {', '.join(update_fields)} WHERE id = %s
            """
            cursor.execute(query, update_values)

        elif button == 'Delete':
            ic("Workflow Delete")
            if 'workflow_id' not in data:
                return False
            query = "DELETE FROM workflow WHERE id = %s"
            cursor.execute(query, (data['workflow_id'],))

        else:
            ic("Button State Undefined")
            return False

        connection.commit()
        cursor.close()
        connection.close()
        return True
    
    def get_questions(workflow_id):
        connection = get_db_connection()
        cursor = connection.cursor()
        try:
            sql = """
            SELECT id, area, question_text, insp_method, input_type, options, required, sort_order
            FROM questions
            WHERE workflow_id = %s
            ORDER BY area, sort_order;
            """
            cursor.execute(sql, (workflow_id,))
            results = cursor.fetchall()

            grouped = {}
            for row in results:
                area = row['area']
                if area not in grouped:
                    grouped[area] = []
                grouped[area].append({
                    'id': row['id'],
                    'question_text': row['question_text'],
                    'insp_method': row['insp_method'],
                    'input_type': row['input_type'],
                    'sort_order': row['sort_order'],
                    'options': json.loads(row['options']) if row['options'] else None,
                    'required': bool(row['required'])
                })
            
            ic(grouped)
            return grouped
        finally:
            cursor.close()
            connection.close()

    def check_auto_downtime():
        connection = get_db_connection()
        cursor = connection.cursor()

        try:
            now = datetime.now()

            time_limits = {
                'case1': timedelta(minutes=10),
                'case2': timedelta(minutes=5),
                'case3': timedelta(minutes=2),
            }

            row_indices = {
                'case1': 60,
                'case2': 30,
                'case3': 12,
            }

            # --- Collect extruder data for all 3 cases ---
            extruder_data_by_case = {'case1': [], 'case2': [], 'case3': []}

            for i in range(1, 5):  # ext1 to ext4
                for case, idx in row_indices.items():
                    query = f"""
                        SELECT timestamp, ext_on{i} AS ext_on, ext_heat{i} AS ext_heat, ext_temp{i} AS ext_temp 
                        FROM l8ext{i}
                        ORDER BY timestamp DESC
                        LIMIT 1 OFFSET {idx - 1}
                    """
                    cursor.execute(query)
                    result = cursor.fetchone()
                    extruder_data_by_case[case].append(result if result else {'ext_on': 0, 'ext_heat': 0, 'ext_temp': 0, 'timestamp': now})
                    
            # ic(extruder_data_by_case)

            # --- Collect common table data for case2 (30th) and case3 (12th) ---
            common_by_case = {}
            for case, idx in {'case2': 30, 'case3': 12}.items():
                query = f"""
                    SELECT datetime, cutter_on, first_piece
                    FROM common
                    ORDER BY datetime DESC
                    LIMIT 1 OFFSET {idx - 1}
                """
                cursor.execute(query)
                result = cursor.fetchone()
                common_by_case[case] = result if result else {'datetime': now, 'cutter_on': 0, 'first_piece': 0}

            # --- Apply downtime logic per case ---
            # Case 1: Heat and Temp ON, Motor OFF for >= 5 min
            case1_all = True
            active_ext1 = 0
            for row in extruder_data_by_case['case1']:
                if row['ext_heat'] == 0:
                    continue
                active_ext1 += 1
                if not (row['ext_temp'] == 1 and row['ext_on'] == 0 and now - row['timestamp'] >= time_limits['case1']):
                    case1_all = False
                    break
            if active_ext1 == 0:
                case1_all = False

            # Case 2: Heat, Temp, Motor ON; Cutter OFF, First Piece OFF; ≥ 10 min
            case2_all = True
            active_ext2 = 0
            for row in extruder_data_by_case['case2']:
                if row['ext_heat'] == 0:
                    continue
                active_ext2 += 1
                if not (row['ext_temp'] == 1 and row['ext_on'] == 1 and now - row['timestamp'] >= time_limits['case2']):
                    case2_all = False
                    break
            common2 = common_by_case['case2']
            if not (common2 and common2['first_piece'] == 0 and common2['cutter_on'] == 0):
                case2_all = False
            if active_ext2 == 0:
                case2_all = False

            # Case 3: Heat, Temp, Motor ON; Cutter OFF, First Piece ON; ≥ 2 min
            case3_all = True
            active_ext3 = 0
            for row in extruder_data_by_case['case3']:
                if row['ext_heat'] == 0:
                    continue
                active_ext3 += 1
                if not (row['ext_temp'] == 1 and row['ext_on'] == 1 and now - row['timestamp'] >= time_limits['case3']):
                    case3_all = False
                    break
            common3 = common_by_case['case3']
            if not (common3 and common3['first_piece'] == 1 and common3['cutter_on'] == 0):
                case3_all = False
            if active_ext3 == 0:
                case3_all = False

            # --- Final Result ---
            downtime_case = None
            if case1_all:
                downtime_case = 'case1'
            elif case2_all:
                downtime_case = 'case2'
            elif case3_all:
                downtime_case = 'case3'

            if downtime_case:
                # Check if there is already an active downtime
                cursor.execute("SELECT id FROM downtime WHERE status = 'active' ORDER BY start_time DESC LIMIT 1")
                active_downtime = cursor.fetchone()

                if not active_downtime:
                    # Insert new downtime entry
                    # Get active plan_id from planning table
                    cursor.execute("SELECT id FROM planning WHERE status = 'active' ORDER BY id DESC LIMIT 1")
                    active_plan = cursor.fetchone()
                    plan_id = active_plan['id'] if active_plan else None

                    cursor.execute("Select id as user_id from login_new where department = 'Production'")
                    all_users = cursor.fetchall()

                    title = 'Downtime for Line 8'
                    message = 'Downtime Occured at Line 8. Please Assign the Department'

                    # Insert into downtime with area and plan_id
                    cursor.execute(
                        "INSERT INTO downtime (start_time, status, area, plan_id) VALUES (%s, %s, %s, %s)",
                        (now - time_limits[downtime_case], 'ACTIVE', 'Line 8', plan_id)
                    )
                    connection.commit()
                    for user in all_users:
                        cursor.execute(
                            "INSERT INTO notifications (user_id, title, message, type, area, related_module, related_id ) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                            (user['user_id'], title, message, 'warning', 'Line 8', 'Downtime Log', plan_id )
                        )
                        connection.commit()
                    ic(f"✅ Downtime started and logged for {downtime_case} at {now}")
                else:
                    ic(f"⏳ Downtime already active. No new insert.")
            else:
                # If no downtime now but a downtime is active, mark it as ended
                cursor.execute("SELECT id FROM downtime WHERE status = 'active' ORDER BY start_time DESC LIMIT 1")
                active_downtime = cursor.fetchone()
                ic(active_downtime)
                if active_downtime:
                    cursor.execute(
                        "UPDATE downtime SET end_time = %s, status = 'ENDED' WHERE id = %s",
                        (now, active_downtime['id'])
                    )
                    connection.commit()
                    ic(f"✅ Downtime ended at {now}")

        finally:
            cursor.close()
            connection.close()

    def maint_downtime():
        connection = get_db_connection()
        cursor = connection.cursor()
        cursor.execute("SELECT id FROM downtime WHERE status = 'active' ORDER BY start_time DESC LIMIT 1")
        active_downtime = cursor.fetchone()
        if active_downtime:
            cursor.execute("SELECT downtime_bit FROM common ORDER BY id DESC LIMIT 1")
            button_result = cursor.fetchone()
            button = button_result['downtime_bit'] if 'downtime_bit' in button_result else button_result
            # button = True
            now = datetime.now()
            if button == True:
                cursor.execute("select id from downtime  where maint_start IS NULL")
                main_startid = cursor.fetchone()
                if main_startid:
                    ic("Please Update Maintenance Downtime Start")
                    cursor.execute("UPDATE downtime set maint_start = %s where status = 'ACTIVE'",(now(),))
                    connection.commit()
                else:
                    ic("Maintenance Already Logged their Time")
            elif button == False:
                cursor.execute("select id from downtime  where maint_end IS NULL")
                main_endid = cursor.fetchone()
                if main_endid:
                    ic("Please Update Maintenance Downtime End")
                    cursor.execute("UPDATE downtime set maint_end = %s where status = 'ACTIVE'", (now(),))
                    connection.commit()
                else:
                    ic("Maintenance Downtime Ended & Logged")
        cursor.close()
        connection.close()
    
    def get_user_notifications(user_id=None):
        connection = get_db_connection()
        cursor = connection.cursor()
        try:
            sql = "SELECT * FROM notifications WHERE user_id = %s ORDER BY created_at DESC"
            cursor.execute(sql, (user_id,))
            result = cursor.fetchall()
            return jsonify(result)
        finally:
            connection.close()

    def get_downtime_log():
        connection = get_db_connection()
        cursor = connection.cursor()

        cursor.execute("SELECT * from downtime")
        downtime_data = cursor.fetchall()

        return downtime_data
    
    def update_downtime_log(department=None, id=None):
        connection = get_db_connection()
        cursor = connection.cursor()

        try:
            # Step 1: Update downtime_log
            dept_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S') 
            ic(dept_time)
            cursor.execute("UPDATE downtime SET department=%s, dept_time = %s WHERE id=%s", (department,dept_time, id))
            connection.commit()
            ic("Downtime Updated")

            # Step 2: Get users from login table with matching department
            cursor.execute("SELECT id FROM login_new WHERE department=%s", (department,))
            users = cursor.fetchall()
            ic("User Get")

            # Step 3: Insert notifications for each user
            for user in users:
                user_id = user['id']
                message = f"You are Assigned the Downtime bearing ID - {id}. Please Update the Reason for the Downtime"
                title = "Reason for Downtime"
                notification_type = "info"
                related_module = "Downtime Log"
                related_id = id
                created_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S') 

                # Log type of each variable
                ic(f"user_id type: {type(user_id)}")
                ic(f"message type: {type(message)}")
                ic(f"title type: {type(title)}")
                ic(f"notification_type type: {type(notification_type)}")
                ic(f"related_module type: {type(related_module)}")
                ic(f"related_id type: {type(related_id)}")
                ic(f"created_at type: {type(created_at)}")

                # Debug log for notification details
                ic(f"Notification details: user_id={user_id}, message={message}, title={title}, notification_type={notification_type}, related_module={related_module}, related_id={related_id}")

                cursor.execute(
                    """
                    INSERT INTO notifications (
                        user_id, title, message, type, related_module, related_id, created_at
                    ) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """, 
                    (user_id, title, message, notification_type, related_module, related_id, created_at)
                )
                ic(f"Notification for user_id {user_id} inserted")

            connection.commit()
            ic("Finished")

            return {"message": "Downtime log updated and notifications sent."}

        except Exception as e:
            connection.rollback()
            ic(f"Error: {str(e)}")
            return {"error": str(e)}

        finally:
            cursor.close()
            connection.close()

    def get_breakdown(department=None):
        try:
            connection = get_db_connection()
            cursor = connection.cursor()
            cursor.execute("SELECT * FROM downtime where department =%s", department)
            rows = cursor.fetchall()
            return rows
        except Exception as e:
            print(f"An error occurred: {e}")
            return None
        finally:
            if connection:
                connection.close()

    def update_downtime_reason(reason=None, id=None):
        connection = get_db_connection()
        cursor = connection.cursor()

        cursor.execute("""update downtime set reason=%s where id=%s""",(reason, id,))
        connection.commit()

        return {'status': 'success', 'message': 'Downtime reason updated successfully'}

    def get_token():
        url = "https://login.thepresence360.com/api/v1/admin-login"
        payload = {
            "username": "ankit@ppapco.com",
            "password": "Welcome@123",
            "grant_type": "admin",
            "client_id": 30,
            "client_secret": "pFOU7fB44ZdZtUu1UQnZBEdd7yk7EHLELgmGKke9"
        }
        headers = {
            "Accept": "application/json",
            "Content-Type": "application/json"
        }
        response = requests.post(url, json=payload, headers=headers)
        ic("Status Code:", response.status_code)
        try:
            token = response.json()["data"]["token"]
            return token
        except Exception as e:
            print("Raw response text:", response.text)
            raise Exception("Failed to get token: " + str(e))

    def fetch_attendance(token, from_date, to_date):
        # BASE_URL = Production.BASE_URL
        url = "https://login.thepresence360.com/api/v1/attendance-records"
        # url = f"{BASE_URL}/v1/attendance-records"
        headers = {
            "Authorization": f"Bearer {token}",
            "Accept": "application/json",
            "Content-Type": "application/json"
        }
        params = {
            "from_date": from_date,
            "to_date": to_date
        }
        response = requests.get(url, headers=headers, params=params)
        ic(response)
        return response.json()
    
    def approve_4m(data):
        connection = get_db_connection()
        cursor = connection.cursor()
        try:
            ic(data)
            sno = data.get('id')
            id = data.get('part_id')
            new_value = data.get('new_value')
            change_type = data.get('change_type')
            approved_on = datetime.now()
            access = data.get('access')
            
            if not all([sno, id, new_value, change_type]):
                return False
            if access =='Super_admin':
                cursor.execute("Select approval_status from four_m_log where id = %s", (sno,))
                approval = cursor.fetchone()
                approval_data = approval['approval_status']
                if approval_data:
                    ic('executing first')
                    cursor.execute("update four_m_log set approval_status_f='Approved', approved_on_f = %s where id = %s", (approved_on, sno,))
                else:
                    ic('executing second')
                    cursor.execute("update four_m_log set approval_status ='Approved', approved_on = %s, approval_status_f='Approved', approved_on_f = %s where id = %s", (approved_on, approved_on, sno,))
            elif access == 'Admin' or access == 'Manager':
                remarks = 'Done by Plant Head' if access =='admin' else 'Done by Manager'
                cursor.execute("update four_m_log set approval_status = 'Approved', approved_on = %s, remarks = %s where id = %s", (approved_on, remarks, sno))

            update_column = ''
            update_column2 = ''
            if change_type == 'Extruder 1 Material':
                update_column = 'ext1'
            elif change_type == 'Extruder 2 Material':
                update_column = 'ext2'
            elif change_type == 'Extruder 3 Material':
                update_column = 'ext3'
            elif change_type == 'Extruder 4 Material':
                update_column = 'ext4'
            elif change_type == 'Machine':
                update_column = 'line'
            elif change_type == 'Man':
                update_column2 = 'man'

            if update_column:
                cursor.execute(
                    f"UPDATE target_sheet SET {update_column} = %s WHERE part_id = %s",
                    (new_value, id)
                )

            if update_column2:
                cursor.execute(
                    f"UPDATE site_map SET {update_column2} = %s WHERE id = %s",
                    (new_value, id)
                )
            connection.commit()
            return True
        except Exception as e:
            connection.rollback()
            ic("Error in approve PM is : ", e)
            return False
        finally:
            cursor.close()
            connection.close()

    def reject_4m(data):
        connection = get_db_connection()
        cursor = connection.cursor()
        try:
            # ic(data)
            sno = data.get('id')
            approved_on = datetime.now()

            if not sno:
                return False
            
            cursor.execute("update four_m_log set approval_status='Rejected', approved_on = %s where id = %s", (approved_on, sno,))
            connection.commit()
            return True
        except Exception as e:
            connection.rollback()
            ic("Error in approve PM is : ", e)
            return False
        finally:
            cursor.close()
            connection.close()

    def co_track():
        connection = get_db_connection()
        cursor = connection.cursor()
        try :
            cursor.execute("Select * from planning where status='Active'")
            data = cursor.fetchone()
            if not data:
                ic("No Active Plans")
                return None
            plan_id = data['id']
            co_end = data['co_end']
            cursor.execute("Select * from common where id = %s order by s_no desc limit 1 ", (plan_id,))
            data2 = cursor.fetchone()
            ext1, ext2, ext3, ext4 = data2['ext1'], data2['ext2'], data2['ext3'], data2['ext4']

            if not co_end :
                if any([ext1, ext2, ext3, ext4]):
                    cursor.execute("Update planning set co_end = NOW() where id = %s", (plan_id,))
                    connection.commit()
                    ic(f"✅ Updated co_end for plan {plan_id}")
                else :
                    ic("Plan Under Change Over")
            return data
        except Exception as e:
            ic("❌ Error in DB co_track:", str(e))
            return None
        finally:
            cursor.close()
            connection.close()
