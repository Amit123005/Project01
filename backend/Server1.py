import os
import pymysql.cursors
import mysql.connector
from mysql.connector import Error
import user_agents
from flask import Flask, request, render_template, redirect, url_for, session, send_from_directory, render_template_string, jsonify, make_response, send_file
from flask import current_app
from flask_mysqldb import MySQL
from datetime import datetime
from flask_cors import CORS
import time
from werkzeug.security import check_password_hash
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import timedelta
import time
from openpyxl import load_workbook
from datetime import datetime
from icecream import ic
from waitress import serve
import logging
import requests
import ollama
from pymodbus.client import ModbusTcpClient
import pandas as pd
from flask_mail import Mail
from config import Config
from models import Production
from Email_Extension import mail
from Scheduler import start_scheduler
from scheduler2 import start_scheduler2
import qrcode
import io
import json
import traceback
# logging.basicConfig(level=logging.DEBUG)
from flask import Flask, jsonify
from pymodbus.client.tcp import ModbusTcpClient
from collections import defaultdict
from datetime import datetime, timedelta

app = Flask(__name__)
CORS(app)
mail.init_app(app)
# MySQL Configuration# MySQL Configuration
DB_HOST = 'localhost'
DB_USER = 'root'
DB_PASSWORD = 'Ankit@321'
DB_NAME = 'line8'

app.config['MYSQL_HOST'] = 'localhost' 
app.config['MYSQL_USER'] = 'root' 
app.config['MYSQL_PASSWORD'] = 'Ankit@321' 
app.config['MYSQL_DB'] = 'line8' 
mysql = MySQL(app)

# Directory to store uploaded files
UPLOAD_FOLDER = 'D:/Projects/Operator मित्र/static'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER



# ===== Helper Function =====
def get_db_connection():
    """Return a pymysql connection with DictCursor."""
    return pymysql.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        cursorclass=pymysql.cursors.DictCursor
    )

def get_db_connection2():
    """Return a pymysql connection with DictCursor."""
    return pymysql.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
    )


def bitupdate2():
    HMI_IP = '163.125.103.169'
    HMI_PORT = 502
    REGISTER_ADDRESS = 50
    # ADDRESS2=2 

    # Initialize the Modbus client
    client = ModbusTcpClient(HMI_IP, port=HMI_PORT)
    if not client.connect():
        return jsonify({"error": "Unable to connect to Modbus server"}), 500

    # Read the current value from the register
    read_response = client.read_holding_registers(REGISTER_ADDRESS, count=1)
    if read_response.isError():
        client.close()
        return jsonify({"error": f"Failed to read from register {REGISTER_ADDRESS}"}), 500
    else:
        current_value = read_response.registers[0]
        print(f"Current value of register {REGISTER_ADDRESS}: {current_value}")

    # Set the specific bit (e.g., M2 bit) to True (1)
    M2_BIT_POSITION = 2  # M2 is typically the 3rd bit (0-indexed)
    
    # updated_value = current_value | (1 << M2_BIT_POSITION)
    first_updated_value = 1
    write_response = client.write_register(REGISTER_ADDRESS, first_updated_value)
    if write_response.isError():
        client.close()
        return jsonify({"error": f"Failed to write first value to register {REGISTER_ADDRESS}"}), 500
    else:
        print(f"Successfully updated register {REGISTER_ADDRESS} with first value: {first_updated_value}")

    # Wait for 2 seconds before sending the second value
    time.sleep(3)

    second_updated_value = 0
    write_response = client.write_register(REGISTER_ADDRESS, second_updated_value)
    if write_response.isError():
        client.close()
        return jsonify({"error": f"Failed to write second value to register {REGISTER_ADDRESS}"}), 500
    else:
        print(f"Successfully updated register {REGISTER_ADDRESS} with second value: {second_updated_value}")

    # counter_reset_value = 0
    # write_response = client.write_register(ADDRESS2, counter_reset_value)
    # if write_response.isError():
    #     return jsonify({"error": f"Failed to reset counter for register {ADDRESS2}"}), 500
    # else:
    #     print(f"Successfully reset counter for register {ADDRESS2}")

    client.close()

    return jsonify({
        "status": "success",
        "register_address": REGISTER_ADDRESS,
        "updated_value": first_updated_value,
        "message": "Successfully set M50 bit to True"
    }), 200

# def bitupdate3():
#     HMI_IP = '163.125.103.169'
#     HMI_PORT = 502
#     REGISTER_ADDRESS = 49

#     # Initialize the Modbus client
#     client = ModbusTcpClient(HMI_IP, port=HMI_PORT)
#     if not client.connect():
#         return jsonify({"error": "Unable to connect to Modbus server"}), 500

#     # Read the current value from the register
#     read_response = client.read_holding_registers(REGISTER_ADDRESS, count=1)
#     if read_response.isError():
#         client.close()
#         return jsonify({"error": f"Failed to read from register {REGISTER_ADDRESS}"}), 500
#     else:
#         current_value = read_response.registers[0]
#         print(f"Current value of register {REGISTER_ADDRESS}: {current_value}")

#     first_updated_value = 1
#     write_response = client.write_register(REGISTER_ADDRESS, first_updated_value)
#     if write_response.isError():
#         client.close()
#         return jsonify({"error": f"Failed to write first value to register {REGISTER_ADDRESS}"}), 500
#     else:
#         ic(f"Successfully updated register {REGISTER_ADDRESS} with first value: {first_updated_value}")

#     return jsonify({
#         "status": "success",
#         "register_address": REGISTER_ADDRESS,
#         "updated_value": first_updated_value,
#         "message": "Successfully set M49 bit to True"
#     }), 200

# def bitupdate4():
#     HMI_IP = '163.125.103.169'
#     HMI_PORT = 502
#     REGISTER_ADDRESS = 49

#     # Initialize the Modbus client
#     client = ModbusTcpClient(HMI_IP, port=HMI_PORT)
#     if not client.connect():
#         return jsonify({"error": "Unable to connect to Modbus server"}), 500

#     # Read the current value from the register
#     read_response = client.read_holding_registers(REGISTER_ADDRESS, count=1)
#     if read_response.isError():
#         client.close()
#         return jsonify({"error": f"Failed to read from register {REGISTER_ADDRESS}"}), 500
#     else:
#         current_value = read_response.registers[0]
#         print(f"Current value of register {REGISTER_ADDRESS}: {current_value}")

#     first_updated_value = 0
#     write_response = client.write_register(REGISTER_ADDRESS, first_updated_value)
#     if write_response.isError():
#         client.close()
#         return jsonify({"error": f"Failed to write first value to register {REGISTER_ADDRESS}"}), 500
#     else:
#         ic(f"Successfully updated register {REGISTER_ADDRESS} with Second value: {first_updated_value}")

#     return jsonify({
#         "status": "success",
#         "register_address": REGISTER_ADDRESS,
#         "updated_value": first_updated_value,
#         "message": "Successfully set M49 bit to False"
#     }), 200

@app.route('/update_plan', methods=['POST'])
def update_plan():
    if request.method == 'POST':
        # Get the form data
        plan_id = request.form['plan_id']
        model = request.form['model']
        partna = request.form['partna']
        partno = request.form['partno']
        plan = request.form['plan']
        target = request.form['target']
        operator = request.form['operator']
        supervisor = request.form['supervisor']
        date = request.form['date']
        time = request.form['time']
        shift = request.form['shift']
        status = request.form['status']

        # Update the plan in the database
        cur = mysql.connection.cursor()
        cur.execute("UPDATE planning SET date=%s, time=%s, shift=%s, model=%s, partna=%s, partno=%s, target_qty=%s, plan_qty=%s, operator_name=%s, supervisor_name=%s, status=%s WHERE id=%s", (date, time, shift, model, partna, partno, target, plan, operator, supervisor, status, plan_id))
        mysql.connection.commit()
        cur.close()
        # Redirect to the running_plans endpoint after updating
        return 'Data updated successfully'

@app.route('/api/create_user', methods=['POST'])
def create_user():
    data = request.get_json()
    name = data.get('name')
    emp_id = data.get('empId')
    department = data.get('department')
    access = data.get('access')

    # Insert the data into the database
    cur = mysql.connection.cursor()
    cur.execute("INSERT INTO users (name, emp_id, department, access) VALUES (%s, %s, %s, %s)", 
                (name, emp_id, department, access))
    mysql.connection.commit()
    cur.close()

    return jsonify({"message": "User created successfully"}), 201

@app.route('/api/ext1', methods=['GET'])
def get_data_ext1():
    connection = get_db_connection()
    cursor = connection.cursor()
    
    query = """
    SELECT  BZ1_Set, BZ1_Actual, BZ2_Set, BZ2_Actual,BZ3_Set, BZ3_Actual,BZ4_Set, BZ4_Actual,
            DZ1_Set, DZ1_Actual,DZ2_Set, DZ2_Actual, DZ3_Set, DZ3_Actual
    FROM L8Ext1
    ORDER BY s_no DESC
    LIMIT 1
    """
    cursor.execute(query)
    rows = cursor.fetchall()

    cursor.close()
    connection.close()

    return jsonify(rows)

@app.route('/api/ext2', methods=['GET'])
def get_data_ext2():
    connection = get_db_connection()
    cursor = connection.cursor()
    
    query = """
    SELECT  BZ1_Set, BZ1_Actual, BZ2_Set, BZ2_Actual,BZ3_Set, BZ3_Actual,BZ4_Set, BZ4_Actual,
            DZ1_Set, DZ1_Actual,DZ2_Set, DZ2_Actual, DZ3_Set, DZ3_Actual
    FROM L8Ext2
    ORDER BY s_no DESC
    LIMIT 1
    """
    cursor.execute(query)
    rows = cursor.fetchall()
    cursor.close()
    connection.close()

    return jsonify(rows)
@app.route('/api/ext3', methods=['GET'])
def get_data_ext3():
    connection = get_db_connection()
    cursor = connection.cursor()
    
    query = """
    SELECT  BZ1_Set, BZ1_Actual, BZ2_Set, BZ2_Actual,BZ3_Set, BZ3_Actual,BZ4_Set, BZ4_Actual,
            DZ1_Set, DZ1_Actual,DZ2_Set, DZ2_Actual
    FROM L8Ext3
    ORDER BY s_no DESC
    LIMIT 1
    """
    cursor.execute(query)
    rows = cursor.fetchall()

    cursor.close()
    connection.close()

    return jsonify(rows)
@app.route('/api/ext4', methods=['GET'])
def get_data_ext4():
    connection = get_db_connection()
    cursor = connection.cursor()
    
    query = """
    SELECT  BZ1_Set, BZ1_Actual, BZ2_Set, BZ2_Actual,BZ3_Set, BZ3_Actual,BZ4_Set, BZ4_Actual,
            DZ1_Set, DZ1_Actual,DZ2_Set, DZ2_Actual, DZ3_Set, DZ3_Actual
    FROM L8Ext4
    ORDER BY s_no DESC
    LIMIT 1
    """
    cursor.execute(query)
    rows = cursor.fetchall()

    cursor.close()
    connection.close()

    return jsonify(rows)

@app.route('/api/partname', methods=['GET'])
def get_partname():
    conn = get_db_connection()
    cursor = conn.cursor()

    query = """
    SELECT part_name FROM L8_Running_Status
    ORDER BY s_no DESC
    LIMIT 1
    """
    cursor.execute(query)
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return rows  # Return only rows fetched from the database
def get_PART():
    data2 = get_partname()
    return jsonify(data2)

@app.route('/api/model', methods=['GET'])
def get_model():
    conn = get_db_connection()
    cursor = conn.cursor()

    query = """
    SELECT model FROM L8_Running_Status
    ORDER BY s_no DESC
    LIMIT 1
    """
    cursor.execute(query)
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return rows  # Return only rows fetched from the database
def get_MOD():
    data2 = get_model()
    return jsonify(data2)

@app.route('/api/linespeed', methods=['GET'])
def get_linespeed():
    conn = get_db_connection()
    cursor = conn.cursor()

    query = """
    SELECT line_speed FROM L8_Running_Status
    ORDER BY s_no DESC
    LIMIT 1
    """
    cursor.execute(query)
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return rows  # Return only rows fetched from the database
def get_LS():
    data2 = get_linespeed()
    return jsonify(data2)

@app.route('/api/factor', methods=['GET'])
def get_factor_data():
    conn = get_db_connection()
    cursor = conn.cursor()
    query = """
    SELECT Factor FROM common
    ORDER BY s_no DESC
    LIMIT 1
    """
    cursor.execute(query)
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(rows)
# def get_factor():
#     rows = get_factor_data()
#     return jsonify(rows)

@app.route('/api/common', methods=['GET'])
def get_common_data():
    conn = get_db_connection()
    cursor = conn.cursor()
    # Fetch common values from all tables
    query = """
    SELECT Main_Motor_Set, Main_Motor_Actual, `Load` FROM L8Ext1
    ORDER BY s_no DESC
    LIMIT 1
    """
    cursor.execute(query)
    data1 = cursor.fetchall()

    query2 = """
    SELECT Main_Motor_Set, Main_Motor_Actual, `Load` FROM L8Ext1
    ORDER BY s_no DESC
    LIMIT 1
    """
    cursor.execute(query2)
    data2 = cursor.fetchall()

    query3 = """
    SELECT Main_Motor_Set, Main_Motor_Actual, `Load` FROM L8Ext3
    ORDER BY s_no DESC
    LIMIT 1
    """
    cursor.execute(query3)
    data3 = cursor.fetchall()

    query4 = """
    SELECT Main_Motor_Set, Main_Motor_Actual, `Load` FROM L8Ext4
    ORDER BY s_no DESC
    LIMIT 1
    """
    cursor.execute(query4)
    data4 = cursor.fetchall()
    
    cursor.close()
    conn.close()
    # Combine data into a single JSON format
    data = {
        "L8Ext1": data1[0] if data1 else None,
        "L8Ext2": data2[0] if data2 else None,
        "L8Ext3": data3[0] if data3 else None,
        "L8Ext4": data4[0] if data4 else None
    }
    return jsonify(data)

def get_active_planning_id():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM planning WHERE status = 'active'")
    result = cursor.fetchone()
    conn.close()
    return result['id'] if result else None

def get_gap_values(id):
    conn = get_db_connection()
    cursor = conn.cursor()

    query = """
    SELECT 
        'L8Ext1', BZ1_Gap, BZ2_Gap, BZ3_Gap, BZ4_Gap, DZ1_Gap, DZ2_Gap, DZ3_Gap
    FROM 
        L8Ext1
    WHERE 
        id = %s AND (BZ1_Gap > 10 OR BZ2_Gap > 10 OR BZ3_Gap > 10 OR BZ4_Gap > 10 OR DZ1_Gap > 10 OR DZ2_Gap > 10 OR DZ3_Gap > 10)
    UNION ALL
    SELECT 
        'L8Ext2', BZ1_Gap, BZ2_Gap, BZ3_Gap, BZ4_Gap, DZ1_Gap, DZ2_Gap, DZ3_Gap
    FROM 
        L8Ext2
    WHERE 
        id = %s AND (BZ1_Gap > 10 OR BZ2_Gap > 10 OR BZ3_Gap > 10 OR BZ4_Gap > 10 OR DZ1_Gap > 10 OR DZ2_Gap > 10 OR DZ3_Gap > 10)
    UNION ALL
    SELECT 
        'L8Ext4', BZ1_Gap, BZ2_Gap, BZ3_Gap, BZ4_Gap, DZ1_Gap, DZ2_Gap, DZ3_Gap
    FROM 
        L8Ext4
    WHERE 
        id = %s AND (BZ1_Gap > 10 OR BZ2_Gap > 10 OR BZ3_Gap > 10 OR BZ4_Gap > 10 OR DZ1_Gap > 10 OR DZ2_Gap > 10 OR DZ3_Gap > 10)
    """

    cursor.execute(query, (id, id, id))
    results = cursor.fetchall()
    conn.close()
    return results

@app.route('/api/gap_values', methods=['GET'])
def gap_values():
    planning_id = get_active_planning_id()
    if not planning_id:
        return jsonify({'error': 'No active planning found'}), 404

    gap_values = get_gap_values(planning_id)
    return jsonify(gap_values)

@app.route('/test', methods=['GET'])
def test():
    conn = get_db_connection()
    cursor = conn.cursor()

    query1 = """
    SELECT BZ1_NC, BZ2_NC, BZ3_NC, BZ4_NC, DZ1_NC, DZ2_NC, DZ3_NC 
    FROM L8Ext1 ORDER BY s_no DESC LIMIT 1
    """
    cursor.execute(query1)
    data1 = cursor.fetchall()

    query2 = """
    SELECT BZ1_NC, BZ2_NC, BZ3_NC, BZ4_NC, DZ1_NC, DZ2_NC, DZ3_NC 
    FROM L8Ext2 ORDER BY s_no DESC LIMIT 1
    """
    cursor.execute(query2)
    data2 = cursor.fetchall()

    query3 = """
    SELECT BZ1_NC, BZ2_NC, BZ3_NC, BZ4_NC, DZ1_NC, DZ2_NC 
    FROM L8Ext3 ORDER BY s_no DESC LIMIT 1
    """
    cursor.execute(query3)
    data3 = cursor.fetchall()

    query4 = """
    SELECT BZ1_NC, BZ2_NC, BZ3_NC, BZ4_NC, DZ1_NC, DZ2_NC, DZ3_NC 
    FROM L8Ext4 ORDER BY s_no DESC LIMIT 1
    """
    cursor.execute(query4)
    data4 = cursor.fetchall()

    cursor.close()
    conn.close()

    # Combine data into a single JSON format
    data = {
        "L8Ext1": data1[0] if data1 else None,
        "L8Ext2": data2[0] if data2 else None,
        "L8Ext3": data3[0] if data3 else None,
        "L8Ext4": data4[0] if data4 else None
    }

    return jsonify(data)


def generate_sql_queries(bit_o, bit_n, machine, col, col2, start_time, end_time):
    queries = []
    for i in range(len(col)):
        if bit_o[col2[i]] == 0 and bit_n[col2[i]] == 1:
            queries.append(f"INSERT INTO params_ng (start_time, machine, machine_zone, status) VALUES ('{start_time}', '{machine}', '{col[i]}', 'Open')")
        elif bit_o[col2[i]] == 1 and bit_n[col2[i]] == 0:
            queries.append(f"UPDATE params_ng SET end_time = '{end_time}', duration = TIMESTAMPDIFF(SECOND, start_time, '{end_time}'), status = 'Closed' WHERE machine = '{machine}' AND machine_zone = '{col[i]}' AND status = 'Open'")
    return queries

def update_params_ng():
    while True:
        with app.app_context():
                conn = get_db_connection()
                cursor = conn.cursor()
                col_common = ['BZ1', 'BZ2', 'BZ3', 'BZ4', 'DZ1', 'DZ2', 'DZ3']
                col2_common = ['BZ1_NC', 'BZ2_NC', 'BZ3_NC', 'BZ4_NC', 'DZ1_NC', 'DZ2_NC', 'DZ3_NC']
                
                col_l8ext3 = ['BZ1', 'BZ2', 'BZ3', 'BZ4', 'DZ1', 'DZ2']
                col2_l8ext3 = ['BZ1_NC', 'BZ2_NC', 'BZ3_NC', 'BZ4_NC', 'DZ1_NC', 'DZ2_NC']

                now = datetime.now()
                start_time = end_time = now.strftime('%Y-%m-%d %H:%M:%S')

                tables = ['L8Ext1', 'L8Ext2', 'L8Ext3', 'L8Ext4']
                
                bit_oe = []
                bit_ne = []
                cursor.execute(f"SELECT BZ1_NC, BZ2_NC, BZ3_NC, BZ4_NC, DZ1_NC, DZ2_NC, DZ3_NC FROM L8Ext4 ORDER BY s_no DESC LIMIT 5")
                data = cursor.fetchone()
                mysql.connection.commit()
                cursor.close()
                # print(f'The Values are: {data}')
                
        time.sleep(1)  # Wait for 1 second before repeating


# def test():
#     while True:
#         print("Success 1")
#         time.sleep(1)

@app.route('/api/plan_id', methods=['GET'])
def get_plan_id():
    conn = get_db_connection()
    cursor = conn.cursor()
    query = """
    SELECT id AS id FROM planning
    where status='active'
    """
    cursor.execute(query)
    rows = cursor.fetchone()
    cursor.close()
    conn.close()
    plan_id = rows['id']
    plan_id_data = "{:.2f}".format(plan_id)
    return jsonify({"Plan_ID": plan_id_data})

@app.route('/api/availability', methods=['GET'])
def get_availability():
    conn = get_db_connection()
    cursor = conn.cursor()
    query = """
    SELECT availability AS availability FROM L8_Running_Status
    ORDER BY s_no DESC
    LIMIT 1
    """
    cursor.execute(query)
    rows = cursor.fetchone()
    cursor.close()
    conn.close()
    availability = rows['availability']
    availability_data = "{:.2f}".format(availability)
    return jsonify({"availability": availability_data})

@app.route('/api/performance', methods=['GET'])
def get_performance():
    conn = get_db_connection()
    cursor = conn.cursor()
    query = """
    SELECT performance AS performance FROM L8_Running_Status
    ORDER BY s_no DESC
    LIMIT 1
    """
    cursor.execute(query)
    rows = cursor.fetchone()
    cursor.close()
    conn.close()
    performance = rows['performance']
    performance_data = "{:.2f}".format(performance)
    return jsonify({"performance": performance_data})

@app.route('/api/quality_rate', methods=['GET'])
def get_quality_rate():
    conn = get_db_connection()
    cursor = conn.cursor()
    query1= """select id from planning where status= 'active'"""
    cursor.execute(query1)
    idget = cursor.fetchone()
    id=idget['id']
    query = """
    SELECT quality_rate AS quality_rate FROM L8_Running_Status
    ORDER BY s_no DESC
    LIMIT 1
    """
    cursor.execute(query)
    rows = cursor.fetchone()
    cursor.close()
    conn.close()
    quality_rate = rows['quality_rate']
    quality = "{:.2f}".format(quality_rate)
    return jsonify({"quality": quality})

@app.route('/api/pms_table', methods=['GET'])
def get_pms_table():
    conn = get_db_connection()
    cursor = conn.cursor()
    query = """
    SELECT plan_id, part_name, model, part_length, part_wt, cutter1_count, cutter2_count, startup_rej, line_rej, availability, performance, quality_rate, oee FROM L8_Running_Status
    ORDER BY s_no DESC
    LIMIT 1
    """
    cursor.execute(query)
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return rows 
def pms_table():
    rows = get_pms_table()
    return jsonify(rows)

@app.route('/api/oee_chart', methods=['GET'])
def get_oee_chart():
    conn = get_db_connection()
    cursor = conn.cursor()
    query = """
    SELECT availability as availability, performance as performance, quality_rate as quality_rate, oee as oee FROM L8_Running_Status
    ORDER BY s_no DESC
    LIMIT 1
    """
    cursor.execute(query)
    rows = cursor.fetchone()
    cursor.close()
    conn.close()
    availability = rows['availability']
    performance = rows['performance']
    quality_rate = rows['quality_rate']
    oee = rows['oee']
    avail_data = f"{availability:.2f}"
    perf_data = f"{performance:.2f}"
    qual_data = f"{quality_rate:.2f}"
    oee_data = f"{oee:.2f}"
    data = {"OEE": oee_data, "Availability" : avail_data, "Quality Rate": qual_data, "Performance": perf_data}
    return jsonify([data])

@app.route('/api/oee', methods=['GET'])
def get_oee():
    conn = get_db_connection()
    cursor = conn.cursor()
    query1 = """select id as id from planning where status='active'"""
    cursor.execute(query1)
    get = cursor.fetchone()

    query = """
    SELECT oee AS oee FROM L8_Running_Status
    ORDER BY s_no DESC
    LIMIT 1
    """
    cursor.execute(query)
    rows = cursor.fetchone()

    cursor.close()
    conn.close()
    oee_value = rows['oee']
    formatted_oee = "{:.2f}".format(oee_value)
    return jsonify({"oee": formatted_oee})

@app.route('/api/test_chart', methods=['GET'])
def get_test_chart():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    shift = request.args.get('shift')
    
    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date"}), 400
    
    start_datetime = f"{start_date} 00:00:00"
    end_datetime = f"{end_date} 23:59:59"

    # Construct base query
    query = """
        SELECT 
            MIN(timestamp) AS first_timestamp,
            part_name AS part_name,
            model AS model
        FROM L8_Running_Status
        WHERE timestamp BETWEEN %s AND %s
    """
    params = [start_datetime, end_datetime]
    # Append shift filter if provided
    if shift:
        # Convert the comma-eparated string into a list
        shift_list = shift.split(',')
        # Dynamically create placeholders for the IN clause
        placeholders = ', '.join(['%s'] * len(shift_list))
        query += f" AND (Shift IN ({placeholders}) OR Shift = 'NA')"
        params.extend(shift_list)
    
    query += """
        GROUP BY part_name, model
    """
    
    # Prepare query parameters
    
    
    # Execute query
    cur = mysql.connection.cursor()
    cur.execute(query, params)
    data = cur.fetchall()
    cur.close()
    
    if not data:
        return jsonify({"error": "No data found"}), 404

    # Prepare the result
    result = []
    for row in data:
        result.append({
            "Date": row[0].strftime("%Y-%m-%d %H:%M:%S"),
            "Part Name": row[1],
            "Model": row[2],
        })
    
    return jsonify(result)


@app.route('/api/yield_filter_chart', methods=['GET'])
def get_yield_filter_chart():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    shift = request.args.get('shift')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date"}), 400
    
    start_datetime = f"{start_date} 00:00:00"
    end_datetime = f"{end_date} 23:59:59"

    cur = mysql.connection.cursor()
    query = """
        SELECT 
            DATE(timestamp) AS date,
            HOUR(timestamp) AS hour,
            MAX(startup_rej) AS max_startup_rej,
            MAX(line_rej) AS max_line_rej,
            MAX(part_wt) AS part_wt,
            MAX(cutter1_count) AS cutter1,
            MAX(cutter2_count) AS cutter2,
            MAX(rej_count) AS rej_count
        FROM L8_Running_Status
        WHERE timestamp BETWEEN %s AND %s
    """
    
    params = [start_datetime, end_datetime]
    
    if shift:
        shift_list = shift.split(',')
        placeholders = ', '.join(['%s'] * len(shift_list))
        query += f" AND (shift IN ({placeholders}) OR shift = 'NA')"
        params.extend(shift_list)
    
    # Group by both date and hour to get hour-wise data
    query += " GROUP BY DATE(timestamp), HOUR(timestamp)"
    
    cur.execute(query, params)
    data = cur.fetchall()
    cur.close()

    result = []
    for row in data:
        date = row[0]
        hour = row[1]
        start = row[2]
        line = row[3]
        part_wt = row[4] / 1000  # Convert part weight to kg
        cutter1 = row[5]
        cutter2 = row[6]
        rej_count = row[7]

        countTotal = cutter1 + cutter2
        line_count = line / part_wt if part_wt > 0 else 0
        start_count = start / part_wt if part_wt > 0 else 0
        # prod_count = countTotal - line_count- start_count
        prod_count = countTotal - line_count- start_count
        actual_prod = prod_count - line_count
        # total = start_count + prod_count
        total = countTotal
        rej = start_count + line_count
        
        yield_per = (actual_prod * 100) / total if total != 0 else 0
        rej_per = (rej * 100) / total if total != 0 else 0
        consumed = actual_prod * part_wt + start + line

        result.append({
            "hour": f"{row[0]} {row[1]:02d}:00:00",
            "Yield Percentage": f"{yield_per:.2f}",
            "Rejection Percentage": f"{rej_per:.2f}",
            "Line Rejection": f"{line:.2f}",
            "Startup Rejection": f"{start:.2f}",
            "Material Consumed": f"{consumed:.2f}"
        })

    return jsonify(result)

@app.route('/api/rej_filter_chart', methods=['GET'])
def get_rej_filter_chart():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    shift = request.args.get('shift')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date"}), 400
    
    start_datetime = f"{start_date} 00:00:00"
    end_datetime = f"{end_date} 23:59:59"

    cur = mysql.connection.cursor()
    query = """
        SELECT 
            DATE(timestamp) AS date,
            HOUR(timestamp) AS hour,
            MAX(startup_rej) AS max_startup_rej,
            MAX(line_rej) AS max_line_rej,
            MAX(part_wt) AS part_wt,
            MAX(cutter1_count) AS cutter1,
            MAX(cutter2_count) AS cutter2,
            MAX(rej_count) AS rej_count
        FROM L8_Running_Status
        WHERE timestamp BETWEEN %s AND %s
    """
    
    params = [start_datetime, end_datetime]
    
    if shift:
        shift_list = shift.split(',')
        placeholders = ', '.join(['%s'] * len(shift_list))
        query += f" AND (shift IN ({placeholders}) OR shift = 'NA')"
        params.extend(shift_list)
    
    # Group by both date and hour to get hour-wise data
    query += " GROUP BY DATE(timestamp), HOUR(timestamp)"
    
    cur.execute(query, params)
    data = cur.fetchall()
    cur.close()

    result = []
    for row in data:
        date = row[0]
        hour = row[1]
        start = row[2]
        line = row[3]
        part_wt = row[4] / 1000  # Convert part weight to kg
        cutter1 = row[5]
        cutter2 = row[6]
        rej_count = row[7]

        countTotal = cutter1 + cutter2
        prod_count = countTotal - rej_count
        line_count = line / part_wt if part_wt > 0 else 0
        start_count = start / part_wt if part_wt > 0 else 0
        actual_prod = prod_count - line_count
        total = start_count + prod_count
        rej = start_count + line_count
        
        yield_per = (actual_prod * 100) / total if total != 0 else 0
        rej_per = (rej * 100) / total if total != 0 else 0
        consumed = actual_prod * part_wt + start + line

        result.append({
            "hour": f"{row[0]} {row[1]:02d}:00:00",
            "Rejection Percentage": f"{rej_per:.2f}"
        })

    return jsonify(result)

@app.route('/api/consumption_filter_chart', methods=['GET'])
def get_consumption_filter_chart():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    shift = request.args.get('shift')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date"}), 400
    
    start_datetime = f"{start_date} 00:00:00"
    end_datetime = f"{end_date} 23:59:59"

    cur = mysql.connection.cursor()
    query = """
        SELECT 
            DATE(timestamp) AS date,
            HOUR(timestamp) AS hour,
            MAX(startup_rej) AS max_startup_rej,
            MAX(line_rej) AS max_line_rej,
            MAX(part_wt) AS part_wt,
            MAX(cutter1_count) AS cutter1,
            MAX(cutter2_count) AS cutter2,
            MAX(rej_count) AS rej_count
        FROM L8_Running_Status
        WHERE timestamp BETWEEN %s AND %s
    """
    
    params = [start_datetime, end_datetime]
    
    if shift:
        shift_list = shift.split(',')
        placeholders = ', '.join(['%s'] * len(shift_list))
        query += f" AND (shift IN ({placeholders}) OR shift = 'NA')"
        params.extend(shift_list)
    
    # Group by both date and hour to get hour-wise data
    query += " GROUP BY DATE(timestamp), HOUR(timestamp)"
    
    cur.execute(query, params)
    data = cur.fetchall()
    cur.close()

    result = []
    for row in data:
        date = row[0]
        hour = row[1]
        start = row[2]
        line = row[3]
        part_wt = row[4] / 1000  # Convert part weight to kg
        cutter1 = row[5]
        cutter2 = row[6]
        rej_count = row[7]

        countTotal = cutter1 + cutter2
        prod_count = countTotal - rej_count
        line_count = line / part_wt if part_wt > 0 else 0
        start_count = start / part_wt if part_wt > 0 else 0
        actual_prod = prod_count - line_count
        total = start_count + prod_count
        rej = start_count + line_count
        
        yield_per = (actual_prod * 100) / total if total != 0 else 0
        rej_per = (rej * 100) / total if total != 0 else 0
        consumed = actual_prod * part_wt + start + line

        result.append({
            "hour": f"{row[0]} {row[1]:02d}:00:00",
            "Material Consumed": f"{consumed:.2f}"
        })

    return jsonify(result)

@app.route('/api/downtime_filter', methods=['GET'])
def downtime_filter():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    conn = get_db_connection()
    cursor = conn.cursor()

    start_datetime = f"{start_date} 00:00:00"
    end_datetime = f"{end_date} 23:59:59"

    query = """
    SELECT 
        DATE(downtime1_start) as date,
        COALESCE(SUM(TIME_TO_SEC(downtime1_duration)), 0) / 60 as downtime1_duration,
        COALESCE(SUM(TIME_TO_SEC(downtime2_duration)), 0) / 60 as downtime2_duration,
        COALESCE(SUM(TIME_TO_SEC(downtime3_duration)), 0) / 60 as downtime3_duration,
        COALESCE(SUM(TIME_TO_SEC(downtime4_duration)), 0) / 60 as downtime4_duration,
        COALESCE(SUM(TIME_TO_SEC(downtime5_duration)), 0) / 60 as downtime5_duration,
        (COALESCE(SUM(TIME_TO_SEC(downtime1_duration)), 0) +
         COALESCE(SUM(TIME_TO_SEC(downtime2_duration)), 0) +
         COALESCE(SUM(TIME_TO_SEC(downtime3_duration)), 0) +
         COALESCE(SUM(TIME_TO_SEC(downtime4_duration)), 0) +
         COALESCE(SUM(TIME_TO_SEC(downtime5_duration)), 0)) / 60 as total_downtime
    FROM planning 
    WHERE downtime1_start BETWEEN %s AND %s
    GROUP BY DATE(downtime1_start)
    ORDER BY DATE(downtime1_start)
    """

    cursor.execute(query, (start_datetime, end_datetime))
    downtime_data = cursor.fetchall()

    cursor.close()
    conn.close()

    if not downtime_data:
        return jsonify({"error": "No active downtime found within the specified date range"}), 404

    return jsonify(downtime_data)

@app.route('/api/first_timestamp_card', methods=['GET'])
def get_first_timestamp_card():
    conn = get_db_connection()
    cursor = conn.cursor()

    query1 = """
        SELECT id
        FROM planning
        WHERE status = 'active';
    """
    cursor.execute(query1)
    results = cursor.fetchall()
    active_plan_ids = results[0]
    id = active_plan_ids['id']
    query2 = """
            SELECT 
                MIN(timestamp) AS first_timestamp,
                MAX(timestamp) AS second_timestamp
            FROM L8_Running_Status
            WHERE plan_id = %s;
        """

    cursor.execute(query2, (id))
    max_values = cursor.fetchone()
    first_timestamp = max_values['first_timestamp']
    return jsonify(first_timestamp)

@app.route('/api/prod_count', methods=['GET'])
def get_prod_count_card():
    conn = get_db_connection()
    cursor = conn.cursor()

    query1 = """
        SELECT id
        FROM planning
        WHERE status = 'active';
    """
    cursor.execute(query1)
    results = cursor.fetchall()

    # if not results:
    #     return jsonify({"error": "No active plans found"}), 404

    active_plan_ids = results[0]
    plan_id = active_plan_ids['id']

    # Query to get cutter counts
    query2 = """
        SELECT 
            MAX(cutter1_count) AS cutter1,
            MAX(cutter2_count) AS cutter2
        FROM L8_Running_Status
        WHERE plan_id = %s;
    """
    cursor.execute(query2, (plan_id,))
    max_values = cursor.fetchone()

    if not max_values:
        return jsonify({"error": "No data found for the active plan"}), 404

    # Safely extract cutter counts with fallback for None
    cutter1 = max_values.get('cutter1', 0) or 0
    cutter2 = max_values.get('cutter2', 0) or 0
    count_total = cutter1 + cutter2

    return jsonify({"count_total": count_total})

    


@app.route('/api/rejection', methods=['GET'])
def get_rej_data():
    conn = get_db_connection()
    cursor = conn.cursor()

    query1 = """
        SELECT id
        FROM planning
        WHERE status = 'active';
    """
    cursor.execute(query1)

    results = cursor.fetchall()

    active_plan_ids = results[0]
    id = active_plan_ids['id']

    query2 = """
            SELECT 
                MAX(startup_rej) AS max_startup_rej,
                MAX(line_rej) AS max_line_rej,
                MAX(part_wt) AS part_wt,
                MAX(cutter1_count) AS cutter1,
                MAX(cutter2_count) AS cutter2,
                MAX(rej_count) AS rej_count,
                MIN(timestamp) AS first_timestamp,
                MAX(timestamp) AS second_timestamp
            FROM L8_Running_Status
            WHERE plan_id = %s;
        """

    cursor.execute(query2, (id))
    max_values = cursor.fetchone()
    first_timestamp = max_values['first_timestamp']
    start = max_values['max_startup_rej']
    line = max_values['max_line_rej']
    if (max_values['part_wt']!=0):
        part_wt= max_values['part_wt']/1000
    else:
        part_wt=0
    countTotal = max_values['cutter1'] + max_values['cutter2']
    rej_count = max_values['rej_count']
    total_rej = line + start
    good_count = countTotal-rej_count
    line_count= line/part_wt
    start_count = start/part_wt
    actual_prod = good_count-line_count
    total = start_count+good_count
    rej= start_count+line_count
    if total != 0:
        yield_per = (actual_prod * 100) / total
    else:
        yield_per = 0
    if total != 0:
        rej_per = (rej*100)/(total)
    else:
        rej_per = 0
    yield_per_formatted = f"{yield_per:.2f}"
    rej_per_formatted = f"{rej_per:.2f}"
    if isinstance(first_timestamp, datetime):
        formatted_timestamp = first_timestamp.strftime('%a, %d %b %Y %H:%M:%S')
    else:
        # Convert string to datetime if needed
        first_timestamp_dt = datetime.strptime(first_timestamp, '%a, %d %b %Y %H:%M:%S %Z')
        formatted_timestamp = first_timestamp_dt.strftime('%a, %d %b %Y %H:%M:%S')
    cursor.close()
    conn.close()
    # print(countTotal, rej_count, good_count, line, line_count, actual_prod)
    # Return the result as JSON
    return jsonify({
        "yield": yield_per_formatted,
        "rejection": rej_per_formatted,
        "timestamp": formatted_timestamp,
        "startup": start,
        "line_rej": line,
        "actual_count": actual_prod,
        "production": countTotal,
        "total_rej": total_rej
    })


@app.route('/api/downtime', methods= ['GET'] )
def downtime():
    conn = get_db_connection()
    cursor = conn.cursor()

    query ="""
    SELECT 
        downtime1, type1, downtime1_start, downtime1_stop, downtime1_duration,
        downtime2, type2, downtime2_start, downtime2_stop, downtime2_duration,
        downtime3, type3, downtime3_start, downtime3_stop, downtime3_duration,
        downtime4, type4, downtime4_start, downtime4_stop, downtime4_duration,
        downtime5, type5, downtime5_start, downtime5_stop, downtime5_duration
    FROM planning where status= 'active'
    """
    cursor.execute(query)

    downtime_data = cursor.fetchone()

    if not downtime_data:
        return jsonify({"error": "No active downtime found"}), 404
    
    def format_datetime(dt):
        return dt.isoformat() if dt else None

    def format_timedelta(td):
        return str(td) if td else "00:00:00"

    downtimes = {
        f"downtime{i+1}": {
            "type": downtime_data.get(f"type{i+1}"),
            "start": format_datetime(downtime_data.get(f"downtime{i+1}_start")),
            "stop": format_datetime(downtime_data.get(f"downtime{i+1}_stop")),
            "duration": format_timedelta(downtime_data.get(f"downtime{i+1}_duration"))
        }
        for i in range(5)
        if downtime_data.get(f"downtime{i+1}") == 'Yes'
    }
    cursor.close()
    conn.close()
    return jsonify({
        "abc": downtimes
    })

@app.route('/api/prod_report', methods=['POST', 'GET'])
def get_prod_report():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    shifts = request.args.get('shift')  # Get a list of shifts
    
    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date"}), 400

    start_datetime = f"{start_date} 00:00:00"
    end_datetime = f"{end_date} 23:59:59"

    query = """
        SELECT 
            DATE(rs.timestamp) as date,
            rs.plan_id as plan_id,
            rs.part_name as Part_name,
            rs.model as Model, 
            MAX(rs.availability) as Availability,
            MAX(rs.performance) as Performance,
            MAX(rs.quality_rate) as Quality_Rate,
            MAX(rs.oee) as OEE,
            MAX(rs.line_rej) as Line_rej,
            MAX(rs.startup_rej) as Startup_rej,
            MAX(rs.cutter1_count) as count1,
            MAX(rs.cutter2_count) as count2,
            MAX(rs.part_wt) as part_wt,
            p.downtime1_duration,
            p.downtime2_duration,
            p.downtime3_duration,
            p.downtime4_duration,
            p.downtime5_duration,
            p.type1,
            p.type2,
            p.type3,
            p.type4,
            p.type5,
            MIN(rs.timestamp) as first_timestamp,
            MAX(rs.timestamp) as last_timestamp,
            rs.shift as shift
        FROM L8_Running_Status rs
        LEFT JOIN planning p ON rs.plan_id = p.id
        WHERE rs.timestamp BETWEEN %s AND %s
    """
    
    # Add conditions for shifts if provided
    params = [start_datetime, end_datetime]
    if shifts:
        shift_placeholders = ", ".join(["%s"] * len(shifts))
        query += f" AND (rs.shift IN ({shift_placeholders}) OR rs.shift = 'NA')"
        params.extend(shifts)
    
    query += """
        GROUP BY DATE(rs.timestamp), rs.part_name, rs.model, rs.plan_id, rs.shift
    """
    
    cur = mysql.connection.cursor()
    cur.execute(query, params)
    data = cur.fetchall()
    cur.close()

    if not data:
        return jsonify({"error": "No data found"}), 404

    result = []
    for row in data:
        # Aggregate downtime durations based on their types
        downtime_duration = {
            "Downtime": 0,
            "Change Over": 0,
            "Setup": 0,
            "Breakdown": 0,
        }

        for i in range(1, 6):
            duration_index = 13 + (i - 1)
            type_index = 18 + (i - 1)     

            downtime_duration_value = row[duration_index]
            downtime_type = row[type_index]

            if isinstance(downtime_type, str):
                downtime_type = downtime_type.strip() if downtime_type else None
                if downtime_type in downtime_duration:
                    downtime_duration[downtime_type] += downtime_duration_value.total_seconds() if downtime_duration_value else 0

        total_downtime = sum(downtime_duration.values())
    
        first_timestamp = row[23]
        last_timestamp = row[24] 
        shift = row[25] 
        
        if first_timestamp and last_timestamp:
            plan_duration = (last_timestamp - first_timestamp).total_seconds()
        else:
            plan_duration = 0

        production_duration = f"{((plan_duration - total_downtime) / 60):.2f}"

        consumption = (
            row[8] + 
            row[9] / 1000 + 
            (row[10] + row[11]) * row[12] / 1000 - 
            row[9] / 1000 
        )
        consumption = f"{consumption:.2f}"
        Availability = f"{row[4]:.2f}"
        Performance = f"{row[5]:.2f}"
        Quality_Rate = f"{row[6]:.2f}"
        OEE = f"{row[7]:.2f}"
        Startup = f"{row[8]:.2f}"
        Line = f"{row[9]:.2f}"
        
        result.append({
            "Date": row[0].strftime("%Y-%m-%d"),
            "Plan ID": row[1],
            "Part Name": row[2],
            "Model": row[3],
            "Availability": Availability,
            "Performance": Performance,
            "Quality Rate": Quality_Rate,
            "OEE": OEE,
            "Startup Rejection": Startup,
            "Line Rejection": Line,
            "Consumption": consumption,
            "Downtime Duration": downtime_duration["Downtime"],
            "Change Over Duration": downtime_duration["Change Over"],
            "Setup Duration": downtime_duration["Setup"],
            "Breakdown Duration": downtime_duration["Breakdown"],
            "Production Duration": production_duration,
            "Shift": shift
        })

    return jsonify(result)


@app.route('/api/prod_plans', methods=['POST', 'GET'])
def get_prod_plans():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    shift = request.args.get('shift')
    
    # print("Data is ", start_date, "end Date is", end_date, "shift is ", shift)
    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date"}), 400

    start_datetime = f"{start_date}"
    end_datetime = f"{end_date}"
    connection = get_db_connection2()
    cur = connection.cursor()

    query = """
        SELECT 
            date as date,
            id as plan_id,
            time as time,
            model as Model, 
            partna as part_name,
            partno as part_number,
            target_qty as target_qty,
            plan_qty as plan_qty,
            status,
            operator_name as operator_name,
            supervisor_name as supervisor_name,
            downtime1 as downtime1,
            type1 as type1,
            downtime1_duration as downtime1_duration,
            downtime2 as downtime2,
            type2 as type2,
            downtime2_duration as downtime2_duration,
            downtime3 as downtime3,
            type3 as type3,
            downtime3_duration as downtime3_duration,
            downtime4 as downtime4,
            type4 as type4,
            downtime4_duration as downtime4_duration,
            downtime5 as downtime5,
            type5 as type5,
            downtime5_duration as downtime5_duration,
            reason1 as reason1,
            reason2 as reason2,
            reason3 as reason3,
            reason4 as reason4,
            reason5 as reason5,
            area as line,
            ext1 as ext1,
            ext2 as ext2,
            ext3 as ext3,
            ext4 as ext4
        FROM planning 
        WHERE date BETWEEN %s AND %s
    """
    params = [start_datetime, end_datetime]
    
    
    if shift:
        # Convert the comma-separated string into a list
        shift_list = shift.split(',')
        # Dynamically create placeholders for the IN clause
        placeholders = ', '.join(['%s'] * len(shift_list))
        query += f" AND (Shift IN ({placeholders}) OR Shift = 'NA')"
        params.extend(shift_list)
    
    query += """
            GROUP BY DATE(date), part_name, model, plan_id
            """
    cur.execute(query, params)
    data = cur.fetchall()
    # ic(data)
    cur.close()
    if not data:
        print("No data found")
        return jsonify({"error": "No data found"}), 404

    def timedelta_to_str(td):
        """Convert a timedelta to a string in HH:MM:SS format."""
        if isinstance(td, timedelta):
            total_seconds = int(td.total_seconds())
            hours, remainder = divmod(total_seconds, 3600)
            minutes, seconds = divmod(remainder, 60)
            return f"{hours:02}:{minutes:02}:{seconds:02}"
        return str(td)

    result = []
    for row in data:

        result.append({
            "Date": row[0].strftime("%Y-%m-%d"),
            "Plan ID": row[1],
            "Time": timedelta_to_str(row[2]),
            "Model": row[3],
            "Part Name": row[4],
            "Part Number": row[5],
            "Target Quantity": row[6],
            "Plan Quantity": row[7],
            "Status": row[8],
            "operator_name": row[9],
            "Supervisor Name": row[10],
            "Downtime 1": row[11],
            "Type 1": row[12],
           "Downtime 1 Duration": timedelta_to_str(row[13]),
            "Downtime 2": row[14],
            "Type 2": row[15],
            "Downtime 2 Duration": timedelta_to_str(row[16]),
            "Downtime 3": row[17],
            "Type 3": row[18],
            "Downtime 3 Duration": timedelta_to_str(row[19]),
            "Downtime 4": row[20],
            "Type 4": row[21],
            "Downtime 4 Duration": timedelta_to_str(row[22]),
            "Downtime 5": row[23],
            "Type 5": row[24],
            "Downtime 5 Duration": timedelta_to_str(row[25]),
            "Reason 1": row[26],
            "Reason 2": row[27],
            "Reason 3": row[28],
            "Reason 4": row[29],
            "Reason 5": row[30],
            "line": row[31],
            "ext1": row[32],
            "ext2": row[33],
            "ext3": row[34],
            "ext4": row[35]
        })

    return jsonify(result)

@app.route('/api/prod_plans_HMI', methods=['GET'])
def get_prod_plansHMI():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # print("Data is ", start_date, "end Date is", end_date, "shift is ", shift)
    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date"}), 400

    start_datetime = f"{start_date}"
    end_datetime = f"{end_date}"

    query = """
        SELECT 
            date as date,
            id as plan_id,
            time as time,
            model as Model, 
            partna as part_name,
            partno as part_number,
            target_qty as target_qty,
            plan_qty as plan_qty,
            status,
            shift
        FROM planning 
        WHERE date BETWEEN %s AND %s
    """
    
    # if shift:
    #     query += " AND (shift = %s OR shift = 'NA')"
    
    query += """
            GROUP BY DATE(date), part_name, model, plan_id
            """
    
    params = [start_datetime, end_datetime]
    # if shift:
    #     params.append(shift)
    
    cur = mysql.connection.cursor()
    cur.execute(query, params)
    data = cur.fetchall()
    cur.close()
    if not data:
        print("No data found")
        return jsonify({"error": "No data found"}), 404

    def timedelta_to_str(td):
        """Convert a timedelta to a string in HH:MM:SS format."""
        if isinstance(td, timedelta):
            total_seconds = int(td.total_seconds())
            hours, remainder = divmod(total_seconds, 3600)
            minutes, seconds = divmod(remainder, 60)
            return f"{hours:02}:{minutes:02}:{seconds:02}"
        return str(td)

    result = []
    for row in data:

        result.append({
            "Date": row[0].strftime("%Y-%m-%d"),
            "Plan ID": row[1],
            "Time": timedelta_to_str(row[2]),
            "Model": row[3],
            "Part Name": row[4],
            "Part Number": row[5],
            "Target Quantity": row[6],
            "Plan Quantity": row[7],
            "Status": row[8],
            "Shift": row[9]
        })

    return jsonify(result)

@app.route('/api/update_plan2', methods=['POST'])
def update_plan2():
    data = request.json
    ic(data)
    plan_id = data.get('plan_id')
    status = data.get('status')
    target = data.get('target')
    model = data.get('model')
    partna = data.get('partna')
    modelnew = data.get('modelnew')
    partnanew = data.get('partnanew')
    man = data.get('man')
    machine = data.get('linenew')
    material = data.get('material')
    material2 = data.get('material2')
    material3 = data.get('material3')
    material4 = data.get('material4')

    # Validation for critical fields
    if not plan_id:
        return jsonify({"error": "Missing 'plan_id'"}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Prepare updates based on valid fields
        updates = []
        values = []

        if man:
            cursor.execute("select man from site_map where machine = %s", (machine,))
            man_data = cursor.fetchone()
            man_fetch = man_data['man']
            # ic(man_fetch)
            operator_fetch = man_fetch.replace('"','').split(',')
            change = 0
            for item in operator_fetch:
                if man == item:
                    change += 1
            if change == 0:
                ic('4M change for Man Detected')
            else:
                ic('Assigned Operator is Switched')

        if machine:
            cursor.execute("select line from target_sheet where partna = %s and model = %s", (partnanew, modelnew))
            machine_data = cursor.fetchone()
            # ic(machine_data)
            machine_fetch = machine_data['line']
            # ic(machine_fetch)
            if machine != machine_fetch:
                ic('4M change for Machine Detected')

        if material:
            cursor.execute("select ext1 from target_sheet where partna = %s and model = %s", (partnanew, modelnew))
            material_data = cursor.fetchone()
            material_fetch = material_data['ext1'].replace('"','').split(',')
            change = 0
            for item in material_fetch:
                if material == item:
                    change += 1
            if change == 0:
                ic('4M change for Extruder 1 Material')
            else:
                ic('Assigned Material is Used')

        if material2:
            cursor.execute("select ext2 from target_sheet where partna = %s and model = %s", (partnanew, modelnew))
            material2_data = cursor.fetchone()
            material2_fetch = material2_data['ext2'].replace('"','').split(',')
            change = 0
            for item in material2_fetch:
                if material2 == item:
                    change += 1
            if change == 0:
                ic('4M change for Extruder 2 Material')
            else:
                ic('Assigned Material is Used')
            
        if material3:
            cursor.execute("select ext3 from target_sheet where partna = %s and model = %s", (partnanew, modelnew))
            material3_data = cursor.fetchone()
            material3_fetch = material3_data['ext3'].replace('"','').split(',')
            change = 0
            for item in material3_fetch:
                if material3 == item:
                    change += 1
            if change == 0:
                ic('4M change for Extruder 3 Material')
            else:
                ic('Assigned Material is Used')

        if material4:
            cursor.execute("select ext4 from target_sheet where partna = %s and model = %s", (partnanew, modelnew))
            material4_data = cursor.fetchone()
            material4_fetch = material4_data['ext4'].replace('"','').split(',')
            change = 0
            for item in material4_fetch:
                if material4 == item:
                    change += 1
            if change == 0:
                ic('4M change for Extruder 4 Material')
            else:
                ic('Assigned Material is Used')


        if status:
            updates.append("status = %s")
            values.append(status)
            bitupdate2()

            if status == 'Active':
                # 1. Close any existing active plans
                cursor.execute("""
                    UPDATE planning
                    SET status = 'Completed',
                        complete_time = NOW()
                    WHERE status = 'Active'
                    AND id = %s
                """, (plan_id,))

                # 2. Set active_time for the new active plan
                cursor.execute("""
                    UPDATE planning
                    SET active_time = NOW()
                    WHERE id = %s
                """, (plan_id,))
                
                conn.commit()

                try:
                    payload = {
                        "model": modelnew,
                        "partna": partnanew,
                        "area": machine,
                        "plant":"PP02"
                    }
                    api_url = "http://163.125.102.142:8500/api/ext_sop"
                    resp = requests.post(api_url, json=payload, timeout=5)
                    resp.raise_for_status()
                    print(f"Sent to ext_sop: {payload} | Response: {resp.text}")
                except requests.RequestException as e:
                    print(f"Error sending to ext_sop API: {e}")

            elif status == "Completed" : 
                cursor.execute("""Update planning set complete_time = NOW()
                               where id = %s""", (plan_id,))
            # bitupdate3()
            # time.sleep(2)
            # bitupdate4()
            # if status == "Active":
            #     query = "SELECT COUNT(plan_id) as test FROM L8_Running_Status WHERE plan_id = %s"
            #     cursor.execute(query, (plan_id,))
            #     result = cursor.fetchone()
            #     ic(result)

            #     # Check if the plan_id exists
            #     if result['test'] > 0:
            #         print(f"Plan ID {plan_id} exists in the table.")
            #     else:
            #         print(f"Plan ID {plan_id} does not exist in the table.")
            #         bitupdate2()

        if target:
            target = float(target)
            updates.append("target_Qty = %s")
            values.append(target)

        if model:
            updates.append("model = %s")
            values.append(model)

        if partna:
            updates.append("partna = %s")
            values.append(partna)

        if target and modelnew and partnanew:
            # Calculate schedule_time
            query1 = "SELECT cycle_time FROM target_sheet WHERE partna = %s AND model = %s"
            cursor.execute(query1, (partnanew, modelnew))
            cycle_time_row = cursor.fetchone()
            cycle_time = cycle_time_row[0] if cycle_time_row else 0
            schedule_time = target * cycle_time / 60
            updates.append("schedule_time = %s")
            values.append(schedule_time)

        # Proceed with the update only if there are valid fields to update
        if updates:
            # Construct the query for updating the planning table
            ic("Executing SQL Query")
            query = f"UPDATE planning SET {', '.join(updates)} WHERE id = %s"
            values.append(plan_id)
            cursor.execute(query, tuple(values))
            ic("Finished Executing SQL Query")

            conn.commit()

            # Additional logic for "On Hold" status
            if status == "On Hold":
                query2 = """
                UPDATE L8_Running_Status 
                SET hold_bit = true 
                WHERE plan_id = %s AND s_no = (
                    SELECT max_s_no FROM (
                        SELECT MAX(s_no) AS max_s_no 
                        FROM L8_Running_Status 
                        WHERE plan_id = %s
                    ) AS subquery
                );
                """
                cursor.execute(query2, (plan_id, plan_id))
                conn.commit()

        else:
            return jsonify({"error": "No valid fields to update"}), 400
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": "An error occurred while updating the plan"}), 500
    finally:
        cursor.close()
        conn.close()

    return jsonify({"message": "Plan updated successfully"})

@app.route('/api/downtimerep', methods=['POST'])
def downtime_rep():
    data = request.json
    plan_id = data.get('plan_id')
    category1 = data.get('category1')
    reason1 = data.get('reason1')
    category2 = data.get('category2')
    reason2 = data.get('reason2')
    category3 = data.get('category3')
    reason3 = data.get('reason3')
    category4 = data.get('category4')
    reason4 = data.get('reason4')
    category5 = data.get('category5')
    reason5 = data.get('reason5')

    # Validation for critical fields
    if not plan_id:
        return jsonify({"error": "Missing 'plan_id'"}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Prepare updates based on valid fields
        updates = []
        values = []

        if category1:
            updates.append("type1 = %s")
            values.append(category1)
        if reason1:
            updates.append("reason1 = %s")
            values.append(reason1)
        if category2:
            updates.append("type2 = %s")
            values.append(category2)
        if reason2:
            updates.append("reason2 = %s")
            values.append(reason2)
        if category3:
            updates.append("type3 = %s")
            values.append(category3)
        if reason3:
            updates.append("reason3 = %s")
            values.append(reason3)
        if category4:
            updates.append("type4 = %s")
            values.append(category4)
        if reason4:
            updates.append("reason4 = %s")
            values.append(reason4)
        if category5:
            updates.append("type5 = %s")
            values.append(category5)
        if reason5:
            updates.append("reason5 = %s")
            values.append(reason5)

        # Proceed with the update only if there are valid fields to update
        if updates:
            query = f"UPDATE planning SET {', '.join(updates)} WHERE id = %s"
            values.append(plan_id)
            cursor.execute(query, tuple(values))
            conn.commit()
        else:
            return jsonify({"error": "No valid fields to update"}), 400
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": "An error occurred while updating the plan"}), 500
    finally:
        cursor.close()
        conn.close()

    return jsonify({"message": "Plan updated successfully"})

def get_shift(time_obj):
    shift_a_start = datetime.strptime("06:00:00", "%H:%M:%S").time()
    shift_a_end = datetime.strptime("14:30:00", "%H:%M:%S").time()
    shift_b_start = datetime.strptime("14:30:00", "%H:%M:%S").time()
    shift_b_end = datetime.strptime("23:00:00", "%H:%M:%S").time()
    shift_c_start = datetime.strptime("23:00:00", "%H:%M:%S").time()
    shift_c_end = datetime.strptime("06:00:00", "%H:%M:%S").time()

    if time_obj is None:
        raise ValueError("time_obj is None")

    # Determine shift
    if shift_a_start <= time_obj < shift_a_end:
        return 'A'
    elif shift_b_start <= time_obj < shift_b_end:
        return 'B'
    elif (shift_c_start <= time_obj or time_obj < shift_c_end):
        return 'C'
    return 'Unknown Shift'

def convert_to_ist(date_str, time_str):
    # Combine date and time into one string
    datetime_str = f"{date_str}T{time_str}"

    # Define the UTC offset in hours (IST is UTC+5:30)
    utc_offset = timedelta(hours=5, minutes=30)
    
    try:
        # Parse datetime string with milliseconds
        utc_datetime = datetime.strptime(datetime_str, "%Y-%m-%dT%H:%M:%S.%f")
    except ValueError:
        try:
            # Parse datetime string without milliseconds
            utc_datetime = datetime.strptime(datetime_str, "%Y-%m-%dT%H:%M:%S")
        except ValueError as e:
            print(f"ValueError: {e}")
            # Return None if parsing fails
            return None, None

    # Apply UTC offset to convert to IST
    ist_datetime = utc_datetime + utc_offset
    
    # Extract date and time in IST
    date_obj = ist_datetime.date()
    time_obj = ist_datetime.time()

    return date_obj, time_obj

@app.route('/api/add_plan2', methods=['POST'])
def add_plan2():
    data = request.json
    # ic(data)
    date_str = data.get('date')
    time_str = data.get('time')
    model = data.get('model')
    partna = data.get('partna')
    partno = data.get('partno')
    target_qty_str = data.get('target_qty')
    plan_qty_str = data.get('plan_qty')
    operator_name = data.get('operator_name')
    supervisor_name = data.get('supervisor_name')
    status = data.get('status')
    target_qty = int(target_qty_str) if target_qty_str.isdigit() else 0
    plan_qty = int(plan_qty_str) if plan_qty_str.isdigit() else 0

    # print('Recieved date is : {date_str}')
    
    date_obj = datetime.fromisoformat(date_str.replace("Z", "+00:00"))
    time_obj = datetime.fromisoformat(time_str.replace("Z", "+00:00"))

    # Combine date and time into a single datetime object
    datetime_obj = datetime.combine(date_obj.date(), time_obj.time())
    
    # Adjust the time by adding 05:30 hours
    adjusted_datetime = datetime_obj + timedelta(hours=5, minutes=30)
    
    # # Subtract one day from the date
    adjusted_datetime += timedelta(days=1)
    
    # Extract adjusted date and time
    formatted_date = adjusted_datetime.strftime('%Y-%m-%d')
    formatted_time = adjusted_datetime.strftime('%H:%M:%S')

    # Convert strings to date and time objects
    date_obj = datetime.strptime(formatted_date, "%Y-%m-%d").date()
    time_obj = datetime.strptime(formatted_time, "%H:%M:%S").time()

    # date_obj = datetime.fromisoformat(date_str.replace("Z", "+05:30"))
    # formatted_date = date_obj.strftime('%Y-%m-%d') 

    # time_obj = datetime.fromisoformat(time_str.replace("Z", "+05:30"))
    # formatted_time = time_obj.strftime('%H:%M:%S')
    
    # date_obj = datetime.strptime(formatted_date, "%Y-%m-%d").date()
    # time_obj = datetime.strptime(formatted_time, "%H:%M:%S").time()

    # date_obj, time_obj = convert_to_ist(date_str, time_str)

    shift = get_shift(time_obj)

    # print(f"date_obj: {date_obj}")
    # print(f"time_obj: {time_obj}")
    # print(shift)

    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute("SELECT cycle_time, ext1, ext2, ext3, ext4, line FROM target_sheet WHERE partna = %s AND model = %s", (partna, model))
    cycle_time_row = cursor.fetchone()
    ic(cycle_time_row)
    if cycle_time_row['cycle_time']:
        cycle_time = cycle_time_row['cycle_time']
        line = cycle_time_row['line']
        ext1 = cycle_time_row['ext1'].split(',')[0].strip().strip('"') if cycle_time_row['ext1'] else None
        ext2 = cycle_time_row['ext2'].split(',')[0].strip().strip('"') if cycle_time_row['ext2'] else None
        ext3 = cycle_time_row['ext3'].split(',')[0].strip().strip('"') if cycle_time_row['ext3'] else None
        ext4 = cycle_time_row['ext4'].split(',')[0].strip().strip('"') if cycle_time_row['ext4'] else None

    else:
        cycle_time = 0

    schedule_time = target_qty * cycle_time/60

    cursor.execute("SELECT man FROM site_map WHERE machine = %s AND plant = 'PP02'", (line,))
    man_data = cursor.fetchone()
    ic(man_data)
    man = None
    if man_data and man_data['man']:
        man = man_data['man'].split(',')[0].strip().strip('"')
    operator_name_new = man if man else operator_name

    if status == 'Active':
        cursor.execute("SELECT * FROM planning WHERE status = 'Active'")
        active_plan = cursor.fetchone()

        if active_plan and data.get('force') != True:
            cursor.close()
            return jsonify({"message": "An active plan already exists. Do you want to mark it as completed and continue?"}), 409

        # Step 2: Mark old active plan as completed (only if forced or none exists)
        if active_plan:
            cursor.execute("UPDATE planning SET status = 'Completed' WHERE status = 'Active'")
            connection.commit()

    update_query = """
    INSERT INTO planning (
        date, time, shift, model, partna, partno, target_qty, plan_qty, 
        operator_name, supervisor_name, status, schedule_time,
        ext1, ext2, ext3, ext4, area
    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
    """
    cursor.execute(update_query, (
        date_obj, time_obj, shift, model, partna, partno, target_qty, plan_qty,
        operator_name_new, supervisor_name, status, schedule_time,
        ext1, ext2, ext3, ext4, line
    ))
    
    connection.commit()
    bitupdate2()
    cursor.close()
    connection.close()

    return jsonify({"message": "Plan updated successfully"}), 200

@app.route('/api/login_new', methods=['POST'])
def login_new():
    data = request.json
    username = data.get('username')
    password = data.get('password')

    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT * FROM login_new WHERE username = %s", (username))
    user = cursor.fetchone()
    # ic(user)

    if user and check_password_hash(user['password_hash'], password):
        # Password matches, return the user's access information
        return jsonify({'status': 'success', 'access_level': user['access'], 'department': user['department'], 'user_id': user['id'], 'username': user['username'], 'name': user['name'], 'email': user['email']})
    else:
        # Invalid credentials
        return jsonify({'status': 'error', 'message': 'Invalid username or password'}), 401

@app.route('/api/password_hash', methods=['GET'])
def password_hash():
    try:
        # Manually hashing and updating passwords
        users = [
            {"username": "login", "password": "login"}
        ]

        conn = get_db_connection()
        cursor = conn.cursor()

        for user in users:
            username = user["username"]
            plain_password = user["password"]
            hashed_password = generate_password_hash(plain_password, method='scrypt')
            
            # Update the user's password in the database with the hashed password
            cursor.execute("UPDATE login_new SET password = %s WHERE username = %s", (hashed_password, username))

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({"message": "Passwords updated successfully"}), 200

    except Exception as e:
        # Return a JSON response with the error message
        return jsonify({"error": str(e)}), 500

@app.route('/api/spr_report', methods=['POST', 'GET'])
def get_spr_report():
    results = None
    if request.method == 'POST':
        data = request.get_json()
        date = data.get('date')
        shift = data.get('shift2')
        line = data.get('line')
        gl_name = data.get('gl_name')
        operator_name = data.get('operator_name')
        supervisor_name = data.get('supervisor_name')
    elif request.method == 'GET':
        date = request.args.get('date')
        shift = request.args.get('shift')
        line = request.args.get('line')
        gl_name = request.args.get('gl_name')
        operator_name = request.args.get('operator_name')
        supervisor_name = request.args.get('supervisor_name')

    date_obj = datetime.strptime(date, '%Y-%m-%dT%H:%M:%S.%fZ')
    date_obj += timedelta(days=1)
    mysql_date = date_obj.strftime('%Y-%m-%d')

    connection = get_db_connection()
    cur = connection.cursor()

    model_count_query = """
    SELECT COUNT(DISTINCT L8.model, L8.part_name) AS model_count
    FROM L8_Running_Status AS L8
    WHERE DATE(L8.timestamp) = %s AND L8.shift = %s
    """
    cur.execute(model_count_query, (mysql_date, shift))
    model_count = cur.fetchone()['model_count']

    query = """
    SELECT 
        MAX(L8.cutter1_count) AS cutter1,
        MAX(L8.cutter2_count) AS cutter2,
        MAX(L8.line_rej) AS line_rej,
        MAX(L8.startup_rej) AS startup_rej,
        L8.model,
        L8.part_name,
        P.target_qty,
        MIN(L8.timestamp) AS first_timestamp,
        MAX(L8.timestamp) AS last_timestamp,
        TS.startup_rej AS startup_rej_target,
        TS.line_rej AS line_rej_target,
        MIN(L8.energy) AS first_energy,  
        MAX(L8.energy) AS last_energy
    FROM 
        L8_Running_Status AS L8
    LEFT JOIN 
        planning AS P 
    ON 
        L8.model = P.model AND DATE(L8.timestamp) = DATE(P.date) AND L8.shift = P.shift
    LEFT JOIN 
        target_sheet AS TS 
    ON 
        L8.model = TS.model AND L8.part_name = TS.partna
    WHERE 
        DATE(L8.timestamp) = %s AND L8.shift = %s 
    GROUP BY 
        L8.model, L8.part_name, P.target_qty, TS.startup_rej, TS.line_rej, L8.plan_id
    ORDER BY 
        L8.model, L8.part_name, L8.plan_id
    """
    cur.execute(query, (mysql_date, shift))
    results = cur.fetchall()

    downtime_query = """
    SELECT plan_id, start_time, end_time, duration, department, reason 
    FROM downtime 
    WHERE DATE(start_time) = %s AND shift = %s
    """
    cur.execute(downtime_query, (mysql_date, shift))
    downtime_results = cur.fetchall()
    cur.close()

    data = []
    for row in results:
        prod_act = (row['cutter1'] or 0) + (row['cutter2'] or 0)
        prod_target = row['target_qty'] or 0
        gap = prod_act - prod_target
        startup_gap = (row['startup_rej_target'] or 0) - (row['startup_rej'] or 0)
        line_gap = (row['line_rej_target'] or 0) - (row['line_rej'] or 0)

        energy_gap = (row['last_energy'] or 0) - (row['first_energy'] or 0)
        gap_prod = (
            (row['last_timestamp'] - row['first_timestamp']).total_seconds() / 60.0
            if row['first_timestamp'] and row['last_timestamp'] else 0
        )

        data.append({
            'model_part': f"{row['model']} {row['part_name']}",
            'prod_act': prod_act,
            'prod_target': prod_target,
            'prod_gap': gap,
            'startup_act': row['startup_rej'],
            'startup_rej_target': row['startup_rej_target'],
            'startup_gap': startup_gap,
            'line_act': row['line_rej'],
            'line_rej_target': row['line_rej_target'],
            'line_gap': line_gap,
            'first_prod': row['first_timestamp'],
            'last_prod': row['last_timestamp'],
            'gap_prod': gap_prod,
            'first_energy': row['first_energy'],
            'last_energy': row['last_energy'],
            'energy_gap': energy_gap,
        })

    file_path = r"D:\\test\\SPR.xlsx"
    new_file_path = f"D:\\test\\SPR_{mysql_date}_{shift}.xlsx"

    wb = load_workbook(file_path)
    ws = wb.active

    ws['A3'] = f"Date : {mysql_date}"
    ws['B3'] = f"Shift : {shift}"
    ws['D3'] = f"Line : {line}"
    ws['G3'] = f"Team Leader : {supervisor_name}"
    ws['K3'] = f"Group Leader : {gl_name}"
    ws['O3'] = f"Operator : {operator_name}"

    for index, item in enumerate(data, start=6):
        ws[f'A{index}'] = item['model_part']
        ws[f'B{index}'] = item['prod_target']
        ws[f'D{index}'] = item['prod_act']
        ws[f'G{index}'] = item['prod_gap']
        ws[f'H{index}'] = item['startup_rej_target']
        ws[f'J{index}'] = item['startup_act']
        ws[f'L{index}'] = item['startup_gap']
        ws[f'M{index}'] = item['line_rej_target']
        ws[f'N{index}'] = item['line_act']
        ws[f'P{index}'] = item['line_gap']

    if data:
        ws['R6'] = data[0]['first_energy']
        ws['R8'] = data[0]['last_energy']
        ws['R10'] = data[0]['energy_gap']

    if results:
        for idx, row in enumerate(results, start=33):
            first_ts = row['first_timestamp']
            last_ts = row['last_timestamp']
            duration = round((last_ts - first_ts).total_seconds() / 60.0, 1) if first_ts and last_ts else 0
            standard_quantity = 480
            gap_standard = f"{(standard_quantity - duration):.1f}"

            ws[f'A{idx}'] = f"{row['model']} {row['part_name']} (production)"
            ws[f'B{idx}'] = first_ts
            ws[f'C{idx}'] = last_ts
            ws[f'D{idx}'] = duration
            ws[f'E{idx}'] = standard_quantity
            ws[f'F{idx}'] = gap_standard

    downtime_start_row = model_count + 33

    if downtime_results:
        for dt in downtime_results:
            if dt['start_time'] and dt['end_time']:
                ws[f'A{downtime_start_row}'] = f"Downtime ({dt['plan_id']})"
                ws[f'B{downtime_start_row}'] = dt['start_time']
                ws[f'C{downtime_start_row}'] = dt['end_time']
                ws[f'D{downtime_start_row}'] = dt['duration']
                ws[f'G{downtime_start_row}'] = dt['department']
                ws[f'L{downtime_start_row}'] = dt['reason']
                downtime_start_row += 1

    wb.save(new_file_path)

    return send_file(
        new_file_path,
        as_attachment=True,
        download_name=f"SPR_{mysql_date}_{shift}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/api/pqcr_report', methods=['POST', 'GET'])
def get_pqcr_report():

    gl_name = "Saurabh Vishwakarma"
    line = "8"
    date = datetime.now()
    mysql_date = date.strftime('%Y-%m-%d')
    time_only = date.time()
    shift_a_start = datetime.strptime("06:00", "%H:%M").time()
    shift_b_start = datetime.strptime("14:00", "%H:%M").time()
    shift_c_start = datetime.strptime("22:00", "%H:%M").time()

    if shift_a_start <= time_only < shift_b_start:
        shift = "A"
        shift_start = date.replace(hour=6, minute=0, second=0)
        shift_end = date.replace(hour=14, minute=0, second=0)

    elif shift_b_start <= time_only < shift_c_start:
        shift = "B"
        shift_start = date.replace(hour=14, minute=0, second=0)
        shift_end = date.replace(hour=22, minute=0, second=0)

    else:
        shift = "C"
        if time_only >= shift_c_start:
            shift_start = date.replace(hour=22, minute=0, second=0)
            shift_end = (shift_start + timedelta(hours=8))
        else:
            # Early morning hours (00:00 to 06:00)
            shift_end = date.replace(hour=6, minute=0, second=0)
            shift_start = shift_end - timedelta(hours=8)

    connection = get_db_connection()
    cursor = connection.cursor()

    unique_ids_query = """
        SELECT DISTINCT id FROM (
            SELECT id FROM l8ext1 WHERE timestamp BETWEEN %s AND %s AND shift = %s
            UNION
            SELECT id FROM l8ext2 WHERE timestamp BETWEEN %s AND %s AND shift = %s
            UNION
            SELECT id FROM l8ext3 WHERE timestamp BETWEEN %s AND %s AND shift = %s
            UNION
            SELECT id FROM l8ext4 WHERE timestamp BETWEEN %s AND %s AND shift = %s
        ) AS combined_ids;
    """

    cursor.execute(unique_ids_query, (
        shift_start, shift_end, shift,
        shift_start, shift_end, shift,
        shift_start, shift_end, shift,
        shift_start, shift_end, shift,
    ))
    all_ids = [row['id'] for row in cursor.fetchall()]
    # ic(all_ids)
    for current_id in all_ids:
        cursor.execute( "Select partna, partno, model from planning where id = %s", (current_id,))
        plannig_data = cursor.fetchone()
        part_name  = plannig_data['partna']
        part_no  = plannig_data['partno']
        model  = plannig_data['model']

        cursor.execute("Select min(hour) as hour from l8ext1 WHERE timestamp BETWEEN %s AND %s AND shift = %s and id = %s", (shift_start, shift_end, shift, current_id))
        hour_result = cursor.fetchone()
        hour = hour_result['hour']
        ic(current_id, plannig_data, hour)

        ext1_query = """SELECT 
                        HOUR(timestamp) AS hour_block,
                        MAX(BZ1_Set) AS BZ1_Set, MAX(BZ1_Actual) AS BZ1_Actual, MAX(BZ2_Set) AS BZ2_Set, MAX(BZ2_Actual) AS BZ2_Actual, MAX(BZ3_Set) AS BZ3_Set,
                        MAX(BZ3_Actual) AS BZ3_Actual, MAX(BZ4_Set) AS BZ4_Set, MAX(BZ4_Actual) AS BZ4_Actual, MAX(DZ1_Set) AS DZ1_Set,
                        MAX(DZ1_Actual) AS DZ1_Actual, MAX(DZ2_Set) AS DZ2_Set, MAX(DZ2_Actual) AS DZ2_Actual, MAX(DZ3_Set) AS DZ3_Set,
                        MAX(DZ3_Actual) AS DZ3_Actual, MAX(Main_Motor_Set) AS Main_Motor_SetE1, MAX(Main_Motor_Actual) AS Main_Motor_ActualE1
                    FROM l8ext1
                    WHERE timestamp BETWEEN %s AND %s and shift = %s AND id = %s
                    GROUP BY HOUR(timestamp)
                    ORDER BY hour_block;
                    """
        ic
        cursor.execute(ext1_query, (shift_start, shift_end, shift, current_id))
        result1= cursor.fetchall()

        ext2_query = """SELECT 
                        HOUR(timestamp) AS hour_block,
                        MAX(BZ1_Set) AS BZ1_Set, MAX(BZ1_Actual) AS BZ1_Actual, MAX(BZ2_Set) AS BZ2_Set, MAX(BZ2_Actual) AS BZ2_Actual, MAX(BZ3_Set) AS BZ3_Set,
                        MAX(BZ3_Actual) AS BZ3_Actual, MAX(BZ4_Set) AS BZ4_Set, MAX(BZ4_Actual) AS BZ4_Actual, MAX(DZ1_Set) AS DZ1_Set,
                        MAX(DZ1_Actual) AS DZ1_Actual, MAX(DZ2_Set) AS DZ2_Set, MAX(DZ2_Actual) AS DZ2_Actual, MAX(DZ3_Set) AS DZ3_Set,
                        MAX(DZ3_Actual) AS DZ3_Actual, MAX(Main_Motor_Set) AS Main_Motor_SetE2, MAX(Main_Motor_Actual) AS Main_Motor_ActualE2
                    FROM l8ext2
                    WHERE timestamp BETWEEN %s AND %s and shift = %s AND id = %s
                    GROUP BY HOUR(timestamp)
                    ORDER BY hour_block;
                    """
        cursor.execute(ext2_query, (shift_start, shift_end, shift, current_id))
        result2= cursor.fetchall()

        ext3_query = """SELECT 
                        HOUR(timestamp) AS hour_block,
                        MAX(BZ1_Set) AS BZ1_Set, MAX(BZ1_Actual) AS BZ1_Actual, MAX(BZ2_Set) AS BZ2_Set, MAX(BZ2_Actual) AS BZ2_Actual, MAX(BZ3_Set) AS BZ3_Set,
                        MAX(BZ3_Actual) AS BZ3_Actual, MAX(BZ4_Set) AS BZ4_Set, MAX(BZ4_Actual) AS BZ4_Actual, MAX(DZ1_Set) AS DZ1_Set,
                        MAX(DZ1_Actual) AS DZ1_Actual, MAX(DZ2_Set) AS DZ2_Set, MAX(DZ2_Actual) AS DZ2_Actual,
                        MAX(Main_Motor_Set) AS Main_Motor_SetE3, MAX(Main_Motor_Actual) AS Main_Motor_ActualE3
                    FROM l8ext3
                    WHERE timestamp BETWEEN %s AND %s and shift = %s AND id = %s
                    GROUP BY HOUR(timestamp)
                    ORDER BY hour_block;
                    """
        cursor.execute(ext3_query, (shift_start, shift_end, shift, current_id))
        result3= cursor.fetchall()

        ext4_query ="""SELECT 
                        HOUR(timestamp) AS hour_block,
                        MAX(BZ1_Set) AS BZ1_Set, MAX(BZ1_Actual) AS BZ1_Actual, MAX(BZ2_Set) AS BZ2_Set, MAX(BZ2_Actual) AS BZ2_Actual, MAX(BZ3_Set) AS BZ3_Set,
                        MAX(BZ3_Actual) AS BZ3_Actual, MAX(BZ4_Set) AS BZ4_Set, MAX(BZ4_Actual) AS BZ4_Actual, MAX(DZ1_Set) AS DZ1_Set,
                        MAX(DZ1_Actual) AS DZ1_Actual, MAX(DZ2_Set) AS DZ2_Set, MAX(DZ2_Actual) AS DZ2_Actual, MAX(DZ3_Set) AS DZ3_Set,
                        MAX(DZ3_Actual) AS DZ3_Actual, MAX(Main_Motor_Set) AS Main_Motor_SetE4, MAX(Main_Motor_Actual) AS Main_Motor_ActualE4
                    FROM l8ext4
                    WHERE timestamp BETWEEN %s AND %s and shift = %s AND id = %s
                    GROUP BY HOUR(timestamp)
                    ORDER BY hour_block;
                    """
        cursor.execute(ext4_query, (shift_start, shift_end, shift, current_id))
        result4= cursor.fetchall()

        common_query = """SELECT 
                        HOUR(datetime) AS hour_block,
                        MAX(Main_Die_Set) AS Main_Die_Set, MAX(Main_Die_Actual) AS Main_Die_Actual, 
                        MAX(Main_Puller_Set) AS Main_Puller_Set, MAX(Main_Puller_Actual) AS Main_Puller_Actual,
                        MAX(Mini_Puller_Set) AS Mini_Puller_Set, MAX(Mini_Puller_Actual) AS Mini_Puller_Actual,
                        MAX(Cutting_Machine_Set) AS Cutting_Machine_Set, MAX(Cutting_Machine_Actual) AS Cutting_Machine_Actual,
                        MAX(Cooling_Trough_Actual) AS Cooling_Trough_Actual,
                        MAX(factor) AS factor
                    FROM common
                    WHERE datetime BETWEEN %s AND %s and shift = %s AND id = %s
                    GROUP BY HOUR(datetime)
                    ORDER BY hour_block;
                    """
        cursor.execute(common_query, (shift_start, shift_end, shift, current_id))
        result5= cursor.fetchall()

        last_query = """SELECT HOUR(timestamp) AS hour_block, MAX(line_speed) as line_speed 
                        FROM L8_Running_Status 
                        WHERE timestamp BETWEEN %s AND %s AND shift = %s AND plan_id = %s
                        GROUP BY HOUR(timestamp)
                        ORDER BY hour_block;
                        """
        cursor.execute(last_query, (shift_start, shift_end, shift, current_id))
        result6= cursor.fetchall()

        file_path = r"D:\PQCR\PQCR2_Format.xlsx"
        # new_file_path = f"D:\\PQCR\\PQCR_{mysql_date}_{shift}.xlsx"
        # download_path = os.path.join(os.path.expanduser("~"), "Downloads", f"PQCR{mysql_date}_{shift}.xlsx")

        wb = load_workbook(file_path)
        ws = wb.active

        # Update header information
        ws['B2'] = f"{mysql_date}"
        ws['B4'] = f"{shift}"
        ws['B6'] = f"{line}"
        ws['K2'] = f"{part_name}"
        ws['K4'] = f"{part_no}"
        ws['K6'] = f"{model}"

        for idx, result in enumerate(result1):
            row = 9 + hour + idx
            ws[f'B{row}'] = result['BZ1_Set']
            ws[f'C{row}'] = result['BZ1_Actual']
            ws[f'D{row}'] = result['BZ2_Set']
            ws[f'E{row}'] = result['BZ2_Actual']
            ws[f'F{row}'] = result['BZ3_Set']
            ws[f'G{row}'] = result['BZ3_Actual']
            ws[f'H{row}'] = result['BZ4_Set']
            ws[f'I{row}'] = result['BZ4_Actual']
            ws[f'J{row}'] = result['DZ1_Set']
            ws[f'K{row}'] = result['DZ1_Actual']
            ws[f'L{row}'] = result['DZ2_Set']
            ws[f'M{row}'] = result['DZ2_Actual']
            ws[f'N{row}'] = result['DZ3_Set']
            ws[f'O{row}'] = result['DZ3_Actual']
            ws[f'P{row}'] = result['Main_Motor_SetE1']
            ws[f'Q{row}'] = result['Main_Motor_ActualE1']
        
        for idx, result in enumerate(result2):
            row = 19 + hour + idx
            ws[f'B{row}'] = result['BZ1_Set']
            ws[f'C{row}'] = result['BZ1_Actual']
            ws[f'D{row}'] = result['BZ2_Set']
            ws[f'E{row}'] = result['BZ2_Actual']
            ws[f'F{row}'] = result['BZ3_Set']
            ws[f'G{row}'] = result['BZ3_Actual']
            ws[f'H{row}'] = result['BZ4_Set']
            ws[f'I{row}'] = result['BZ4_Actual']
            ws[f'J{row}'] = result['DZ1_Set']
            ws[f'K{row}'] = result['DZ1_Actual']
            ws[f'L{row}'] = result['DZ2_Set']
            ws[f'M{row}'] = result['DZ2_Actual']
            ws[f'N{row}'] = result['DZ3_Set']
            ws[f'O{row}'] = result['DZ3_Actual']
            ws[f'P{row}'] = result['Main_Motor_SetE2']
            ws[f'Q{row}'] = result['Main_Motor_ActualE2']

        for idx, result in enumerate(result3):
            row = 29 + hour + idx
            ws[f'B{row}'] = result['BZ1_Set']
            ws[f'C{row}'] = result['BZ1_Actual']
            ws[f'D{row}'] = result['BZ2_Set']
            ws[f'E{row}'] = result['BZ2_Actual']
            ws[f'F{row}'] = result['BZ3_Set']
            ws[f'G{row}'] = result['BZ3_Actual']
            ws[f'H{row}'] = result['BZ4_Set']
            ws[f'I{row}'] = result['BZ4_Actual']
            ws[f'J{row}'] = result['DZ1_Set']
            ws[f'K{row}'] = result['DZ1_Actual']
            ws[f'L{row}'] = result['DZ2_Set']
            ws[f'M{row}'] = result['DZ2_Actual']
            ws[f'N{row}'] = result['Main_Motor_SetE3']
            ws[f'P{row}'] = result['Main_Motor_ActualE3']
        
        for idx, result in enumerate(result4):
            row = 39 + hour + idx
            ws[f'B{row}'] = result['BZ1_Set']
            ws[f'C{row}'] = result['BZ1_Actual']
            ws[f'D{row}'] = result['BZ2_Set']
            ws[f'E{row}'] = result['BZ2_Actual']
            ws[f'F{row}'] = result['BZ3_Set']
            ws[f'G{row}'] = result['BZ3_Actual']
            ws[f'H{row}'] = result['BZ4_Set']
            ws[f'I{row}'] = result['BZ4_Actual']
            ws[f'J{row}'] = result['DZ1_Set']
            ws[f'K{row}'] = result['DZ1_Actual']
            ws[f'L{row}'] = result['DZ2_Set']
            ws[f'M{row}'] = result['DZ2_Actual']
            ws[f'N{row}'] = result['DZ3_Set']
            ws[f'O{row}'] = result['DZ3_Actual']
            ws[f'P{row}'] = result['Main_Motor_SetE4']
            ws[f'Q{row}'] = result['Main_Motor_ActualE4']

        for idx, result in enumerate(result5):
            row = 49 + hour + idx
            ws[f'D{row}'] = result['factor']
            ws[f'F{row}'] = result['Main_Die_Set']
            ws[f'G{row}'] = result['Main_Die_Actual']
            ws[f'H{row}'] = result['Main_Puller_Set']
            ws[f'I{row}'] = result['Main_Puller_Actual']
            ws[f'J{row}'] = result['Mini_Puller_Set']
            ws[f'K{row}'] = result['Mini_Puller_Actual']
            ws[f'L{row}'] = result['Cutting_Machine_Set']
            ws[f'M{row}'] = result['Cutting_Machine_Actual']
            ws[f'N{row}'] = result['Cooling_Trough_Actual']

        for idx, result in enumerate(result6):
            row = 49 + hour + idx
            ws[f'C{row}'] = result['line_speed']
        # new_file_path = f"D:\\PQCR\\PQCR_{mysql_date}({shift})_{model}_{part_name}({current_id}).xlsx"
        save_folder = r"D:\PQCR"  # You can change this path as needed
        os.makedirs(save_folder, exist_ok=True)  # Ensure the folder exists
        save_path = os.path.join(save_folder, f"PQCR_{mysql_date}({shift})_{model}_{part_name}({current_id}).xlsx")
        wb.save(save_path)

        # return send_file(
        #     new_file_path,
        #     as_attachment=True,
        #     download_name=f"PQCR_{mysql_date}({shift})_{model}_{part_name}({current_id}).xlsx",
        #     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        # )
    cursor.close()
    return jsonify({"message": "File generated successfully"}), 200

@app.route('/api/main_die', methods=['GET'])
def get_main_die():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    shifts = request.args.get('shift')  # Retrieve shifts as a comma-separated string

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date"}), 400

    start_datetime = f"{start_date} 00:00:00"
    end_datetime = f"{end_date} 23:59:59"

    query = """
        SELECT DATE(`datetime`) AS date, 
            HOUR(`datetime`) AS hour,
            MAX(Main_Die_Set) AS Main_Die_Set,
            MAX(Main_Die_Actual) AS Main_Die_Actual
        FROM common
        WHERE `datetime` BETWEEN %s AND %s
    """

    params = [start_datetime, end_datetime]

    if shifts:
        # Convert the comma-separated string into a list
        shift_list = shifts.split(',')
        # Dynamically create placeholders for the IN clause
        placeholders = ', '.join(['%s'] * len(shift_list))
        query += f" AND (Shift IN ({placeholders}) OR Shift = 'NA')"
        params.extend(shift_list)

    query += " GROUP BY DATE(`datetime`), HOUR(`datetime`) ORDER BY DATE(`datetime`), HOUR(`datetime`)"

    try:
        cur = mysql.connection.cursor()
        cur.execute(query, params)
        rows = cur.fetchall()
        cur.close()

        # Convert the fetched data to a list of dictionaries
        data = []
        for row in rows:
            data.append({
                "hour": f"{row[0]} {row[1]:02d}:00:00",
                "Main_Die_Actual": row[2],
                "Main_Die_Set": row[3]
            })

        return jsonify(data)

    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": "Internal server error"}), 500


@app.route('/api/common_para_trend', methods=['GET'])
def get_common_para_trend():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    shifts = request.args.get('shift')  # Retrieve shifts as a comma-separated string

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date"}), 400

    start_datetime = f"{start_date} 00:00:00"
    end_datetime = f"{end_date} 23:59:59"

    query = """
        SELECT DATE(`datetime`) AS date, 
            HOUR(`datetime`) AS hour,
            MAX(Main_Puller_Actual) AS Main_Puller_Actual,
            MAX(Main_Puller_Set) AS Main_Puller_Set,
            MAX(Mini_Puller_Actual) AS Mini_Puller_Actual,
            MAX(Mini_Puller_Set) AS Mini_Puller_Set
        FROM common
        WHERE `datetime` BETWEEN %s AND %s
    """

    params = [start_datetime, end_datetime]

    if shifts:
        # Convert the comma-separated string into a list
        shift_list = shifts.split(',')
        # Dynamically create placeholders for the IN clause
        placeholders = ', '.join(['%s'] * len(shift_list))
        query += f" AND (Shift IN ({placeholders}) OR Shift = 'NA')"
        params.extend(shift_list)

    query += " GROUP BY DATE(`datetime`), HOUR(`datetime`) ORDER BY DATE(`datetime`), HOUR(`datetime`)"

    try:
        cur = mysql.connection.cursor()
        cur.execute(query, params)
        rows = cur.fetchall()
        cur.close()

        # Convert the fetched data to a list of dictionaries
        data = []
        for row in rows:
            data.append({
                "hour": f"{row[0]} {row[1]:02d}:00:00",
                "Main_Puller_Actual": row[2],
                "Main_Puller_Set": row[3],
                "Mini_Puller_Actual": row[4],
                "Mini_Puller_Set": row[5]
            })

        return jsonify(data)

    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": "Internal server error"}), 500


@app.route('/api/line_speed_trend', methods=['GET'])
def get_line_speed_trend():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    shifts = request.args.get('shift')  # Retrieve shifts as a comma-separated string

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date"}), 400

    start_datetime = f"{start_date} 00:00:00"
    end_datetime = f"{end_date} 23:59:59"

    query = """
        SELECT DATE(`timestamp`) AS date, 
            HOUR(`timestamp`) AS hour,
            MAX(line_speed) AS Line_Speed
        FROM L8_Running_Status
        WHERE `timestamp` BETWEEN %s AND %s
    """

    params = [start_datetime, end_datetime]

    if shifts:
        # Convert the comma-separated string into a list
        shift_list = shifts.split(',')
        # Dynamically create placeholders for the IN clause
        placeholders = ', '.join(['%s'] * len(shift_list))
        query += f" AND (Shift IN ({placeholders}) OR Shift = 'NA')"
        params.extend(shift_list)

    query += " GROUP BY DATE(`timestamp`), HOUR(`timestamp`) ORDER BY DATE(`timestamp`), HOUR(`timestamp`)"

    try:
        cur = mysql.connection.cursor()
        cur.execute(query, params)
        rows = cur.fetchall()
        cur.close()

        # Convert the fetched data to a list of dictionaries
        data = []
        for row in rows:
            data.append({
                "hour": f"{row[0]} {row[1]:02d}:00:00",
                "Line_Speed": row[2]
            })
        return jsonify(data)

    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": "Internal server error"}), 500


@app.route('/api/ext1_bz_trend', methods=['GET'])
def get_ext1_bz_trend():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    shift = request.args.get('shift')
    # print(f"Received parameters: start_date={start_date}, end_date={end_date}, shift={shift}")

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date"}), 400

    start_datetime = f"{start_date} 00:00:00"
    end_datetime = f"{end_date} 23:59:59"

    query = """
        SELECT DATE(`timestamp`) AS date, 
            HOUR(`timestamp`) AS hour,
            MAX(BZ1_Set) AS BZ1_Set,
            MAX(BZ1_Actual) AS BZ1_Actual,
            MAX(BZ2_Actual) AS BZ2_Actual,
            MAX(BZ2_Set) AS BZ2_Set,
            MAX(BZ3_Actual) AS BZ3_Actual,
            MAX(BZ3_Set) AS BZ3_Set,
            MAX(BZ4_Actual) AS BZ4_Actual,
            MAX(BZ4_Set) AS BZ4_Set,
            model as model,
            part_name as part_name
        FROM L8Ext1
        WHERE `timestamp` BETWEEN %s AND %s
    """
    
    if shift:
        query += " AND (Shift = %s OR Shift = 'NA')"
    
    query += "GROUP BY DATE(`timestamp`), HOUR(`timestamp`), model, part_name ORDER BY DATE(`timestamp`), HOUR(`timestamp`), model, part_name"
    
    params = [start_datetime, end_datetime]
    if shift:
        params.append(shift)
    
    try:
        cur = mysql.connection.cursor()
        cur.execute(query, params)
        rows = cur.fetchall()
        cur.close()
        
        # Convert the fetched data to a list of dictionaries
        data = []
        for row in rows:
            data.append({
                "hour": f"{row[0]} {row[1]:02d}:00:00",
                "BZ1_Actual": row[2],
                "BZ1_Set": row[3],
                "BZ2_Set": row[4],
                "BZ2_Actual": row[5],
                "BZ3_Actual": row[6],
                "BZ3_Set": row[7],
                "BZ4_Actual": row[8],
                "BZ4_Set": row[9],
                "Model": row[10],
                "Part Name": row[11]
            })
        return jsonify(data)
    
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": "Internal server error"}), 500
    
@app.route('/api/ext1_dz_trend', methods=['GET'])
def get_ext1_dz_trend():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    shift = request.args.get('shift')
    # print(f"Received parameters: start_date={start_date}, end_date={end_date}, shift={shift}")

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date"}), 400

    start_datetime = f"{start_date} 00:00:00"
    end_datetime = f"{end_date} 23:59:59"

    query = """
        SELECT DATE(`timestamp`) AS date, 
            HOUR(`timestamp`) AS hour,
            MAX(DZ1_Actual) AS DZ1_Actual,
            MAX(DZ1_Set) AS DZ1_Set,
            MAX(DZ2_Actual) AS DZ2_Actual,
            MAX(DZ2_Set) AS DZ2_Set,
            MAX(DZ3_Actual) AS DZ3_Actual,
            MAX(DZ3_Set) AS DZ3_Set,
            model as model,
            part_name as part_name
        FROM L8Ext1
        WHERE `timestamp` BETWEEN %s AND %s
    """
    
    if shift:
        query += " AND (Shift = %s OR Shift = 'NA')"
    
    query += "GROUP BY DATE(`timestamp`), HOUR(`timestamp`), model, part_name ORDER BY DATE(`timestamp`), HOUR(`timestamp`), model, part_name"
    
    params = [start_datetime, end_datetime]
    if shift:
        params.append(shift)
    
    try:
        cur = mysql.connection.cursor()
        cur.execute(query, params)
        rows = cur.fetchall()
        cur.close()
        
        # Convert the fetched data to a list of dictionaries
        data = []
        for row in rows:
            data.append({
                "hour": f"{row[0]} {row[1]:02d}:00:00",
                "DZ1_Actual": row[2],
                "DZ1_Set": row[3],
                "DZ2_Actual": row[4],
                "DZ2_Set": row[5],
                "DZ3_Actual": row[6],
                "DZ3_Set": row[7],
                "Model": row[8],
                "Part Name": row[9]
            })
        return jsonify(data)
    
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/ext1_trend', methods=['GET'])
def get_ext1_trend():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    shift = request.args.get('shift')
    # print(f"Received parameters: start_date={start_date}, end_date={end_date}, shift={shift}")

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date"}), 400

    start_datetime = f"{start_date} 00:00:00"
    end_datetime = f"{end_date} 23:59:59"

    query = """
        SELECT DATE(`timestamp`) AS date, 
            HOUR(`timestamp`) AS hour,
            MAX(BZ1_Set) AS BZ1_Set,
            MAX(BZ1_Actual) AS BZ1_Actual,
            MAX(BZ2_Actual) AS BZ2_Actual,
            MAX(BZ2_Set) AS BZ2_Set,
            MAX(BZ3_Actual) AS BZ3_Actual,
            MAX(BZ3_Set) AS BZ3_Set,
            MAX(BZ4_Actual) AS BZ4_Actual,
            MAX(BZ4_Set) AS BZ4_Set,
            MAX(DZ1_Actual) AS DZ1_Actual,
            MAX(DZ1_Set) AS DZ1_Set,
            MAX(DZ2_Actual) AS DZ2_Actual,
            MAX(DZ2_Set) AS DZ2_Set,
            MAX(DZ3_Actual) AS DZ3_Actual,
            MAX(DZ3_Set) AS DZ3_Set,
            model as model,
            part_name as part_name,
            MAX(factor) as factor,
            MAX(Main_Motor_Set) as Main_Motor_Set,
            MAX(Main_Motor_Actual) as Main_Motor_Actual
        FROM L8Ext1
        WHERE `timestamp` BETWEEN %s AND %s
    """
    
    if shift:
        query += " AND (Shift = %s OR Shift = 'NA')"
    
    query += "GROUP BY DATE(`timestamp`), HOUR(`timestamp`), model, part_name ORDER BY DATE(`timestamp`), HOUR(`timestamp`), model, part_name"
    
    params = [start_datetime, end_datetime]
    if shift:
        params.append(shift)
    
    try:
        cur = mysql.connection.cursor()
        cur.execute(query, params)
        rows = cur.fetchall()
        cur.close()
        
        # Convert the fetched data to a list of dictionaries
        data = []
        for row in rows:
            data.append({
                "hour": f"{row[0]} {row[1]:02d}:00:00",
                "BZ1_Actual": row[2],
                "BZ1_Set": row[3],
                "BZ2_Set": row[4],
                "BZ2_Actual": row[5],
                "BZ3_Actual": row[6],
                "BZ3_Set": row[7],
                "BZ4_Actual": row[8],
                "BZ4_Set": row[9],
                "DZ1_Actual": row[11],
                "DZ1_Set": row[10],
                "DZ2_Actual": row[13],
                "DZ2_Set": row[12],
                "DZ3_Actual": row[15],
                "DZ3_Set": row[14],
                "Model": row[16],
                "Part Name": row[17],
                "Factor": row[18],
                "Main_Motor_Actual": row[19],
                "Main_Motor_Set": row[20],
            })
        return jsonify(data)
    
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": "Internal server error"}), 500
    
@app.route('/api/ext2_bz_trend', methods=['GET'])
def get_ext2_bz_trend():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    shift = request.args.get('shift')
    # print(f"Received parameters: start_date={start_date}, end_date={end_date}, shift={shift}")

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date"}), 400

    start_datetime = f"{start_date} 00:00:00"
    end_datetime = f"{end_date} 23:59:59"

    query = """
        SELECT DATE(`timestamp`) AS date, 
            HOUR(`timestamp`) AS hour,
            MAX(BZ1_Set) AS BZ1_Set,
            MAX(BZ1_Actual) AS BZ1_Actual,
            MAX(BZ2_Actual) AS BZ2_Actual,
            MAX(BZ2_Set) AS BZ2_Set,
            MAX(BZ3_Actual) AS BZ3_Actual,
            MAX(BZ3_Set) AS BZ3_Set,
            MAX(BZ4_Actual) AS BZ4_Actual,
            MAX(BZ4_Set) AS BZ4_Set,
            model as model,
            part_name as part_name
        FROM L8Ext2
        WHERE `timestamp` BETWEEN %s AND %s
    """
    
    if shift:
        query += " AND (Shift = %s OR Shift = 'NA')"
    
    query += "GROUP BY DATE(`timestamp`), HOUR(`timestamp`), model, part_name ORDER BY DATE(`timestamp`), HOUR(`timestamp`), model, part_name"
    
    params = [start_datetime, end_datetime]
    if shift:
        params.append(shift)
    
    try:
        cur = mysql.connection.cursor()
        cur.execute(query, params)
        rows = cur.fetchall()
        cur.close()
        
        # Convert the fetched data to a list of dictionaries
        data = []
        for row in rows:
            data.append({
                "hour": f"{row[0]} {row[1]:02d}:00:00",
                "BZ1_Actual": row[2],
                "BZ1_Set": row[3],
                "BZ2_Set": row[4],
                "BZ2_Actual": row[5],
                "BZ3_Actual": row[6],
                "BZ3_Set": row[7],
                "BZ4_Actual": row[8],
                "BZ4_Set": row[9],
                "Model": row[10],
                "Part Name": row[11]
            })
        return jsonify(data)
    
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": "Internal server error"}), 500
    
@app.route('/api/ext2_dz_trend', methods=['GET'])
def get_ext2_dz_trend():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    shift = request.args.get('shift')
    # print(f"Received parameters: start_date={start_date}, end_date={end_date}, shift={shift}")

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date"}), 400

    start_datetime = f"{start_date} 00:00:00"
    end_datetime = f"{end_date} 23:59:59"

    query = """
        SELECT DATE(`timestamp`) AS date, 
            HOUR(`timestamp`) AS hour,
            MAX(DZ1_Actual) AS DZ1_Actual,
            MAX(DZ1_Set) AS DZ1_Set,
            MAX(DZ2_Actual) AS DZ2_Actual,
            MAX(DZ2_Set) AS DZ2_Set,
            MAX(DZ3_Actual) AS DZ3_Actual,
            MAX(DZ3_Set) AS DZ3_Set,
            model as model,
            part_name as part_name
        FROM L8Ext2
        WHERE `timestamp` BETWEEN %s AND %s
    """
    
    if shift:
        query += " AND (Shift = %s OR Shift = 'NA')"
    
    query += "GROUP BY DATE(`timestamp`), HOUR(`timestamp`), model, part_name ORDER BY DATE(`timestamp`), HOUR(`timestamp`), model, part_name"
    
    params = [start_datetime, end_datetime]
    if shift:
        params.append(shift)
    
    try:
        cur = mysql.connection.cursor()
        cur.execute(query, params)
        rows = cur.fetchall()
        cur.close()
        
        # Convert the fetched data to a list of dictionaries
        data = []
        for row in rows:
            data.append({
                "hour": f"{row[0]} {row[1]:02d}:00:00",
                "DZ1_Actual": row[2],
                "DZ1_Set": row[3],
                "DZ2_Actual": row[4],
                "DZ2_Set": row[5],
                "DZ3_Actual": row[6],
                "DZ3_Set": row[7],
                "Model": row[8],
                "Part Name": row[9]
            })
        return jsonify(data)
    
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/ext2_trend', methods=['GET'])
def get_ext2_trend():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    shift = request.args.get('shift')
    # print(f"Received parameters: start_date={start_date}, end_date={end_date}, shift={shift}")

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date"}), 400

    start_datetime = f"{start_date} 00:00:00"
    end_datetime = f"{end_date} 23:59:59"

    query = """
        SELECT DATE(`timestamp`) AS date, 
            HOUR(`timestamp`) AS hour,
            MAX(BZ1_Set) AS BZ1_Set,
            MAX(BZ1_Actual) AS BZ1_Actual,
            MAX(BZ2_Actual) AS BZ2_Actual,
            MAX(BZ2_Set) AS BZ2_Set,
            MAX(BZ3_Actual) AS BZ3_Actual,
            MAX(BZ3_Set) AS BZ3_Set,
            MAX(BZ4_Actual) AS BZ4_Actual,
            MAX(BZ4_Set) AS BZ4_Set,
            MAX(DZ1_Actual) AS DZ1_Actual,
            MAX(DZ1_Set) AS DZ1_Set,
            MAX(DZ2_Actual) AS DZ2_Actual,
            MAX(DZ2_Set) AS DZ2_Set,
            MAX(DZ3_Actual) AS DZ3_Actual,
            MAX(DZ3_Set) AS DZ3_Set,
            model as model,
            part_name as part_name,
            MAX(factor) as factor,
            MAX(Main_Motor_Set) as Main_Motor_Set,
            MAX(Main_Motor_Actual) as Main_Motor_Actual
        FROM L8Ext2
        WHERE `timestamp` BETWEEN %s AND %s
    """
    
    if shift:
        query += " AND (Shift = %s OR Shift = 'NA')"
    
    query += "GROUP BY DATE(`timestamp`), HOUR(`timestamp`), model, part_name ORDER BY DATE(`timestamp`), HOUR(`timestamp`), model, part_name"
    
    params = [start_datetime, end_datetime]
    if shift:
        params.append(shift)
    
    try:
        cur = mysql.connection.cursor()
        cur.execute(query, params)
        rows = cur.fetchall()
        cur.close()
        
        # Convert the fetched data to a list of dictionaries
        data = []
        for row in rows:
            data.append({
                "hour": f"{row[0]} {row[1]:02d}:00:00",
                "BZ1_Actual": row[2],
                "BZ1_Set": row[3],
                "BZ2_Set": row[4],
                "BZ2_Actual": row[5],
                "BZ3_Actual": row[6],
                "BZ3_Set": row[7],
                "BZ4_Actual": row[8],
                "BZ4_Set": row[9],
                "DZ1_Actual": row[11],
                "DZ1_Set": row[10],
                "DZ2_Actual": row[13],
                "DZ2_Set": row[12],
                "DZ3_Actual": row[15],
                "DZ3_Set": row[14],
                "Model": row[16],
                "Part Name": row[17],
                "Factor": row[18],
                "Main_Motor_Actual": row[19],
                "Main_Motor_Set": row[20],
            })
        return jsonify(data)
    
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/ext3_bz_trend', methods=['GET'])
def get_ext3_bz_trend():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    shift = request.args.get('shift')
    # print(f"Received parameters: start_date={start_date}, end_date={end_date}, shift={shift}")

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date"}), 400

    start_datetime = f"{start_date} 00:00:00"
    end_datetime = f"{end_date} 23:59:59"

    query = """
        SELECT DATE(`timestamp`) AS date, 
            HOUR(`timestamp`) AS hour,
            MAX(BZ1_Set) AS BZ1_Set,
            MAX(BZ1_Actual) AS BZ1_Actual,
            MAX(BZ2_Actual) AS BZ2_Actual,
            MAX(BZ2_Set) AS BZ2_Set,
            MAX(BZ3_Actual) AS BZ3_Actual,
            MAX(BZ3_Set) AS BZ3_Set,
            MAX(BZ4_Actual) AS BZ4_Actual,
            MAX(BZ4_Set) AS BZ4_Set,
            model as model,
            part_name as part_name
        FROM L8Ext3
        WHERE `timestamp` BETWEEN %s AND %s
    """
    
    if shift:
        query += " AND (Shift = %s OR Shift = 'NA')"
    
    query += "GROUP BY DATE(`timestamp`), HOUR(`timestamp`), model, part_name ORDER BY DATE(`timestamp`), HOUR(`timestamp`), model, part_name"
    
    params = [start_datetime, end_datetime]
    if shift:
        params.append(shift)
    
    try:
        cur = mysql.connection.cursor()
        cur.execute(query, params)
        rows = cur.fetchall()
        cur.close()
        
        # Convert the fetched data to a list of dictionaries
        data = []
        for row in rows:
            data.append({
                "hour": f"{row[0]} {row[1]:02d}:00:00",
                "BZ1_Actual": row[2],
                "BZ1_Set": row[3],
                "BZ2_Set": row[4],
                "BZ2_Actual": row[5],
                "BZ3_Actual": row[6],
                "BZ3_Set": row[7],
                "BZ4_Actual": row[8],
                "BZ4_Set": row[9],
                "Model": row[10],
                "Part Name": row[11]
            })
        return jsonify(data)
    
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": "Internal server error"}), 500
    
@app.route('/api/ext3_dz_trend', methods=['GET'])
def get_ext3_dz_trend():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    shift = request.args.get('shift')
    # print(f"Received parameters: start_date={start_date}, end_date={end_date}, shift={shift}")

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date"}), 400

    start_datetime = f"{start_date} 00:00:00"
    end_datetime = f"{end_date} 23:59:59"

    query = """
        SELECT DATE(`timestamp`) AS date, 
            HOUR(`timestamp`) AS hour,
            MAX(DZ1_Actual) AS DZ1_Actual,
            MAX(DZ1_Set) AS DZ1_Set,
            MAX(DZ2_Actual) AS DZ2_Actual,
            MAX(DZ2_Set) AS DZ2_Set,
            model as model,
            part_name as part_name
        FROM L8Ext3
        WHERE `timestamp` BETWEEN %s AND %s
    """
    
    if shift:
        query += " AND (Shift = %s OR Shift = 'NA')"
    
    query += "GROUP BY DATE(`timestamp`), HOUR(`timestamp`), model, part_name ORDER BY DATE(`timestamp`), HOUR(`timestamp`), model, part_name"
    
    params = [start_datetime, end_datetime]
    if shift:
        params.append(shift)
    
    try:
        cur = mysql.connection.cursor()
        cur.execute(query, params)
        rows = cur.fetchall()
        cur.close()
        
        # Convert the fetched data to a list of dictionaries
        data = []
        for row in rows:
            data.append({
                "hour": f"{row[0]} {row[1]:02d}:00:00",
                "DZ1_Actual": row[2],
                "DZ1_Set": row[3],
                "DZ2_Actual": row[4],
                "DZ2_Set": row[5],
                "Model": row[6],
                "Part Name": row[7]
            })
        return jsonify(data)
    
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/ext3_trend', methods=['GET'])
def get_ext3_trend():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    shift = request.args.get('shift')
    # print(f"Received parameters: start_date={start_date}, end_date={end_date}, shift={shift}")

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date"}), 400

    start_datetime = f"{start_date} 00:00:00"
    end_datetime = f"{end_date} 23:59:59"

    query = """
        SELECT DATE(`timestamp`) AS date, 
            HOUR(`timestamp`) AS hour,
            MAX(BZ1_Set) AS BZ1_Set,
            MAX(BZ1_Actual) AS BZ1_Actual,
            MAX(BZ2_Actual) AS BZ2_Actual,
            MAX(BZ2_Set) AS BZ2_Set,
            MAX(BZ3_Actual) AS BZ3_Actual,
            MAX(BZ3_Set) AS BZ3_Set,
            MAX(BZ4_Actual) AS BZ4_Actual,
            MAX(BZ4_Set) AS BZ4_Set,
            MAX(DZ1_Actual) AS DZ1_Actual,
            MAX(DZ1_Set) AS DZ1_Set,
            MAX(DZ2_Actual) AS DZ2_Actual,
            MAX(DZ2_Set) AS DZ2_Set,
            model as model,
            part_name as part_name,
            MAX(factor) as factor,
            MAX(Main_Motor_Set) as Main_Motor_Set,
            MAX(Main_Motor_Actual) as Main_Motor_Actual
        FROM L8Ext3
        WHERE `timestamp` BETWEEN %s AND %s
    """
    
    if shift:
        query += " AND (Shift = %s OR Shift = 'NA')"
    
    query += "GROUP BY DATE(`timestamp`), HOUR(`timestamp`), model, part_name ORDER BY DATE(`timestamp`), HOUR(`timestamp`), model, part_name"
    
    params = [start_datetime, end_datetime]
    if shift:
        params.append(shift)
    
    try:
        cur = mysql.connection.cursor()
        cur.execute(query, params)
        rows = cur.fetchall()
        cur.close()
        
        # Convert the fetched data to a list of dictionaries
        data = []
        for row in rows:
            data.append({
                "hour": f"{row[0]} {row[1]:02d}:00:00",
                "BZ1_Actual": row[2],
                "BZ1_Set": row[3],
                "BZ2_Set": row[4],
                "BZ2_Actual": row[5],
                "BZ3_Actual": row[6],
                "BZ3_Set": row[7],
                "BZ4_Actual": row[8],
                "BZ4_Set": row[9],
                "DZ1_Actual": row[11],
                "DZ1_Set": row[10],
                "DZ2_Actual": row[13],
                "DZ2_Set": row[12],
                "Model": row[13],
                "Part Name": row[14],
                "Factor": row[15],
                "Main_Motor_Actual": row[16],
                "Main_Motor_Set": row[17],
            })
        return jsonify(data)
    
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": "Internal server error"}), 500
    
@app.route('/api/ext4_bz_trend', methods=['GET'])
def get_ext4_bz_trend():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    shift = request.args.get('shift')
    # print(f"Received parameters: start_date={start_date}, end_date={end_date}, shift={shift}")

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date"}), 400

    start_datetime = f"{start_date} 00:00:00"
    end_datetime = f"{end_date} 23:59:59"

    query = """
        SELECT DATE(`timestamp`) AS date, 
            HOUR(`timestamp`) AS hour,
            MAX(BZ1_Set) AS BZ1_Set,
            MAX(BZ1_Actual) AS BZ1_Actual,
            MAX(BZ2_Actual) AS BZ2_Actual,
            MAX(BZ2_Set) AS BZ2_Set,
            MAX(BZ3_Actual) AS BZ3_Actual,
            MAX(BZ3_Set) AS BZ3_Set,
            MAX(BZ4_Actual) AS BZ4_Actual,
            MAX(BZ4_Set) AS BZ4_Set,
            model as model,
            part_name as part_name
        FROM L8Ext4
        WHERE `timestamp` BETWEEN %s AND %s
    """
    
    if shift:
        query += " AND (Shift = %s OR Shift = 'NA')"
    
    query += "GROUP BY DATE(`timestamp`), HOUR(`timestamp`), model, part_name ORDER BY DATE(`timestamp`), HOUR(`timestamp`), model, part_name"
    
    params = [start_datetime, end_datetime]
    if shift:
        params.append(shift)
    
    try:
        cur = mysql.connection.cursor()
        cur.execute(query, params)
        rows = cur.fetchall()
        cur.close()
        
        # Convert the fetched data to a list of dictionaries
        data = []
        for row in rows:
            data.append({
                "hour": f"{row[0]} {row[1]:02d}:00:00",
                "BZ1_Actual": row[2],
                "BZ1_Set": row[3],
                "BZ2_Set": row[4],
                "BZ2_Actual": row[5],
                "BZ3_Actual": row[6],
                "BZ3_Set": row[7],
                "BZ4_Actual": row[8],
                "BZ4_Set": row[9],
                "Model": row[10],
                "Part Name": row[11]
            })
        return jsonify(data)
    
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": "Internal server error"}), 500
    
@app.route('/api/ext4_dz_trend', methods=['GET'])
def get_ext4_dz_trend():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    shift = request.args.get('shift')
    # print(f"Received parameters: start_date={start_date}, end_date={end_date}, shift={shift}")

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date"}), 400

    start_datetime = f"{start_date} 00:00:00"
    end_datetime = f"{end_date} 23:59:59"

    query = """
        SELECT DATE(`timestamp`) AS date, 
            HOUR(`timestamp`) AS hour,
            MAX(DZ1_Actual) AS DZ1_Actual,
            MAX(DZ1_Set) AS DZ1_Set,
            MAX(DZ2_Actual) AS DZ2_Actual,
            MAX(DZ2_Set) AS DZ2_Set,
            MAX(DZ3_Actual) AS DZ3_Actual,
            MAX(DZ3_Set) AS DZ3_Set,
            model as model,
            part_name as part_name
        FROM L8Ext4
        WHERE `timestamp` BETWEEN %s AND %s
    """
    
    if shift:
        query += " AND (Shift = %s OR Shift = 'NA')"
    
    query += "GROUP BY DATE(`timestamp`), HOUR(`timestamp`), model, part_name ORDER BY DATE(`timestamp`), HOUR(`timestamp`), model, part_name"
    
    params = [start_datetime, end_datetime]
    if shift:
        params.append(shift)
    
    try:
        cur = mysql.connection.cursor()
        cur.execute(query, params)
        rows = cur.fetchall()
        cur.close()
        
        # Convert the fetched data to a list of dictionaries
        data = []
        for row in rows:
            data.append({
                "hour": f"{row[0]} {row[1]:02d}:00:00",
                "DZ1_Actual": row[2],
                "DZ1_Set": row[3],
                "DZ2_Actual": row[4],
                "DZ2_Set": row[5],
                "DZ3_Actual": row[6],
                "DZ3_Set": row[7],
                "Model": row[8],
                "Part Name": row[9]
            })
        return jsonify(data)
    
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/ext4_trend', methods=['GET'])
def get_ext4_trend():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    shift = request.args.get('shift')
    # print(f"Received parameters: start_date={start_date}, end_date={end_date}, shift={shift}")

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date"}), 400

    start_datetime = f"{start_date} 00:00:00"
    end_datetime = f"{end_date} 23:59:59"

    query = """
        SELECT DATE(`timestamp`) AS date, 
            HOUR(`timestamp`) AS hour,
            MAX(BZ1_Set) AS BZ1_Set,
            MAX(BZ1_Actual) AS BZ1_Actual,
            MAX(BZ2_Actual) AS BZ2_Actual,
            MAX(BZ2_Set) AS BZ2_Set,
            MAX(BZ3_Actual) AS BZ3_Actual,
            MAX(BZ3_Set) AS BZ3_Set,
            MAX(BZ4_Actual) AS BZ4_Actual,
            MAX(BZ4_Set) AS BZ4_Set,
            MAX(DZ1_Actual) AS DZ1_Actual,
            MAX(DZ1_Set) AS DZ1_Set,
            MAX(DZ2_Actual) AS DZ2_Actual,
            MAX(DZ2_Set) AS DZ2_Set,
            MAX(DZ3_Actual) AS DZ3_Actual,
            MAX(DZ3_Set) AS DZ3_Set,
            model as model,
            part_name as part_name,
            MAX(factor) as factor,
            MAX(Main_Motor_Set) as Main_Motor_Set,
            MAX(Main_Motor_Actual) as Main_Motor_Actual
        FROM L8Ext4
        WHERE `timestamp` BETWEEN %s AND %s
    """
    
    if shift:
        query += " AND (Shift = %s OR Shift = 'NA')"
    
    query += "GROUP BY DATE(`timestamp`), HOUR(`timestamp`), model, part_name ORDER BY DATE(`timestamp`), HOUR(`timestamp`), model, part_name"
    
    params = [start_datetime, end_datetime]
    if shift:
        params.append(shift)
    
    try:
        cur = mysql.connection.cursor()
        cur.execute(query, params)
        rows = cur.fetchall()
        cur.close()
        
        # Convert the fetched data to a list of dictionaries
        data = []
        for row in rows:
            data.append({
                "hour": f"{row[0]} {row[1]:02d}:00:00",
                "BZ1_Actual": row[2],
                "BZ1_Set": row[3],
                "BZ2_Set": row[4],
                "BZ2_Actual": row[5],
                "BZ3_Actual": row[6],
                "BZ3_Set": row[7],
                "BZ4_Actual": row[8],
                "BZ4_Set": row[9],
                "DZ1_Actual": row[11],
                "DZ1_Set": row[10],
                "DZ2_Actual": row[13],
                "DZ2_Set": row[12],
                "DZ3_Actual": row[15],
                "DZ3_Set": row[14],
                "Model": row[16],
                "Part Name": row[17],
                "Factor": row[18],
                "Main_Motor_Actual": row[19],
                "Main_Motor_Set": row[20],
            })
        return jsonify(data)
    
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/line_chart', methods=['GET'])
def get_line_chart():
    # Retrieve start_date and end_date from the query parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    shift = request.args.get('shift')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date"}), 400

    start_datetime = f"{start_date} 00:00:00"
    end_datetime = f"{end_date} 23:59:59"


    cur = mysql.connection.cursor()
    query = """
        SELECT 
            DATE(timestamp) as date,
            HOUR(timestamp) as hour,
            model as model,
            part_name as part_name,
            MAX(availability) as max_availability, 
            MAX(performance) as max_performance, 
            MAX(quality_rate) as max_quality_rate,
            MAX(oee) as max_oee  
        FROM L8_Running_Status
        WHERE timestamp BETWEEN %s AND %s
    """
    params = [start_datetime, end_datetime]

    if shift:
        # Convert the comma-separated string into a list
        shift_list = shift.split(',')
        # Dynamically create placeholders for the IN clause
        placeholders = ', '.join(['%s'] * len(shift_list))
        query += f" AND (Shift IN ({placeholders}) OR Shift = 'NA')"
        params.extend(shift_list)
    
    query += "GROUP BY DATE(timestamp), HOUR(timestamp), model, part_name ORDER BY DATE(timestamp), HOUR(timestamp), model, part_name"
    

    cur.execute(query, params)
    data = cur.fetchall()
    cur.close()


    result = []
    for row in data:
        result.append({
            "hour": f"{row[0]} {row[1]:02d}:00:00",
            "max_availability": row[4],
            "max_performance": row[5],
            "max_quality_rate": row[6],
            "max_oee": row[7]
        })

    return jsonify(result)

@app.route("/api/seq_plans", methods=["GET"])
def get_plans():
    selected_date = request.args.get("date")
    if not selected_date:
        return jsonify({"error": "Date parameter is required"}), 400

    cursor = mysql.connection.cursor()
    cursor.execute("""
        SELECT id, date, time, partno, model, area, schedule_time, status, active_time, co_end, complete_time
        FROM planning
        WHERE date = %s
        ORDER BY time ASC
    """, (selected_date,))
    rows = cursor.fetchall()
    cursor.close()

    plans = []
    for idx, row in enumerate(rows, start=1):
        plan_date = row[1]
        plan_time = row[2]
        complete_time = row[10]

        if isinstance(plan_time, timedelta):
            plan_time = (datetime.min + plan_time).time()

        # Scheduled start datetime
        start_dt = datetime.combine(plan_date, plan_time)
        start_dt += timedelta(hours=5, minutes=30)  # adjust for timezone

        # Changeover: difference between co_end and active_time
        active_time = row[8]
        co_end = row[9]
        changeover_duration = None
        if active_time and co_end:
            changeover_duration = (co_end - active_time).total_seconds() / 60.0  # minutes

        plans.append({
            "sequence_no": row[0],
            "start": start_dt.strftime("%Y-%m-%dT%H:%M:%S"),
            "duration": row[6],  # minutes
            "partno": row[3],
            "model": row[4],
            "area": row[5],
            "status": row[7],
            "complete_time": complete_time.strftime("%Y-%m-%dT%H:%M:%S") if complete_time else None,
            "active_time": active_time.strftime("%Y-%m-%dT%H:%M:%S") if active_time else None,
            "co_end": co_end.strftime("%Y-%m-%dT%H:%M:%S") if co_end else None,
            "changeover_duration": changeover_duration  # in minutes
        })

    return jsonify(plans)

@app.route('/api/update_rej', methods=['POST'])
def update_rej():
    data = request.json
    plan_id = data.get('plan_id')
    startup_rej = data.get('startup_rej')
    line_rej = data.get('line_rej')

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        updates = []
        values = []

        if startup_rej:
            startup_rej = float(startup_rej)
            updates.append("startup_rej=%s")
            values.append(startup_rej)
        if line_rej:
            line_rej = float(line_rej)
            updates.append("line_rej=%s")
            values.append(line_rej)

        if updates:
            query = f"update L8_Running_Status set {','.join(updates)} where plan_id = %s"
            values.append(plan_id)
            cursor.execute(query, tuple(values))
            conn.commit()
        
        # update_query = """
        #    UPDATE L8_Running_Status 
        #    SET startup_rej = %s
        #    WHERE plan_id = %s """
        # cursor.execute(update_query, (startup_rej, plan_id))
        # update_query2 = """
        #    UPDATE L8_Running_Status 
        #    SET line_rej = %s
        #    WHERE plan_id = %s """
        # cursor.execute(update_query2, (line_rej, plan_id))
        # conn.commit()
        
    except ValueError as e:
        print(f"ValueError: {e}")
        return jsonify({"error": str(e)}), 400
    except Error as e:
        print(f"Error: {e}")
        return jsonify({"error": "An error occurred while updating the plan"}), 500
    finally:
        cursor.close()
        conn.close()

    return jsonify({"message": "Plan updated successfully"})

@app.route('/api/delete_plan', methods=['POST'])
def delete_plan():
    data = request.get_json()  # Parse the JSON body
    plan_id = data.get('plan_id')  # Extract plan_id from the body
    if not plan_id:
        return {"error": "plan_id is required"}, 400

    connection = get_db_connection()
    cursor = connection.cursor()
    
    try:
        query = "DELETE FROM planning WHERE id = %s"
        cursor.execute(query, (plan_id,))
        connection.commit()
        return {"message": "Plan deleted successfully"}, 200
    except Exception as e:
        connection.rollback()
        return {"error": str(e)}, 500
    finally:
        cursor.close()
        connection.close()

@app.route('/api/bitupdate', methods=['POST', 'GET'])
def bitupdate():
    HMI_IP = '163.125.103.169'
    HMI_PORT = 502
    REGISTER_ADDRESS = 50  # Register address

    # Initialize the Modbus client
    client = ModbusTcpClient(HMI_IP, port=HMI_PORT)
    if not client.connect():
        return jsonify({"error": "Unable to connect to Modbus server"}), 500

    # Read the current value from the register
    read_response = client.read_holding_registers(REGISTER_ADDRESS, count=1)
    if read_response.isError():
        client.close()
        return jsonify({"error": f"Failed to read from register {REGISTER_ADDRESS}"}), 500
    else:
        current_value = read_response.registers[0]
        print(f"Current value of register {REGISTER_ADDRESS}: {current_value}")

    # Set the specific bit (e.g., M2 bit) to True (1)
    M2_BIT_POSITION = 2  # M2 is typically the 3rd bit (0-indexed)
    
    # updated_value = current_value | (1 << M2_BIT_POSITION)
    first_updated_value = 1
    write_response = client.write_register(REGISTER_ADDRESS, first_updated_value)
    if write_response.isError():
        client.close()
        return jsonify({"error": f"Failed to write first value to register {REGISTER_ADDRESS}"}), 500
    else:
        print(f"Successfully updated register {REGISTER_ADDRESS} with first value: {first_updated_value}")

    # Wait for 2 seconds before sending the second value
    time.sleep(2)

    # Send second updated value (0)
    second_updated_value = 0
    write_response = client.write_register(REGISTER_ADDRESS, second_updated_value)
    if write_response.isError():
        client.close()
        return jsonify({"error": f"Failed to write second value to register {REGISTER_ADDRESS}"}), 500
    else:
        print(f"Successfully updated register {REGISTER_ADDRESS} with second value: {second_updated_value}")

    client.close()

    return jsonify({
        "status": "success",
        "register_address": REGISTER_ADDRESS,
        "updated_value": first_updated_value,
        "message": "Successfully set M2 bit to True"
    }), 200

@app.route('/api/downtime_status', methods=['GET'])
def get_latest_downtime():
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        query = """
            SELECT status, start_time, end_time
            FROM downtime 
            WHERE status = 'active'
            ORDER BY id DESC LIMIT 1;
        """
        cursor.execute(query)
        result = cursor.fetchone()
        cursor.close()

        if not result:
            return jsonify({"message": "No active downtime", "downtime_active": False}), 200

        # Extract start and stop times for all 5 downtime slots
        start_times = result['start_time']
        stop_times = result['end_time']

        # Find the latest downtime that has not stopped (stop_time is NULL)
        latest_downtime = None
        if start_times and not stop_times:  # Ongoing downtime
            latest_downtime = start_times

        if not latest_downtime:
            return jsonify({"message": "No ongoing downtime", "downtime_active": False}), 200

        # Calculate elapsed time
        current_time = datetime.now()
        elapsed_time = current_time - latest_downtime
        elapsed_hours, remainder = divmod(elapsed_time.seconds, 3600)
        elapsed_minutes, elapsed_seconds = divmod(remainder, 60)

        downtime_since = f"{elapsed_hours}h {elapsed_minutes}m {elapsed_seconds}s"

        return jsonify({"downtime_active": True, "downtime_since": downtime_since}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
@app.route('/api/bulk_plan', methods=['POST'])
def bulk_plan():
    if 'file' not in request.files:
        return jsonify({'message': 'No file part'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'message': 'No selected file'}), 400

    try:
        # Read Excel file from memory
        df = pd.read_excel(file)

        required_columns = [
            'shift', 'date', 'time', 'model', 'partna', 'partno',
            'target_qty', 'plan_qty', 'operator_name', 'supervisor_name'
        ]

        # Filter DataFrame to include only the necessary columns
        # df = df[required_columns]
        df = df.dropna(subset=required_columns)

        # Debug: Print columns
        print("Excel columns:", df.columns.tolist())

        # Connect to DB
        connection = get_db_connection()
        cursor = connection.cursor()

        # Insert each row into the planning table
        for _, row in df.iterrows():
            shift = row['shift'] if pd.notna(row['shift']) else None
            date = row['date'] if pd.notna(row['date']) else None
            time = row['time'] if pd.notna(row['time']) else None
            model = row['model'] if pd.notna(row['model']) else None
            partna = row['partna'] if pd.notna(row['partna']) else None
            partno = row['partno'] if pd.notna(row['partno']) else None
            target_qty = int(row['target_qty']) if pd.notna(row['target_qty']) else 0
            plan_qty = int(row['plan_qty']) if pd.notna(row['plan_qty']) else 0
            operator_name = row['operator_name'] if pd.notna(row['operator_name']) else None
            supervisor_name = row['supervisor_name'] if pd.notna(row['supervisor_name']) else None
            status = 'Inactive'  # Default status, or get from row if available
            cursor.execute("SELECT cycle_time FROM target_sheet WHERE partna = %s AND model = %s", (partna, model,))
            result = cursor.fetchone()

            if result:
                cycle_time = int(result['cycle_time'])
                schedule_time = cycle_time * target_qty / 60  # target_qty must be defined earlier
                print("Schedule Time:", schedule_time)
            else:
                print("No data found for given partna and model")
            cursor.execute("""
                INSERT INTO planning (
                    shift, date, time, model, partna, partno,
                    target_qty, plan_qty, operator_name,
                    supervisor_name, status, schedule_time
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                shift, date, time, model, partna, partno,
                target_qty, plan_qty, operator_name,
                supervisor_name, status, schedule_time
            ))

        connection.commit()
        cursor.close()
        connection.close()

        return jsonify({'message': 'Planning data uploaded successfully'}), 200

    except Exception as e:
        print("Error occurred:", str(e))
        return jsonify({'message': f'Error: {str(e)}'}), 500
    
@app.route('/api/target_sheet', methods=['GET'])
def fetch_target_sheet():
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        cursor.execute("SELECT * FROM target_sheet")
        data = cursor.fetchall()
        cursor.close()
        return jsonify(data), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
@app.route('/api/four_m_table', methods=['GET'])
def fetch_4M_sheet():
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        cursor.execute("SELECT * FROM four_m_log")
        data = cursor.fetchall()
        cursor.close()
        return jsonify(data), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
@app.route('/api/site_map', methods=['GET'])
def fetch_site_map():
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        cursor.execute("SELECT * FROM site_map")
        data = cursor.fetchall()
        cursor.close()
        return jsonify(data), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
@app.route('/api/digital_twin', methods=['GET'])
def fetch_digital_twin():
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        cursor.execute("SELECT * FROM login_new")
        data = cursor.fetchall()
        cursor.close()
        return jsonify(data), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/add_site', methods= ['POST'])
def add_site():
    try:
        data = request.json
        # ic(data)
        button = data['button']
        Add = Production.Add_site(button, data)
        if Add:
            return jsonify({"message": "Success"}), 200
        else:
            return jsonify({"error": "Invalid operation or missing fields"}), 400
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500

@app.route('/api/add_part', methods= ['POST'])
def add_part():
    try:
        data = request.json
        ic(data)
        button = data['button']
        Add = Production.Add_part(button, data)
        if Add:
            return jsonify({"message": "Success"}), 200
        else:
            return jsonify({"error": "Invalid operation or missing fields"}), 400
    except Exception as exc:
        traceback.print_exc()
        return jsonify({'error': str(exc)}), 500

@app.route('/api/approve_4m', methods= ['POST'])
def approve_4m():
    try:
        data = request.json
        Add = Production.approve_4m(data)
        if Add:
            return jsonify({"message": "Success"}), 200
        else:
            return jsonify({"error": "Invalid operation or missing fields"}), 400
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500

@app.route('/api/reject_4m', methods= ['POST'])
def reject_4m():
    try:
        data = request.json
        Add = Production.reject_4m(data)
        if Add:
            return jsonify({"message": "Success"}), 200
        else:
            return jsonify({"error": "Invalid operation or missing fields"}), 400
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500
    
@app.route('/api/crud_user', methods= ['POST'])
def crud_user():
    try:
        data = request.json
        ic(data)
        button = data['button']
        Add = Production.crud_user(button, data)
        if Add:
            return jsonify({"message": "Success"}), 200
        else:
            return jsonify({"error": "Invalid operation or missing fields"}), 400
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500
    
@app.route('/api/workflow', methods=['GET'])
def fetch_workflow():
    connection = get_db_connection()
    cursor = connection.cursor()  # Get data as dicts
    try:
        cursor.execute("SELECT * FROM workflow")
        rows = cursor.fetchall()
        cursor.close()

        # Convert datetime objects to string
        for row in rows:
            if isinstance(row['created_at'], datetime):
                row['created_at'] = row['created_at'].strftime("%Y-%m-%d %H:%M:%S")

        return jsonify(rows), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
@app.route('/api/workflow_qr', methods=['POST'])
def generate_qr():
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data received"}), 400

    # Extract workflow id, fallback to 'unknown' if missing
    workflow_id = data.get('id', 'unknown')

    # Create the URL to encode in the QR code
    qr_url = f"http://163.125.102.142:3500/fill?id={workflow_id}"

    # Generate QR code for the URL string
    img = qrcode.make(qr_url)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)

    # Send the QR code image as a file attachment, filename based on workflow id
    return send_file(
        buf,
        mimetype='image/png',
        as_attachment=True,
        download_name=f"workflow_{workflow_id}_qr.png"
    )

@app.route('/api/add_workflow', methods= ['POST'])
def add_workflow():
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        data = request.json
        ic(data)
        button = data['button']
        Add = Production.add_workflow(button, data)
        if Add:
            return jsonify({"message": "Success"}), 200
        else:
            return jsonify({"error": "Invalid operation or missing fields"}), 400
    except Exception as exc:
        traceback.print_exc()  # Log full traceback
        return jsonify({'error': str(exc)}), 500
    
@app.route('/api/forms', methods= ['GET'])
def forms():
    try:
        workflow_id = request.args.get('id')
        questions = Production.get_questions(workflow_id)
        # ic(questions)
        return jsonify(questions), 200
    except Exception as exc:
        traceback.print_exc()  # Log full traceback
        return jsonify({'error': str(exc)}), 500
    
@app.route('/api/notifications', methods=['GET'])
def get_notifications():
    try:
        user_id = request.args.get('user_id')
        ic(user_id)
        if not user_id:
            raise ValueError("Missing user_id parameter")

        return Production.get_user_notifications(user_id)  
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
@app.route('/api/notifications/<int:notification_id>/read', methods=['PATCH'])
def mark_notification_as_read(notification_id):
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        sql = "UPDATE notifications SET is_read = True WHERE id = %s"
        cursor.execute(sql, (notification_id,))
        connection.commit()
        return jsonify({"message": "Notification marked as read"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        connection.close()
    
@app.route('/api/downtime_logs', methods=['GET'])
def get_downtime_log():
    try:
        data = Production.get_downtime_log()
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
@app.route('/api/downtime_update', methods=['POST'])
def update_downtime():
    try:
        data = request.get_json()
        id = data.get('id')
        department = data.get('department')
        ic(department, id)
        data = Production.update_downtime_log(department, id)
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
@app.route('/api/get_breakdown', methods=['GET', 'POST'])
def get_breakdown():
    try:
        department = request.json.get("department")
        # ic(department)
        bd_data = Production.get_breakdown(department)
        return jsonify(bd_data)

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
@app.route('/api/downtime_reason', methods=['POST'])
def update_downtime_reason():
    try:
        data = request.get_json()
        id = data.get('id')
        reason = data.get('reason')
        ic(reason, id)
        data = Production.update_downtime_reason(reason, id)
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/getapi', methods=['GET'])
def getapi():
    year_range = request.args.get('year')
    ic(year_range)
    try:
        token = Production.get_token()
        ic(token)
        attendance = Production.fetch_attendance(token, "2025-08-01", "2025-08-20")
        print(attendance)

        return jsonify(attendance), 200

    except Exception as e:
        print("Error in get_annual_plan_excel_data:", str(e))
        return jsonify({"error": str(e)}), 500

@app.route('/api/change_options', methods= ['GET'])
def get_change_options():
    options = {
        "Operator on leave with information" : {
            "type_of_change": "Planned",
            "action": ["Deploy same skill team member at change work area", 
                       "Operator of same skill deputed if the operator has been deputed on the station after long time(eg Afetr 3 Months)",
                       "Give refresher training to team member on process control parameters and check points referring SOP/SIP",
                       "Verify part after operation and give guidance to team member",
                       "Insert change flag in respective work area in 4M change control sheet for communication and necessary action to all",
                       "In case of Existing manpower is not available, Provide proper training of process and key point befor deploying back up member/ Level2 team members at work station and evaluate their work periodically."
                       ],
            "abnormal_situation": "No",
            "setup_approval": "Yes",
            "retroactive_inspection": "No",
            "suspected_lot": "No"
        },
        "Extended / extra duty" : {
            "type_of_change": "Planned",
            "action": ["4 Hour Extra duty deploy same skill member at work area.", 
                       "8 Hr. Extra duty deploy same skill member at work area. & inspection frequency increase from 1/hr to 1/half hour for aesthatic defects"
                       ],
            "abnormal_situation": "No",
            "setup_approval": "Yes",
            "retroactive_inspection": "No",
            "suspected_lot": "Yes"
        },
        "New operator after DOJO training" : {
            "type_of_change": "Planned",
            "action": ["Depute as per skill matrix chart"
                       ],
            "abnormal_situation": "No",
            "setup_approval": "Yes",
            "retroactive_inspection": "No",
            "suspected_lot": "Yes"
        },
        "Sudden change in operator" : {
            "type_of_change": "Unplanned",
            "action": ["Deploy same skill team member at change work areas"
                       ],
            "abnormal_situation": "No",
            "setup_approval": "Yes",
            "retroactive_inspection": "No",
            "suspected_lot": "No"
        },
        "Operator suddenly leave work place due to accident/illness (after start of production)" : {
            "type_of_change": "Unplanned",
            "action": ["Operator of same skill deputed who has been working on the station frequently."
                       ],
            "abnormal_situation": "No",
            "setup_approval": "Yes",
            "retroactive_inspection": "No",
            "suspected_lot": "No"
        },
        "Operator on leave without information" : {
            "type_of_change": "Unplanned",
            "action": ["Operator of same skill deputed if the operator has been deputed on the station after long time(eg After 3 Months.)",
                       "Give refresher training to team member on process control parameters and check points referring SOP/SIP",
                       "Verify part after operation and give guidance to team member",
                       "Insert change flag in respective work area in 4M change control sheet for communication and necessary action to all"
                       ],
            "abnormal_situation": "No",
            "setup_approval": "Yes",
            "retroactive_inspection": "Yes",
            "suspected_lot": "Yes"
        },
        "Operator injured" : {
            "type_of_change": "Unplanned",
            "action": ["Stop production and deputed same skilled operator"
                       ],
            "abnormal_situation": "No",
            "setup_approval": "Yes",
            "retroactive_inspection": "Yes",
            "suspected_lot": "Yes"
        },
        "Non-conformance during operator observance" : {
            "type_of_change": "Unplanned",
            "action": ["Stop production and deputed same skilled operator"
                       ],
            "abnormal_situation": "No",
            "setup_approval": "Yes",
            "retroactive_inspection": "Yes",
            "suspected_lot": "Yes"
        },
        "During job rotation and multiskilling" : {
            "type_of_change": "Planned",
            "action": ["Impart training on Procedure, Process parameters, control items and check items by referring SOP/SIP,Jishu Hosen check sheet and 5'S check sheet",
                       "Verification of output",
                       "First and last piece ok",
                       "Hourly inspection",
                       "Final product audit",
                       "Insert change flag in respective work area in 4M change control sheet for communication and necessary action to all.",
                       "Record the change in 4M Tracking sheet",
                       "Evaluate skill level"
                       ],
            "abnormal_situation": "No",
            "setup_approval": "Yes",
            "retroactive_inspection": "Yes",
            "suspected_lot": "Yes"
        },
        "During GL/ TL change in the team" : {
            "type_of_change": "Planned",
            "action": ["Ensure availability of check item list to new GL/ TL"
                       ],
            "abnormal_situation": "No",
            "setup_approval": "Yes",
            "retroactive_inspection": "NA",
            "suspected_lot": "NA"
        },
        "Deployment of underskill manpower in above mentioned Man changes" : {
            "type_of_change": ["Planned","Unplanned"],
            "action": ["Increase inspection frequency from 1/Hour to 2/Hour"
                       ],
            "abnormal_situation": "Yes",
            "setup_approval": "Yes",
            "retroactive_inspection": ["Yes","No"],
            "suspected_lot": "Yes"
        },
    }
    return jsonify(options)

@app.route('/api/change_options_mat', methods= ['GET'])
def get_change_options_mat():
    options = {
        "Introduction of new grade / supplier" : {
            "type_of_change": "Planned",
            "action": ["Plan trial with PE", 
                       "Get the samples approved from laboratory for all performance testing as per drawing",
                       "Submit samples and test reports to customer for approval",
                       "Insert change flag in respective work area in 4m change control sheet for communication and necessary action to all",
                       "Record in 4M change and take approval from Department head or Quality head / Plant Head",
                       "Record the change",
                       ],
            "abnormal_situation": "No",
            "setup_approval": "Yes",
            "retroactive_inspection": "No",
            "suspected_lot": "No"
        },
        "Material received from other approved source already under regular consumption" : {
            "type_of_change": "Planned",
            "action": ["NA"],
            "abnormal_situation": "No",
            "setup_approval": "No",
            "retroactive_inspection": "No",
            "suspected_lot": "No"
        },
        "Material received from other approved souce not used earlier" : {
            "type_of_change": "Planned",
            "action": ["Start production after setup approval"
                       ],
            "abnormal_situation": "No",
            "setup_approval": "Yes",
            "retroactive_inspection": "No",
            "suspected_lot": "No"
        },
        "Material received from other unapproved source without intimation" : {
            "type_of_change": "Unplanned",
            "action": ["Start production after taking approval from quality",
                       "Check last part produced with the material"
                       ],
            "abnormal_situation": "No",
            "setup_approval": "Yes",
            "retroactive_inspection": "Yes",
            "suspected_lot": "Yes"
        },
        "Continous non conformance in product" : {
            "type_of_change": "Unplanned",
            "action": ["stop the M/C",
                       "Reset the process After correction of abnormility",
                       "Take setup approval",
                       "Reject the NG parts",
                       "Start production with OK Parts"
                       ],
            "abnormal_situation": "No",
            "setup_approval": "Yes",
            "retroactive_inspection": "No",
            "suspected_lot": "No"
        },
        "Operator on leave without information" : {
            "type_of_change": "Unplanned",
            "action": ["Operator of same skill deputed if the operator has been deputed on the station after long time(eg After 3 Months.)",
                       "Give refresher training to team member on process control parameters and check points referring SOP/SIP",
                       "Verify part after operation and give guidance to team member",
                       "Insert change flag in respective work area in 4M change control sheet for communication and necessary action to all"
                       ],
            "abnormal_situation": "No",
            "setup_approval": "Yes",
            "retroactive_inspection": "Yes",
            "suspected_lot": "Yes"
        },
        "Operator injured" : {
            "type_of_change": "Unplanned",
            "action": ["Stop production and deputed same skilled operator"
                       ],
            "abnormal_situation": "No",
            "setup_approval": "Yes",
            "retroactive_inspection": "Yes",
            "suspected_lot": "Yes"
        },
        "Non-conformance during operator observance" : {
            "type_of_change": "Unplanned",
            "action": ["Stop production and deputed same skilled operator"
                       ],
            "abnormal_situation": "No",
            "setup_approval": "Yes",
            "retroactive_inspection": "Yes",
            "suspected_lot": "Yes"
        },
        "During job rotation and multiskilling" : {
            "type_of_change": "Planned",
            "action": ["Impart training on Procedure, Process parameters, control items and check items by referring SOP/SIP,Jishu Hosen check sheet and 5'S check sheet",
                       "Verification of output",
                       "First and last piece ok",
                       "Hourly inspection",
                       "Final product audit",
                       "Insert change flag in respective work area in 4M change control sheet for communication and necessary action to all.",
                       "Record the change in 4M Tracking sheet",
                       "Evaluate skill level"
                       ],
            "abnormal_situation": "No",
            "setup_approval": "Yes",
            "retroactive_inspection": "Yes",
            "suspected_lot": "Yes"
        },
        "During GL/ TL change in the team" : {
            "type_of_change": "Planned",
            "action": ["Ensure availability of check item list to new GL/ TL"
                       ],
            "abnormal_situation": "No",
            "setup_approval": "Yes",
            "retroactive_inspection": "NA",
            "suspected_lot": "NA"
        },
        "Deployment of underskill manpower in above mentioned Man changes" : {
            "type_of_change": ["Planned","Unplanned"],
            "action": ["Increase inspection frequency from 1/Hour to 2/Hour"
                       ],
            "abnormal_situation": "Yes",
            "setup_approval": "Yes",
            "retroactive_inspection": ["Yes","No"],
            "suspected_lot": "Yes"
        },
    }
    return jsonify(options)

@app.route('/api/chat', methods=['POST'])
def get_insights():
    try:
        data = request.json
        query_text = data.get("query", "").strip()
        ic(data, query_text)

        if not query_text:
            return jsonify({"error": "Query is empty"}), 400

        # Fetch production data from MySQL
        # conn = get_db_connection()
        # cursor = conn.cursor()
        # cursor.execute("SELECT timestamp, part_name, model, cutter1_count, cutter2_count, startup_rej, line_rej FROM L8_Running_Status LIMIT 100")
        # production_data = cursor.fetchall()
        # conn.close()

        # if not production_data:
        #     return jsonify({"error": "No production data available"}), 404

        # # Format data for AI
        # formatted_data = f"User Query: {query_text}\nProduction Data: {production_data}. Provide insights and improvement suggestions."

        formatted_data = f"User Query: {query_text}\n Table in mysql: L8_Running_Status \n columns in table: timestamp, part_name, model, cutter1_count . Provide mysql query ."

        print("Generating response from Ollama...")
        response = ollama.generate(
            model=data.get("model", "llama3.2-vision:latest"),  # Use model from request or default
            prompt=formatted_data
        )
        print(f"Ollama response: {response}")

        ai_response = response.get("response", "No response generated.")
        return jsonify({"insights": ai_response})

    except mysql.connector.Error as db_err:
        return jsonify({"error": f"Database Error: {str(db_err)}"}), 500
    except requests.exceptions.RequestException as req_err:
        return jsonify({"error": f"Ollama API Error: {str(req_err)}"}), 500
    except Exception as e:
        return jsonify({"error": f"Internal Server Error: {str(e)}"}), 500

MODBUS_IP = "163.125.103.169"
MODBUS_PORT = 502
MODBUS_UNIT = 1

last_states = [None, None, None]  # For inputs 0, 1, 2
last_holding_139 = None           # For holding register 139


def read_inputs():
    client = ModbusTcpClient(MODBUS_IP, port=MODBUS_PORT)
    client.connect()
    rr = client.read_discrete_inputs(0, count=3, slave=MODBUS_UNIT)
    client.close()

    if rr is None or rr.isError():
        return None

    return [{"address": i, "status": int(rr.bits[i])} for i in range(3)]


def read_holding_139():
    client = ModbusTcpClient(MODBUS_IP, port=MODBUS_PORT)
    client.connect()
    rr = client.read_holding_registers(139, count=1, slave=MODBUS_UNIT)
    client.close()

    if rr is None or rr.isError():
        return None

    return rr.registers[0]  # single value


@app.route("/api/pokayoke_status", methods=["GET"])
def poka_yoke_status():
    global last_states, last_holding_139

    # Read discrete inputs
    current_states = read_inputs()
    if current_states is None:
        return jsonify({"error": "Failed to read Modbus inputs"}), 500

    # Read holding register 139
    holding_139_value = read_holding_139()
    if holding_139_value is None:
        return jsonify({"error": "Failed to read holding register 139"}), 500

    # Process discrete inputs 0,1,2
    for i in range(3):
        current_status = current_states[i]["status"]

        if last_states[i] is None:
            last_states[i] = current_status
        elif last_states[i] != current_status:
            nc_present = 1 if (last_states[i] == 0 and current_status == 1) else 0

            cursor = mysql.connection.cursor()
            cursor.execute("""
                INSERT INTO poka_yoke (input_no, bit_status, nc_present, timestamp)
                VALUES (%s, %s, %s, %s)
            """, (i, current_status, nc_present, datetime.now()))
            mysql.connection.commit()
            cursor.close()

            last_states[i] = current_status

    # Process holding register 139 with same logic
    if last_holding_139 is None:
        last_holding_139 = holding_139_value
    elif last_holding_139 != holding_139_value:
        nc_present = 1 if (last_holding_139 == 0 and holding_139_value == 1) else 0

        cursor = mysql.connection.cursor()
        cursor.execute("""
            INSERT INTO poka_yoke (input_no, bit_status, nc_present, timestamp)
            VALUES (%s, %s, %s, %s)
        """, (139, holding_139_value, nc_present, datetime.now()))
        mysql.connection.commit()
        cursor.close()

        last_holding_139 = holding_139_value

    return jsonify({
        "inputs": current_states,
        "holding_139": holding_139_value
    })

@app.route("/api/glboard", methods=["GET"])
def get_glboard():
    try:
        connection = get_db_connection()
        cursor = connection.cursor()

        month_param = request.args.get("month")  # expecting format YYYY-MM
        line_param = request.args.get("line")
        area_param = request.args.get("area")

        sql = """SELECT * FROM gl_board_kpis WHERE 1=1"""
        params = []

        # ✅ Month filter
        if month_param:
            sql += " AND DATE_FORMAT(date, '%%Y-%%m') = %s"
            params.append(month_param)

        # ✅ Line filter
        if line_param:
            sql += " AND (CAST(line AS CHAR) = %s OR line = %s)"
            params.extend([str(line_param), f"Line {line_param}"])
        
        if area_param : 
            sql += " AND plant = %s"
            params.append(area_param)

        sql += " ORDER BY date DESC, created_at DESC LIMIT 200"

        cursor.execute(sql, params)
        data = cursor.fetchall()

        return jsonify(data), 200

    except Exception as ex:
        return jsonify({"error": str(ex)}), 500

    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals():
            connection.close()

@app.route("/api/4mboard", methods=["GET"])
def get_4mboard():
    try:
        connection = get_db_connection()
        cursor = connection.cursor()

        month_param = request.args.get("month")  # expecting format YYYY-MM
        line_param = request.args.get("line")
        area_param = request.args.get("area")

        sql = """SELECT * FROM four_m_change WHERE 1=1"""
        params = []

        # ✅ Month filter
        if month_param:
            sql += " AND DATE_FORMAT(date, '%%Y-%%m') = %s"
            params.append(month_param)

        # ✅ Line filter
        if line_param:
            sql += " AND (CAST(sub_area AS CHAR) = %s OR sub_area = %s)"
            params.extend([str(line_param), f"Line {line_param}"])
        
        if area_param : 
            sql += " AND plant = %s"
            params.append(area_param)

        sql += " ORDER BY date DESC, created_at DESC LIMIT 200"

        cursor.execute(sql, params)
        data = cursor.fetchall()
        ic(data)

        return jsonify(data), 200

    except Exception as ex:
        return jsonify({"error": str(ex)}), 500

    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals():
            connection.close()

@app.route("/api/gl_form2", methods=['POST'])
def gl_form2():
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        data = request.get_json()
        ic(data)
        numeric_fields = [
            'fr_act','rr_act','fr_target','rr_target','fr_wt','rr_wt','fr_len','rr_len',
            'co_time','setup_time','dt_time','duration','startup_rej','line_rej'
        ]


        for field in numeric_fields:
            data[field] = float(data[field]) if data.get(field) else 0
        efficiency = (data['fr_act'] + data['rr_act']) * 100 / (data['fr_target'] + data['rr_target'])
        runtime = data['duration'] - (data['co_time'] + data['setup_time'] + data['dt_time'])
        prod_target = (data['fr_target'] * data['fr_wt'] + data['rr_target'] * data['rr_wt']) * 480 / (runtime * 1000)
        prod_act = (data['fr_act'] * data['fr_wt'] + data['rr_act'] * data['rr_wt']) * 480 / (runtime * 1000)
        cosump = ((data['fr_act'] * data['fr_wt'] + data['rr_act'] * data['rr_wt']) / 1000) + data['startup_rej'] + data['line_rej']
        total_len = (data['fr_len'] * data['fr_act'] + data['rr_len'] * data['rr_act']) / 1000
        gpm = cosump / total_len if total_len else 0
        line_rej_perc = (data['line_rej'] / (data['fr_act'] + data['rr_act'])) * 100 if (data['fr_act'] + data['rr_act']) else 0
        overall_rej_perc = ((data['line_rej'] + data['startup_rej']) / (data['fr_act'] + data['rr_act'])) * 100 if (data['fr_act'] + data['rr_act']) else 0
        indirect_cost = int(data['hand_gloves'])*10 + int(data['hand_gloves'])*10 + int(data['hand_gloves'])*270
        # ic(indirect_cost)
        cursor.execute("""
            INSERT INTO gl_board_kpis 
            (date, line, model, part_name, customer, line_rej, startup_rej, fr_target, rr_target, 
             fr_act, rr_act, fr_wt, rr_wt, fr_len, rr_len, co_time, setup_time, dt_time, 
             duration, run_time, efficiency, prod_target, prod_act, gpm, consumption, 
             line_rej_perc, overall_rej_perc, indirect_cost, old_dhoti, hand_gloves, corrogard, start_reading, end_reading, plant) 
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            data['date'], data['line'], data['model'], data['part_name'],  data['customer'], data['line_rej'], data['startup_rej'],
            data['fr_target'], data['rr_target'], data['fr_act'], data['rr_act'], data['fr_wt'], data['rr_wt'],
            data['fr_len'], data['rr_len'], data['co_time'], data['setup_time'], data['dt_time'], 
            data['duration'], runtime, efficiency, prod_target, prod_act, gpm, cosump,
            line_rej_perc, overall_rej_perc, indirect_cost, data['old_dhoti'], data['hand_gloves'], data['corrogard'], data['start_reading'], data['end_reading'], data['plant']
        ))
        connection.commit()
        cursor.close()
        connection.close()
        return {"status" : "success"},200
    except Exception as E:
        ic(E)
        return {"status" : "fail"},500
    
@app.route("/api/gl_qual_form", methods=['POST'])
def gl_qual_form():
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        data = request.get_json()
        ic(data)

        cursor.execute("""
            INSERT INTO gl_qual 
            (date, plant, area, problem, action, target_date, responsibility, status ) 
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            data['date'], data['plant'], data['line'], data['problem'], data['action'], data['target_date'], data['responsibility'], "Open"
        ))
        connection.commit()
        cursor.close()
        connection.close()
        return {"status" : "success"},200
    except Exception as E:
        ic(E)
        return {"status" : "fail"},500
    
@app.route("/api/four_m_add", methods=["POST"])
def add_four_m_change():
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        data = request.get_json()

        # Debug incoming payload
        print("Received 4M Change Data:", data)

        cursor.execute("""
            INSERT INTO four_m_change 
            (date, shift, plant, area, sub_area, category, detail, bd_reason, bd_time,
             total_qty, action_taken, retro_total, retro_ok, retro_ng,
             suspected_total, suspected_ok, suspected_ng)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            data.get("date"),
            data.get("shift"),
            data.get("plant"),
            data.get("area"),
            data.get("sub_area"),
            data.get("category"),
            data.get("detail"),
            data.get("bd_reason"),
            data.get("bd_time"),
            data.get("total_qty"),
            data.get("action_taken"),
            data.get("retro_total"),
            data.get("retro_ok"),
            data.get("retro_ng"),
            data.get("suspected_total"),
            data.get("suspected_ok"),
            data.get("suspected_ng")
        ))

        connection.commit()
        cursor.close()
        connection.close()
        return {"status": "success"}, 200

    except Exception as e:
        print("Error while inserting 4M Change:", e)
        return {"status": "fail", "error": str(e)}, 500

    
    
@app.route("/api/glsaform", methods=['POST'])
def glsaform():
    try:
        data = request.get_json()
        if not data:
            return {"status": "fail", "error": "No JSON payload"}, 400

        form_type = data.get("formType")
        if form_type == "Accident":
            required_keys = ['date','category','area','sub_area','detail','cause','counter_measure']
        elif form_type == "Safety":
            required_keys = ['date','area','sub_area','detail','responsibility','target_date']
        else:
            return {"status": "fail", "error": "Wrong Submission"}, 400

        # Check for missing keys
        missing_keys = [k for k in required_keys if k not in data]
        if missing_keys:
            return {"status": "fail", "error": f"Missing keys: {missing_keys}"}, 400

        # Use context manager for connection
        with get_db_connection() as connection:
            with connection.cursor() as cursor:
                if form_type == "Accident":
                    cursor.execute("""
                        INSERT INTO gl_safety
                        (date, category, incident_cat, area, sub_area, detail, cause, counter_measure) 
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        data['date'],
                        form_type,
                        data['category'],
                        data['area'],
                        data['sub_area'],
                        data['detail'],
                        data['cause'],
                        data['counter_measure']
                    ))
                elif form_type == "Safety":
                    cursor.execute("""
                        INSERT INTO gl_safety 
                        (date, category, area, sub_area, detail, responsibility, target_date) 
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """, (
                        data['date'],
                        form_type,
                        data['area'],
                        data['sub_area'],
                        data['detail'],
                        data['responsibility'],
                        data['target_date']
                    ))
            connection.commit()  # commit outside the cursor context

        return {"status": "success"}, 200

    except Exception as e:
        print("Error:", e)
        return {"status": "fail", "error": str(e)}, 500

@app.route("/api/gl_safety2", methods=["GET"])
def gl_safety2():
    try:
        connection = get_db_connection()
        cursor = connection.cursor()

        month_param = request.args.get("month")  # expecting format YYYY-MM
        line_param = request.args.get("line")
        area_param = request.args.get("area")
        ic(month_param, line_param, area_param)

        sql = """SELECT * FROM gl_safety WHERE 1=1"""
        params = []

        # ✅ Month filter
        if month_param:
            sql += " AND DATE_FORMAT(date, '%%Y-%%m') = %s"
            params.append(month_param)

        # ✅ Line filter
        if line_param:
            sql += " AND (CAST(sub_area AS CHAR) = %s OR sub_area = %s)"
            params.extend([f"Line {line_param}",str(line_param)])
        
        if area_param : 
            sql += " AND plant = %s"
            params.append(area_param)

        sql += " ORDER BY date DESC, created_at DESC LIMIT 200"

        cursor.execute(sql, params)
        data = cursor.fetchall()
        # ic(data)

        return jsonify(data), 200

    except Exception as ex:
        return jsonify({"error": str(ex)}), 500

    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals():
            connection.close()


# ---- GET route for fetching safety data ----
@app.route("/api/gl_safety", methods=['GET'])
def get_gl_safety():
    try:
        month = request.args.get("month")
        if not month:
            return jsonify({"status": "fail", "error": "month parameter required"}), 400

        connection = get_db_connection()
        cursor = connection.cursor()

        query = """
            SELECT category, COUNT(*) as count
            FROM gl_safety
            WHERE DATE_FORMAT(date, '%%Y-%%m') = %s
            GROUP BY category
        """
        cursor.execute(query, (month,))
        rows = cursor.fetchall()

        cursor.close()
        connection.close()

        return jsonify(rows), 200

    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "fail", "error": str(e)}), 500


# ---- GET route for fetching incident data ----
@app.route("/api/gl_incident", methods=['GET'])
def get_gl_incident():
    try:
        month = request.args.get("month")  # format: YYYY-MM
        if not month:
            return jsonify({"status": "fail", "error": "month parameter required"}), 400

        connection = get_db_connection()
        cursor = connection.cursor()

        query = """
            SELECT incident_cat, COUNT(*) as count
            FROM gl_safety
            WHERE incident_cat IS NOT NULL
            AND DATE_FORMAT(date, '%%Y-%%m') = %s
            GROUP BY incident_cat
        """
        cursor.execute(query, (month,))
        rows = cursor.fetchall()

        cursor.close()
        connection.close()

        return jsonify(rows), 200

    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "fail", "error": str(e)}), 500

@app.route("/api/glboard", methods=["POST"])
def submit_glboard():
    try:
        data = request.get_json()
 
        connection = get_db_connection()
        cursor = connection.cursor()
        now = datetime.now()
        timestamp_str = now.strftime("%Y-%m-%d %H:%M:%S")
 
        sql = """
            INSERT INTO gl_board_kpis (
                date, line, safety, safety_action_plan,
                line_rejection_actual, total_material_consumption, line_rejection_percent,
                startup_scrap, overall_rejection_percent,
                productivity_target, part_production, productivity_efficiency_percent,
                changeover_time, setup_time, breakdown_time,
                abnormality_management,
                cost_energy, cost_direct_material,
                old_dhoti_consumption, hand_gloves, coroguard_ml,created_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s)
        """

        values = (
            data.get("date"),
            data.get("shift"),
            data.get("safety"),
            data.get("safetyActionPlan"),
            data.get("lineRejectionActual"),
            data.get("totalMaterialConsumption"),
            data.get("lineRejectionPercent"),
            data.get("startupScrap"),
            data.get("overallRejectionPercent"),
            data.get("productivityTarget"),
            data.get("partProduction"),
            data.get("productivityEfficiency"),
            data.get("changeoverTime"),
            data.get("setupTime"),
            data.get("breakdownTime"),
            data.get("abnormalityManagement"),
            data.get("costEnergy"),
            data.get("costDirectMaterial"),
            data.get("oldDhoti"),
            data.get("handGloves"),
            data.get("coroguard"),
            timestamp_str
        )
 
        cursor.execute(sql, values)
        connection.commit()
 
        return jsonify({"message": "Data inserted successfully"}), 201
 
    except Error as e:
        print("Database Error:", e)
        return jsonify({"error": str(e)}), 500
 
    except Exception as ex:
        print("Error:", ex)
        return jsonify({"error": str(ex)}), 500
 
    finally:
        if 'cursor' in locals(): cursor.close()
        if 'connection' in locals(): connection.close()

@app.route("/api/points", methods=["GET"])
def get_points():
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
 
        sql = """
            SELECT *
            FROM points
            ORDER BY date DESC
        """
 
        cursor.execute(sql)
        data = cursor.fetchall()
 
        return jsonify(data), 200
 
    except Error as e:
        print("Database Error:", e)
        return jsonify({"error": str(e)}), 500
 
    finally:
        if 'cursor' in locals(): cursor.close()
        if 'connection' in locals(): connection.close()

@app.route("/api/gl_points", methods=["GET"])
def get_gl_points():
    try:
        connection = get_db_connection()
        cursor = connection.cursor()

        # ✅ Get filters from query params
        month = request.args.get("month")  # format: YYYY-MM
        area = request.args.get("area")
        line = request.args.get("line")

        # ✅ Base query
        sql = "SELECT * FROM gl_safety WHERE 1=1"
        params = []

        # ✅ Apply filters dynamically
        if month:
            # Convert month to start and end dates
            start_date = f"{month}-01"
            # Calculate last day of month
            end_date = datetime.strptime(start_date, "%Y-%m-%d").replace(day=28)  # temporary
            end_date = end_date.replace(day=28)  # fallback if month has <31 days
            sql += " AND DATE_FORMAT(date, '%%Y-%%m') = %s"
            params.append(month)

        if area:
            sql += " AND plant = %s"
            params.append(area)

        if line:
            sql += " AND sub_area = %s"
            params.append(line)

        sql += " ORDER BY id DESC"

        cursor.execute(sql, params)
        data = cursor.fetchall()
        # ic(data)

        return jsonify(data), 200

    except Exception as e:
        print("Database Error:", e)
        return jsonify({"error": str(e)}), 500

    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals():
            connection.close()

@app.route("/api/gl_points/<int:id>", methods=["PUT"])
def update_gl_point_status(id):
    try:
        connection = get_db_connection()
        cursor = connection.cursor()

        data = request.get_json()
        status = data.get("status")

        sql = "UPDATE gl_safety SET status=%s WHERE id=%s"
        cursor.execute(sql, (status, id))
        connection.commit()

        return jsonify({"status": "success"}), 200

    except Exception as e:
        print("Update Error:", e)
        return jsonify({"status": "fail", "error": str(e)}), 500

    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals():
            connection.close()

@app.route("/api/gl_qa_points", methods=["GET"])
def get_gl_qa_points():
    try:
        connection = get_db_connection()
        cursor = connection.cursor()

        # ✅ Get filters from query params
        month = request.args.get("month")  # format: YYYY-MM
        area = request.args.get("area")
        line = request.args.get("line")

        # ✅ Base query
        sql = "SELECT * FROM gl_qual WHERE 1=1"
        params = []

        # ✅ Apply filters dynamically
        if month:
            # Convert month to start and end dates
            start_date = f"{month}-01"
            # Calculate last day of month
            end_date = datetime.strptime(start_date, "%Y-%m-%d").replace(day=28)  # temporary
            end_date = end_date.replace(day=28)  # fallback if month has <31 days
            sql += " AND DATE_FORMAT(date, '%%Y-%%m') = %s"
            params.append(month)

        if area:
            sql += " AND plant = %s"
            params.append(area)

        if line:
            sql += " AND area = %s"
            params.append(line)

        sql += " ORDER BY id DESC"

        cursor.execute(sql, params)
        data = cursor.fetchall()
        # ic(data)

        return jsonify(data), 200

    except Exception as e:
        print("Database Error:", e)
        return jsonify({"error": str(e)}), 500

    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals():
            connection.close()

@app.route("/api/gl_4m_points", methods=["GET"])
def get_gl_4m_points():
    try:
        connection = get_db_connection()
        cursor = connection.cursor()

        # ✅ Get filters from query params
        month = request.args.get("month")  # format: YYYY-MM
        area = request.args.get("area")
        line = request.args.get("line")

        # ✅ Base query
        sql = "SELECT * FROM four_m_change WHERE 1=1"
        params = []

        # ✅ Apply filters dynamically
        if month:
            # Convert month to start and end dates
            start_date = f"{month}-01"
            # Calculate last day of month
            end_date = datetime.strptime(start_date, "%Y-%m-%d").replace(day=28)  # temporary
            end_date = end_date.replace(day=28)  # fallback if month has <31 days
            sql += " AND DATE_FORMAT(date, '%%Y-%%m') = %s"
            params.append(month)

        if area:
            sql += " AND plant = %s"
            params.append(area)

        if line:
            sql += " AND sub_area = %s"
            params.append(line)

        sql += " ORDER BY id DESC"

        cursor.execute(sql, params)
        data = cursor.fetchall()
        # ic(data)

        return jsonify(data), 200

    except Exception as e:
        print("Database Error:", e)
        return jsonify({"error": str(e)}), 500

    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals():
            connection.close()

@app.route("/api/gl_tab", methods=["GET"])
def get_gl_tab():
    try:
        connection = get_db_connection()
        cursor = connection.cursor()

        # ✅ Get filters from query params
        month = request.args.get("month")  # format: YYYY-MM
        area = request.args.get("area")
        line = request.args.get("line")
        date = request.args.get("date")

        # ✅ Base query
        sql = "SELECT * FROM gl_board_kpis WHERE 1=1"
        params = []

        # ✅ Apply filters dynamically
        if date:  # 👈 Exact date filter (when coming from chart click)
            sql += " AND DATE(date) = %s"
            params.append(date)

        elif month:  # 👈 Month filter
            sql += " AND DATE_FORMAT(date, '%%Y-%%m') = %s"
            params.append(month)

        if area:
            sql += " AND plant = %s"
            params.append(area)

        if line:
            sql += " AND line = %s"
            params.append(line)

        sql += " ORDER BY id DESC"

        cursor.execute(sql, params)
        data = cursor.fetchall()

        return jsonify(data), 200

    except Exception as e:
        print("Database Error:", e)
        return jsonify({"error": str(e)}), 500

    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals():
            connection.close()

@app.route("/api/gl_qa_points/<int:id>", methods=["PUT"])
def update_gl_qa_points_status(id):
    try:
        connection = get_db_connection()
        cursor = connection.cursor()

        data = request.get_json()
        status = data.get("status")

        sql = "UPDATE gl_qual SET status=%s WHERE id=%s"
        cursor.execute(sql, (status, id))
        connection.commit()

        return jsonify({"status": "success"}), 200

    except Exception as e:
        print("Update Error:", e)
        return jsonify({"status": "fail", "error": str(e)}), 500

    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals():
            connection.close()


@app.route("/api/points", methods=["POST"])
def submit_points():
    try:
        data = request.get_json()
 
        connection = get_db_connection()
        cursor = connection.cursor()
 
        sql = """
            INSERT INTO points (
                date, points_identified, area, target_date, responsibility, status
            ) VALUES (%s, %s, %s, %s, %s, %s)
        """
 
        values = (
            data.get("date"),
            data.get("pointsIdentified"),
            data.get("area"),
            data.get("targetDate"),
            data.get("responsibility"),
            data.get("status")
        )
 
        cursor.execute(sql, values)
        connection.commit()
 
        return jsonify({"message": "Point submitted successfully"}), 201
 
    except Error as e:
        return jsonify({"error": str(e)}), 500
 
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500
 
    finally:
        if 'cursor' in locals(): cursor.close()
        if 'connection' in locals(): connection.close()

# def get_shift_slots(shift):
#     if shift == "A":
#         return [
#             "06:00-07:00", "07:00-08:00", "08:00-09:00", "09:00-10:00",
#             "10:00-11:00", "11:00-12:00", "12:00-13:00", "13:00-14:00",
#             "14:00-14:30"
#         ]
#     elif shift == "B":
#         return [
#             "14:30-15:30", "15:30-16:30", "16:30-17:30", "17:30-18:30",
#             "18:30-19:30", "19:30-20:30", "20:30-21:30", "21:30-22:30",
#             "22:30-23:00"
#         ]
#     elif shift == "C":
#         return [
#             "23:00-00:00", "00:00-01:00", "01:00-02:00", "02:00-03:00",
#             "03:00-04:00", "04:00-05:00", "05:00-06:00"
#         ]
    

# def get_shift_slots(shift):
#     if shift == "A":
#         slots = []

#         # Full 8 hours: 06:00-14:00
#         start = datetime.strptime("06:00", "%H:%M")
#         for i in range(8):
#             slot_start = start + timedelta(hours=i)
#             slot_end = slot_start + timedelta(hours=1)
#             slots.append(f"{slot_start.strftime('%H:%M')}-{slot_end.strftime('%H:%M')}")

#         # Last half-hour 14:00-14:30
#         slots.append("14:00-14:30")

#         return slots

#     elif shift == "B":
#         slots = []

#         # Example B shift: 14:30-23:00
#         start = datetime.strptime("14:30", "%H:%M")
#         for i in range(8):  # adjust number of hours as per your shift
#             slot_start = start + timedelta(hours=i)
#             slot_end = slot_start + timedelta(hours=1)
#             slots.append(f"{slot_start.strftime('%H:%M')}-{slot_end.strftime('%H:%M')}")
#         slots.append("22:30-23:00")  # last half-hour
#         return slots

#     elif shift == "C":
#         slots = []

#         # Example C shift: 23:00-06:00 (overnight)
#         start = datetime.strptime("23:00", "%H:%M")
#         for i in range(7):  # 7 hours to 06:00
#             slot_start = start + timedelta(hours=i)
#             slot_end = slot_start + timedelta(hours=1)
#             # wrap around midnight
#             slots.append(f"{slot_start.strftime('%H:%M')}-{slot_end.strftime('%H:%M')}")
#         slots.append("06:00-06:00")  # last half-hour if needed
#         return slots

#     return []

    
# def get_hour_slot(ts: datetime, shift_slots):
#     """Find which slot this timestamp belongs to"""
#     for slot in shift_slots:
#         start_str, end_str = slot.split("-")
#         start = datetime.strptime(start_str, "%H:%M").time()
#         end = datetime.strptime(end_str, "%H:%M").time()

#         if start <= ts.time() < end:
#             return slot
#     return None

def get_shift_slots(shift_start_time, schedule_minutes):
    """
    Returns hourly slots starting from shift_start_time for schedule_minutes duration.
    Each slot is a dict with 'start' and 'end' datetime objects.
    """
    slots = []
    start_time = shift_start_time
    remaining_minutes = schedule_minutes

    while remaining_minutes > 0:
        slot_minutes = min(60, remaining_minutes)  # max 1 hour
        end_time = start_time + timedelta(minutes=slot_minutes)
        slots.append({'start': start_time, 'end': end_time})
        start_time = end_time
        remaining_minutes -= slot_minutes

    return slots

def get_slot_for_timestamp(ts, slots):
    """
    Given a timestamp, find the slot it belongs to.
    """
    for slot in slots:
        if slot['start'] <= ts < slot['end']:
            return slot
    return None

@app.route("/api/pab", methods=["GET"])
def get_pab():
    date_str = request.args.get("date")
    shift = request.args.get("shift", "A")

    if not date_str:
        return jsonify({"error": "date parameter required"}), 400

    conn = get_db_connection()
    cursor = conn.cursor()

    # Fetch the active plan
    cursor.execute("""
        SELECT id, partna, plan_qty, active_time, model, schedule_time
        FROM planning
        WHERE status='Active'
        LIMIT 1
    """)
    plan = cursor.fetchone()
    if not plan:
        return jsonify([])

    planned_qty = plan["plan_qty"] or 0
    schedule_minutes = plan["schedule_time"] or 1  # total planned minutes
    target_per_min = planned_qty / schedule_minutes

    # active_time is the actual part start datetime
    active_time = plan["active_time"]  # assuming this is a datetime object from DB

    # Create slots starting from active_time
    slots = get_shift_slots(active_time, schedule_minutes)
    
    # Initialize slot_map
    slot_map = {}
    for slot in slots:
        slot_label = f"{slot['start'].strftime('%H:%M')} - {slot['end'].strftime('%H:%M')}"
        slot_minutes = int((slot['end'] - slot['start']).total_seconds() / 60)
        slot_target = round(slot_minutes * target_per_min)
        slot_map[slot_label] = {
            "id": slot_label,
            "partna": plan["partna"],
            "model": plan["model"],
            "hour_slot": slot_label,
            "target": slot_target,
            "actual": 0,
            "efficiency": 0,
            "gap": 0
        }

    # Fetch actual running counts
    cursor.execute("""
        SELECT timestamp, cutter1_count, cutter2_count
        FROM L8_Running_Status
        WHERE plan_id=%s
        AND DATE(timestamp)=%s
        ORDER BY timestamp
    """, (plan["id"], date_str))
    db_data = cursor.fetchall()
    cursor.close()
    conn.close()

    # Track first & last reading per slot
    slot_first, slot_last = {}, {}
    for row in db_data:
        ts = row["timestamp"]
        slot = get_slot_for_timestamp(ts, slots)
        if not slot:
            continue
        slot_label = f"{slot['start'].strftime('%H:%M')} - {slot['end'].strftime('%H:%M')}"
        val = (row["cutter1_count"] or 0) + (row["cutter2_count"] or 0)
        if slot_label not in slot_first:
            slot_first[slot_label] = val
        slot_last[slot_label] = val

    # Calculate actual, gap, efficiency
    for slot_label, s in slot_map.items():
        if slot_label in slot_first and slot_label in slot_last:
            actual = slot_last[slot_label] - slot_first[slot_label]
            s["actual"] = actual
        s["gap"] = s["target"] - s["actual"]
        s["efficiency"] = round((s["actual"] / s["target"]) * 100, 2) if s["target"] else 0

    return jsonify(list(slot_map.values()))

mode ="dev"

if __name__ == '__main__':
    app.secret_key = 'AnkitHarshAmit@123'
    start_scheduler()
    start_scheduler2()
    if mode == "dev":
        app.run(host='0.0.0.0', port=5000, debug=False)
    else:
        serve(app, host='0.0.0.0', port=5000, threads = 4 )